"""
Dialogs Components - Diálogos y ventanas emergentes

Responsabilidades:
- CustomerDetailDialog: Detalles de cliente
- ProductDetailDialog: Detalles de producto
- DayPurchasesDialog: Compras de un día específico
"""

import os
from datetime import datetime
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QMessageBox, QListWidget, QListWidgetItem
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QFont

THEME = {
    "bg_app": "#F6F7F9",
    "card_bg": "#FFFFFF",
    "primary": "#1F2937",
    "text_main": "#111827",
    "accent": "#111827",
    "border": "#E5E7EB",
}


def _load_sales_data(username, allow_remote_restore=True):
    try:
        from utils.file_handler import cargar_ventas_dashboard

        ventas = cargar_ventas_dashboard(
            username,
            allow_remote_restore=allow_remote_restore,
        ) or []
    except Exception:
        ventas = []

    if isinstance(ventas, dict):
        ventas = list(ventas.values())
    return ventas if isinstance(ventas, list) else []


class CustomerDetailDialog(QDialog):
    """Detalles de cliente"""
    
    def __init__(self, cliente_nombre, username=None, sales_data=None, allow_remote_restore=True, parent=None):
        super().__init__(parent)
        self.cliente_nombre = cliente_nombre
        self.username = username
        self.sales_data = list(sales_data) if isinstance(sales_data, list) else None
        self.allow_remote_restore = bool(allow_remote_restore)
        self.setWindowTitle(f"Detalles - {cliente_nombre}")
        self.setGeometry(100, 100, 800, 550)
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {THEME['bg_app']};
            }}
            QLabel {{
                color: {THEME['text_main']};
            }}
        """)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Título
        titulo = QLabel(f"Compras de {self.cliente_nombre}")
        titulo.setFont(QFont("Segoe UI", 14, QFont.Bold))
        titulo.setStyleSheet(f"color: {THEME['accent']};")
        layout.addWidget(titulo)
        
        # Tabla
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Fecha", "Producto", "Cantidad", "Monto (S/."])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {THEME['card_bg']};
                border: 1px solid {THEME['border']};
            }}
            QHeaderView::section {{
                background-color: {THEME['primary']};
                color: white;
                padding: 8px;
                font-weight: bold;
            }}
        """)
        layout.addWidget(self.table)
        
        self.load_customer_data()
        
        # Botones
        btn_layout = QHBoxLayout()
        btn_excel = QPushButton("Exportar a Excel")
        btn_excel.setStyleSheet(f"""
            QPushButton {{
                background-color: {THEME['primary']};
                color: white;
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: bold;
                border: none;
            }}
        """)
        btn_excel.clicked.connect(self.export_excel)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_excel)
        layout.addLayout(btn_layout)
    
    def load_customer_data(self):
        try:
            ventas = self.sales_data if self.sales_data is not None else _load_sales_data(
                self.username,
                allow_remote_restore=self.allow_remote_restore,
            )
            compras = []
            total_general = 0
            
            for venta in ventas:
                try:
                    if venta.get('cliente', '') == self.cliente_nombre:
                        fecha = venta.get('fecha', '')
                        producto = venta.get('producto', '')
                        cant = venta.get('cantidad', 0)
                        monto = float(venta.get('total', 0) or 0)
                        compras.append((fecha, producto, cant, monto))
                        total_general += monto
                except:
                    continue
            
            self.table.setRowCount(len(compras))
            for row, (fecha, producto, cant, monto) in enumerate(compras):
                self.table.setItem(row, 0, QTableWidgetItem(fecha))
                self.table.setItem(row, 1, QTableWidgetItem(producto))
                self.table.setItem(row, 2, QTableWidgetItem(str(cant)))
                self.table.setItem(row, 3, QTableWidgetItem(f"S/. {monto:,.2f}"))
            
            # Total
            self.table.insertRow(len(compras))
            total_label = QTableWidgetItem("TOTAL")
            total_label.setFont(QFont("Segoe UI", 10, QFont.Bold))
            total_label.setBackground(QColor(THEME['primary']))
            total_label.setForeground(QColor("white"))
            self.table.setItem(len(compras), 2, total_label)
            
            total_item = QTableWidgetItem(f"S/. {total_general:,.2f}")
            total_item.setFont(QFont("Segoe UI", 10, QFont.Bold))
            total_item.setBackground(QColor(THEME['primary']))
            total_item.setForeground(QColor("white"))
            self.table.setItem(len(compras), 3, total_item)
            
        except Exception as e:
            print(f"[CustomerDetailDialog] Error: {e}")
    
    def export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Detalles Cliente"
            
            ws['A1'] = f"Cliente: {self.cliente_nombre}"
            ws['A1'].font = Font(bold=True, size=12)
            
            headers = ["Fecha", "Producto", "Cantidad", "Monto (S/."]
            for col, header in enumerate(headers, 1):
                ws.cell(row=3, column=col).value = header
                ws.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=3, column=col).fill = PatternFill(start_color=THEME['primary'].lstrip('#'), 
                                                               end_color=THEME['primary'].lstrip('#'), fill_type="solid")
            
            row = 4
            for i in range(self.table.rowCount() - 1):
                for col in range(self.table.columnCount()):
                    item = self.table.item(i, col)
                    if item:
                        ws.cell(row=row, column=col+1).value = item.text()
                row += 1
            
            ws['D{}'.format(row)] = self.table.item(self.table.rowCount() - 1, 3).text()
            ws['D{}'.format(row)].font = Font(bold=True, size=11)
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            
            filename = f"{self.username}_cliente_{self.cliente_nombre.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(os.path.expanduser('~/Desktop'), filename)
            wb.save(filepath)
            
            QMessageBox.information(self, "Éxito", f"Archivo exportado a:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar: {e}")


class ProductDetailDialog(QDialog):
    """Detalles de producto"""
    
    def __init__(self, producto_nombre, username=None, sales_data=None, allow_remote_restore=True, parent=None):
        super().__init__(parent)
        self.producto_nombre = producto_nombre
        self.username = username
        self.sales_data = list(sales_data) if isinstance(sales_data, list) else None
        self.allow_remote_restore = bool(allow_remote_restore)
        self.setWindowTitle(f"Detalles - {producto_nombre}")
        self.setGeometry(100, 100, 800, 550)
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {THEME['bg_app']};
            }}
            QLabel {{
                color: {THEME['text_main']};
            }}
        """)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        titulo = QLabel(f"Ventas de {self.producto_nombre}")
        titulo.setFont(QFont("Segoe UI", 14, QFont.Bold))
        titulo.setStyleSheet(f"color: {THEME['accent']};")
        layout.addWidget(titulo)
        
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Fecha", "Cliente", "Cantidad", "Monto (S/."])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {THEME['card_bg']};
                border: 1px solid {THEME['border']};
            }}
            QHeaderView::section {{
                background-color: {THEME['primary']};
                color: white;
                padding: 8px;
                font-weight: bold;
            }}
        """)
        layout.addWidget(self.table)
        
        self.load_product_data()
        
        btn_layout = QHBoxLayout()
        btn_excel = QPushButton("Exportar a Excel")
        btn_excel.setStyleSheet(f"""
            QPushButton {{
                background-color: {THEME['primary']};
                color: white;
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: bold;
                border: none;
            }}
        """)
        btn_excel.clicked.connect(self.export_excel)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_excel)
        layout.addLayout(btn_layout)
    
    def load_product_data(self):
        try:
            ventas = self.sales_data if self.sales_data is not None else _load_sales_data(
                self.username,
                allow_remote_restore=self.allow_remote_restore,
            )
            ventas_producto = []
            total_general = 0
            
            for venta in ventas:
                try:
                    if venta.get('producto', '') == self.producto_nombre:
                        fecha = venta.get('fecha', '')
                        cliente = venta.get('cliente', '')
                        cant = venta.get('cantidad', 0)
                        monto = float(venta.get('total', 0) or 0)
                        ventas_producto.append((fecha, cliente, cant, monto))
                        total_general += monto
                except:
                    continue
            
            self.table.setRowCount(len(ventas_producto))
            for row, (fecha, cliente, cant, monto) in enumerate(ventas_producto):
                self.table.setItem(row, 0, QTableWidgetItem(fecha))
                self.table.setItem(row, 1, QTableWidgetItem(cliente))
                self.table.setItem(row, 2, QTableWidgetItem(str(cant)))
                self.table.setItem(row, 3, QTableWidgetItem(f"S/. {monto:,.2f}"))
            
            self.table.insertRow(len(ventas_producto))
            total_label = QTableWidgetItem("TOTAL")
            total_label.setFont(QFont("Segoe UI", 10, QFont.Bold))
            total_label.setBackground(QColor(THEME['primary']))
            total_label.setForeground(QColor("white"))
            self.table.setItem(len(ventas_producto), 2, total_label)
            
            total_item = QTableWidgetItem(f"S/. {total_general:,.2f}")
            total_item.setFont(QFont("Segoe UI", 10, QFont.Bold))
            total_item.setBackground(QColor(THEME['primary']))
            total_item.setForeground(QColor("white"))
            self.table.setItem(len(ventas_producto), 3, total_item)
            
        except Exception as e:
            print(f"[ProductDetailDialog] Error: {e}")
    
    def export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Detalles Producto"
            
            ws['A1'] = f"Producto: {self.producto_nombre}"
            ws['A1'].font = Font(bold=True, size=12)
            
            headers = ["Fecha", "Cliente", "Cantidad", "Monto (S/."]
            for col, header in enumerate(headers, 1):
                ws.cell(row=3, column=col).value = header
                ws.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=3, column=col).fill = PatternFill(start_color=THEME['primary'].lstrip('#'),
                                                               end_color=THEME['primary'].lstrip('#'), fill_type="solid")
            
            row = 4
            for i in range(self.table.rowCount() - 1):
                for col in range(self.table.columnCount()):
                    item = self.table.item(i, col)
                    if item:
                        ws.cell(row=row, column=col+1).value = item.text()
                row += 1
            
            ws['D{}'.format(row)] = self.table.item(self.table.rowCount() - 1, 3).text()
            ws['D{}'.format(row)].font = Font(bold=True, size=11)
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            
            filename = f"{self.username}_producto_{self.producto_nombre.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(os.path.expanduser('~/Desktop'), filename)
            wb.save(filepath)
            
            QMessageBox.information(self, "Éxito", f"Archivo exportado a:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar: {e}")


class DayPurchasesDialog(QDialog):
    """Compras de un día específico"""
    
    def __init__(self, day, purchases, sales_data=None, username=None, parent=None):
        super().__init__(parent)
        self.day = day
        self.purchases = purchases
        self.sales_data = sales_data or []
        self.username = username
        
        self.setWindowTitle(f"Compras - Día {day}")
        self.setFixedSize(500, 450)
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {THEME['bg_app']};
            }}
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        title = QLabel(f"Compras realizadas - Día {day}")
        title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title.setStyleSheet(f"color: {THEME['accent']};")
        layout.addWidget(title)
        
        info = QLabel(f"Total de clientes: {len(self.purchases)}")
        info.setFont(QFont("Segoe UI", 11))
        info.setStyleSheet(f"color: {THEME['text_main']};")
        layout.addWidget(info)
        
        list_widget = QListWidget()
        list_widget.setFont(QFont("Segoe UI", 10))
        
        if self.purchases:
            for i, customer in enumerate(self.purchases, 1):
                list_widget.addItem(QListWidgetItem(f"{i}. {customer}"))
        else:
            empty_item = QListWidgetItem("Sin compras este día")
            list_widget.addItem(empty_item)
        
        layout.addWidget(list_widget)
        
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        btn_export = QPushButton("Exportar a Excel")
        btn_export.setFixedHeight(35)
        btn_export.setFont(QFont("Segoe UI", 11))
        btn_export.setStyleSheet(f"""
            QPushButton {{
                background-color: #10B981;
                color: white;
                border: none;
                font-weight: bold
            }}
        """)
        btn_export.clicked.connect(self.export_to_excel)
        btn_layout.addWidget(btn_export)
        
        btn_close = QPushButton("Cerrar")
        btn_close.setFixedHeight(35)
        btn_close.setFont(QFont("Segoe UI", 11))
        btn_close.setStyleSheet(f"""
            QPushButton {{
                background-color: {THEME['primary']};
                color: white;
                border: none;
                font-weight: bold
            }}
        """)
        btn_close.clicked.connect(self.accept)
        btn_layout.addWidget(btn_close)
        
        layout.addLayout(btn_layout)
    
    def export_to_excel(self):
        """Exporta compras del día a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"Día {self.day}"
            
            headers = ["Nro", "Cliente", "Productos", "Total (S/.)"]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num).value = header
                ws.cell(row=1, column=col_num).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=1, column=col_num).fill = PatternFill(start_color=THEME['primary'].lstrip('#'),
                                                                   end_color=THEME['primary'].lstrip('#'), fill_type="solid")
            
            row_num = 2
            total_general = 0
            
            from datetime import datetime as dt
            for venta in self.sales_data:
                try:
                    fecha_str = venta.get('fecha', '').split()[0]
                    fecha = dt.strptime(fecha_str, "%d/%m/%Y")
                    if fecha.day == self.day:
                        nro = row_num - 1
                        cliente = venta.get('cliente', '')
                        producto = venta.get('producto', '')
                        total = float(venta.get('total', 0) or 0)
                        
                        ws.cell(row=row_num, column=1).value = nro
                        ws.cell(row=row_num, column=2).value = cliente
                        ws.cell(row=row_num, column=3).value = producto
                        ws.cell(row=row_num, column=4).value = round(total, 2)
                        
                        total_general += total
                        row_num += 1
                except:
                    continue
            
            total_row = row_num + 1
            ws.cell(row=total_row, column=3).value = "TOTAL:"
            ws.cell(row=total_row, column=3).font = Font(bold=True)
            ws.cell(row=total_row, column=3).alignment = Alignment(horizontal="right")
            
            ws.cell(row=total_row, column=4).value = round(total_general, 2)
            ws.cell(row=total_row, column=4).font = Font(bold=True)
            ws.cell(row=total_row, column=4).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 15
            
            desktop_path = os.path.expanduser("~/Desktop")
            filename = f"Ventas_Día_{self.day}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(desktop_path, filename)
            
            wb.save(filepath)
            QMessageBox.information(self, "Éxito", f"Archivo exportado:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}")
