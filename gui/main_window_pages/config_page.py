import sys
import os
import socket
import shutil
import json
import zipfile
import webbrowser
import datetime
import uuid
import csv
try:
    import pandas as pd # type: ignore
except ImportError:
    pd = None
try:
    import openpyxl # type: ignore
except ImportError:
    openpyxl = None
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QGroupBox, QLineEdit, QPushButton,
    QListWidget, QHBoxLayout, QMessageBox, QFormLayout, QApplication,
    QFileDialog, QTabWidget, QAbstractItemView, QFrame, QProgressDialog,
    QScrollArea, QGridLayout, QCheckBox
)
from PyQt5.QtCore import Qt, QTimer

# Importar páginas de gestión
# Materiales/Tallas/Tipos de Lente han sido movidos a Inventario > Configuración
from gui.main_window_pages.helpers_page import HelpersPage

# Importaciones para el entorno de desarrollo y empaquetado
from utils.file_handler import (
    cargar_nombre_optica, guardar_nombre_optica, cargar_optometras,
    guardar_optometras, agregar_optometra as agregar_optometra_json, eliminar_optometra as eliminar_optometra_json, cargar_metodos_pago, guardar_metodos_pago,
    comprimir_y_subir_datos, descargar_y_descomprimir_datos,
    cargar_password_setup, guardar_password_setup,
    get_user_file_path, VISO_DIR, crear_tabla_usuarios_remoto, resource_path,
    cargar_logo_optica, guardar_logo_optica, cargar_tamano_logo, guardar_tamano_logo,
    cargar_datos_generales, guardar_datos_generales, actualizar_datos_sunat,
    cargar_ruc, guardar_ruc, cargar_razon_social, guardar_razon_social,
    cargar_configuracion_optica, guardar_configuracion_optica,
    cargar_token_sunat, guardar_token_sunat,
    cargar_datos_optica, guardar_datos_optica, cargar_whatsapp_optica,
    is_modo_basico, set_modo_basico, get_modo_basico_config,
    save_modo_basico_config, MODO_BASICO_PAGE_OPTIONS, MODO_BASICO_HOME_ACTION_OPTIONS,
    get_plantillas_ventas_disponibles, cargar_plantilla_ventas_seleccionada,
    guardar_plantilla_ventas_seleccionada
)
from utils.lan_service import LanServerWorker, LanClient, LanAutoSyncWorker
from utils.text_normalizer import maybe_normalize_ui_text
import threading

LAN_DISABLED_MESSAGE = (
    "El servicio LAN fue deshabilitado completamente por mantenimiento. "
    "Usa solo sincronizacion normal/local hasta nuevo aviso."
)


class SaveChildDeviceWorker(QtCore.QThread):
    """Guarda/actualiza dispositivo hijo en hilo separado para no bloquear la UI."""
    result_ready = QtCore.pyqtSignal(dict)

    def __init__(
        self,
        username: str,
        selected_child_id: str,
        nombre_optica: str,
        ciudad: str,
        codigo: str,
        estado: str,
        parent=None
    ):
        super().__init__(parent)
        self.username = str(username or "").strip()
        self.selected_child_id = str(selected_child_id or "").strip()
        self.nombre_optica = str(nombre_optica or "").strip()
        self.ciudad = str(ciudad or "").strip()
        self.codigo = str(codigo or "").strip().upper()
        self.estado = "activo" if str(estado or "").lower() == "activo" else "bloqueado"

    def _normalize_limit(self, value, default=0):
        try:
            parsed = int(value)
            if parsed > 0:
                return parsed
        except Exception:
            pass
        return max(0, int(default or 0))

    def run(self):
        result = {
            "ok": False,
            "kind": "unknown_error",
            "message": "Error desconocido guardando dispositivo.",
            "max_sucursales": 0,
            "total_actual": 0
        }

        try:
            try:
                from utils.api_handler import listar_dispositivos_hijos_remoto_con_limite
                ok_remote, dispositivos, max_remote, msg_remote = listar_dispositivos_hijos_remoto_con_limite(self.username)
                max_sucursales = self._normalize_limit(max_remote, default=0)
            except Exception:
                from utils.api_handler import listar_dispositivos_hijos_remoto
                ok_remote, dispositivos, msg_remote = listar_dispositivos_hijos_remoto(self.username)
                max_sucursales = 0

            if not ok_remote:
                result.update({
                    "kind": "load_error",
                    "message": f"No se pudo consultar dispositivos en nube.\nDetalle: {msg_remote}"
                })
                self.result_ready.emit(result)
                return

            dispositivos = dispositivos if isinstance(dispositivos, list) else []
            dispositivos = [d for d in dispositivos if isinstance(d, dict)]
            total_actual = len(dispositivos)
            activos_actual = len([
                d for d in dispositivos
                if str(d.get("estado", "activo")).strip().lower() != "bloqueado"
            ])
            bloqueados_actual = max(0, total_actual - activos_actual)
            result["max_sucursales"] = max_sucursales
            # Usar activas como "usadas" para el limite (bloqueadas no consumen cupo).
            result["total_actual"] = activos_actual

            if not self.selected_child_id and max_sucursales == 1:
                result.update({
                    "kind": "single_branch_plan",
                    "message": (
                        "Tu cuenta esta configurada para una sola sucursal.\n\n"
                        "Este equipo ya es tu sucursal principal, por lo que no puedes crear dispositivos hijos."
                    )
                })
                self.result_ready.emit(result)
                return

            if not self.selected_child_id and max_sucursales > 0 and activos_actual >= max_sucursales:
                result.update({
                    "kind": "limit_reached",
                    "message": (
                        f"Ya alcanzaste el limite de sucursales para tu cuenta.\n\n"
                        f"Activas: {activos_actual}/{max_sucursales}\n"
                        f"Bloqueadas: {bloqueados_actual} | Total: {total_actual}\n"
                        f"Para agregar mas sucursales, aumenta max_sucursales en la BD."
                    )
                })
                self.result_ready.emit(result)
                return

            for d in dispositivos:
                if str(d.get("codigo_dispositivo", "")).upper() == self.codigo:
                    if self.selected_child_id and str(d.get("id", "")) == self.selected_child_id:
                        continue
                    result.update({
                        "kind": "duplicate",
                        "message": "Ya existe un dispositivo con ese codigo."
                    })
                    self.result_ready.emit(result)
                    return

            now_iso = datetime.datetime.now().isoformat()
            dispositivo_payload = None
            action_msg = ""

            if self.selected_child_id:
                updated = False
                for d in dispositivos:
                    if str(d.get("id", "")) == self.selected_child_id:
                        d["nombre_optica"] = self.nombre_optica
                        d["ciudad"] = self.ciudad
                        d["codigo_dispositivo"] = self.codigo
                        d["estado"] = self.estado
                        d["updated_at"] = now_iso
                        dispositivo_payload = dict(d)
                        updated = True
                        break
                if not updated:
                    result.update({
                        "kind": "not_found",
                        "message": "No se encontro el dispositivo seleccionado."
                    })
                    self.result_ready.emit(result)
                    return
                action_msg = "Dispositivo hijo actualizado correctamente."
            else:
                nuevo_dispositivo = {
                    "id": str(uuid.uuid4()),
                    "nombre_optica": self.nombre_optica,
                    "ciudad": self.ciudad,
                    "codigo_dispositivo": self.codigo,
                    "estado": self.estado,
                    "cloud_sync_enabled": True,
                    "ultima_sincronizacion": None,
                    "created_at": now_iso,
                    "updated_at": now_iso
                }
                dispositivo_payload = dict(nuevo_dispositivo)
                action_msg = "Dispositivo hijo agregado correctamente."

            from utils.api_handler import sync_dispositivo_hijo_remoto
            remote_ok, remote_msg, _remote_device = sync_dispositivo_hijo_remoto(
                self.username,
                dispositivo_payload or {}
            )

            if remote_ok:
                result.update({
                    "ok": True,
                    "kind": "ok",
                    "message": f"{action_msg}\n\nSincronizado en nube correctamente."
                })
            else:
                result.update({
                    "kind": "save_error",
                    "message": f"No se pudo guardar el dispositivo en la base de datos remota.\nDetalle: {remote_msg}"
                })

        except Exception as e:
            result.update({
                "kind": "exception",
                "message": f"No se pudo guardar el dispositivo hijo.\n{e}"
            })

        self.result_ready.emit(result)


class ConfigPage(QWidget):
    def exportar_datos_generales_ui(self):
        """Crea una copia de seguridad y abre el administrador de archivos en la carpeta."""
        try:
            # Crear un nombre de archivo con fecha y hora
            fecha_hora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"VISO_backup_{fecha_hora}.zip"
            
            # Mostrar diálogo para seleccionar carpeta
            carpeta_destino = QFileDialog.getExistingDirectory(
                self,
                "Seleccionar carpeta para guardar la copia de seguridad",
                os.path.expanduser("~\\Documents"),
                QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks
            )
            
            if not carpeta_destino:
                return
            
            ruta_guardado = os.path.join(carpeta_destino, nombre_archivo)

            # Mostrar diálogo de progreso
            progress = QProgressDialog("Preparando archivos...", "Cancelar", 0, 100, self)
            progress.setWindowTitle("Exportando Datos")
            progress.setWindowModality(Qt.WindowModal)
            progress.setAutoClose(True)
            progress.setMinimumDuration(0)
            progress.show()
            QApplication.processEvents()

            # Crear ZIP temporal
            progress.setValue(10)
            progress.setLabelText("Creando directorio temporal...")
            temp_dir = os.path.join(VISO_DIR, "temp_backup")
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            os.makedirs(temp_dir)

            try:
                # Copiar archivos al directorio temporal
                progress.setValue(20)
                progress.setLabelText("Copiando archivos...")
                user_data_dir = os.path.join(VISO_DIR, str(self.username))
                if os.path.exists(user_data_dir):
                    files_to_copy = []
                    for root, dirs, files in os.walk(user_data_dir):
                        for file in files:
                            if file.lower() != "clave_activacion.txt":
                                files_to_copy.append((root, file))
                    
                    total_files = len(files_to_copy)
                    for i, (root, file) in enumerate(files_to_copy):
                        if progress.wasCanceled():
                            raise Exception('Operación cancelada por el usuario')
                        
                        src_path = os.path.join(root, file)
                        rel_path = os.path.relpath(src_path, user_data_dir)
                        dst_path = os.path.join(temp_dir, rel_path)
                        os.makedirs(os.path.dirname(dst_path), exist_ok=True)
                        shutil.copy2(src_path, dst_path)
                        
                        # Actualizar progreso (20-50%)
                        current_progress = 20 + int((i / total_files) * 30)
                        progress.setValue(current_progress)
                        progress.setLabelText(f"Copiando archivos... ({i+1}/{total_files})")

                # Crear archivo ZIP
                progress.setValue(60)
                progress.setLabelText("Comprimiendo archivos...")
                
                # Obtener lista de archivos a comprimir
                files_to_zip = []
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, temp_dir)
                        files_to_zip.append((file_path, arcname))
                
                total_files = len(files_to_zip)
                with zipfile.ZipFile(ruta_guardado, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for i, (file_path, arcname) in enumerate(files_to_zip):
                        if progress.wasCanceled():
                            raise Exception('Operación cancelada por el usuario')
                            
                        zipf.write(file_path, arcname)
                        
                        # Actualizar progreso (60-90%)
                        current_progress = 60 + int((i / total_files) * 30)
                        progress.setValue(current_progress)
                        progress.setLabelText(f"Comprimiendo archivos... ({i+1}/{total_files})")
                
                progress.setValue(100)
                progress.setLabelText("Completado!")
                progress.close()
                
                # Mostrar mensaje de Éxito
                resultado = QMessageBox.information(
                    self,
                    'Éxito',
                    f"Copia de seguridad creada exitosamente en:\n{ruta_guardado}\n\n"
                    'Nota: La clave de activación no se incluye en la copia de seguridad\n'
                    "por razones de seguridad.\n\n"
                    '¿Deseas abrir la carpeta?',
                    QMessageBox.Yes | QMessageBox.No
                )
                
                # Abrir el administrador de archivos en la carpeta si el usuario lo desea
                if resultado == QMessageBox.Yes:
                    import subprocess
                    subprocess.Popen(f'explorer /select,"{ruta_guardado}"')

            except Exception as e:
                progress.close()
                QMessageBox.critical(
                    self,
                    "Error",
                    f"Error al crear la copia de seguridad:\n{str(e)}"
                )
            finally:
                # Limpiar archivos temporales
                if os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir)

        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error inesperado:\n{str(e)}"
            )

    def guardar_numero_whatsapp(self):
        'Guarda el número de WhatsApp en whatsapp.json y muestra confirmación.'
        numero = self.entry_whatsapp.text().strip()
        try:
            payload = {
                "nombre_optica": self.entry_nombre_optica.text().strip() if hasattr(self, "entry_nombre_optica") else "",
                "slogan": self.entry_slogan_optica.text().strip() if hasattr(self, "entry_slogan_optica") else "",
                "direccion": self.entry_direccion_optica.text().strip() if hasattr(self, "entry_direccion_optica") else "",
                "correo_electronico": self.entry_correo_optica.text().strip() if hasattr(self, "entry_correo_optica") else "",
                "whatsapp": numero,
            }
            guardar_datos_optica(self.username, payload, sync_remote=False)
            remote_ok, remote_msg = self._guardar_datos_optica_en_bd(payload)
            if remote_ok:
                QMessageBox.information(self, "WhatsApp guardado", "El numero de WhatsApp se ha guardado correctamente en local y BD.")
            else:
                QMessageBox.warning(
                    self,
                    "WhatsApp guardado localmente",
                    f"El número se guardó localmente, pero no se pudo guardar en la BD remota.\n\nDetalle: {remote_msg}"
                )
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo guardar el número de WhatsApp.\n{e}")

    def _guardar_datos_optica_en_bd(self, payload):
        """Guarda datos comerciales en la BD remota y devuelve resultado explícito."""
        try:
            from utils.api_handler import guardar_datos_optica_remoto
            from utils.file_handler import cargar_usuarios

            usuario_id = str(self.username or "")
            usuarios = cargar_usuarios() or {}
            for uid, info in usuarios.items():
                if isinstance(info, dict) and info.get('username') == self.username:
                    usuario_id = str(uid)
                    break

            ok, msg, _data = guardar_datos_optica_remoto(
                username=str(self.username or ""),
                usuario_id=usuario_id,
                datos=dict(payload or {}),
            )
            return bool(ok), str(msg or "")
        except Exception as e:
            return False, str(e)

    def _get_device_config_path(self):
        'Obtiene la ruta del archivo de configuración del tipo de dispositivo.'
        return os.path.join(VISO_DIR, self.username, "data", "config_dispositivo.json")

    def guardar_tipo_dispositivo_ui(self):
        'Guarda la selección de tipo de dispositivo y actualiza la UI local.'
        try:
            if not self.username:
                QMessageBox.warning(self, "Error", "No se pudo identificar el usuario actual.")
                return

            tipo_dispositivo = "madre" if self.device_type_combo.currentIndex() == 0 else "trabajador"
            tipo_label = self.device_type_combo.currentText()

            payload = {
                "tipo_dispositivo": tipo_dispositivo,
                "tipo_dispositivo_label": tipo_label,
                "updated_at": datetime.datetime.now().isoformat()
            }

            config_path = self._get_device_config_path()
            os.makedirs(os.path.dirname(config_path), exist_ok=True)
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)

            self.actualizar_modo_dispositivo_ui()

            if hasattr(self, "parent_app") and hasattr(self.parent_app, "on_device_role_changed"):
                try:
                    self.parent_app.on_device_role_changed(tipo_dispositivo)
                except Exception:
                    pass

            QMessageBox.information(
                self,
                'Configuración Guardada',
                f"Se guardó correctamente: {tipo_label}.\n\n"
                'Para aplicar todo el menú del modo madre/trabajador, reinicia la aplicación.'
            )
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo guardar la configuración del dispositivo.\n{e}")

    def cargar_tipo_dispositivo_ui(self):
        """Carga el tipo de dispositivo guardado y actualiza el selector."""
        try:
            if not self.username or not hasattr(self, "device_type_combo"):
                return

            config_path = self._get_device_config_path()
            if not os.path.exists(config_path):
                return

            with open(config_path, "r", encoding="utf-8") as f:
                config = json.load(f)

            tipo_dispositivo = str(config.get("tipo_dispositivo", "madre")).strip().lower()
            if tipo_dispositivo == "trabajador":
                self.device_type_combo.setCurrentIndex(1)
            else:
                self.device_type_combo.setCurrentIndex(0)
        except Exception:
            # Si el archivo está corrupto o vacío, mantener selección por defecto.
            pass
        finally:
            self.actualizar_modo_dispositivo_ui()

    def es_dispositivo_madre_ui(self):
        return hasattr(self, "device_type_combo") and self.device_type_combo.currentIndex() == 0

    def actualizar_modo_dispositivo_ui(self):
        'Habilita/deshabilita funciones de dispositivo madre en la página de configuración.'
        es_madre = self.es_dispositivo_madre_ui()

        if hasattr(self, "device_mode_note"):
            if es_madre:
                self.device_mode_note.setText(
                    "Modo activo: Dispositivo madre. Se habilitan opciones administrativas."
                )
            else:
                self.device_mode_note.setText(
                    "Modo activo: Dispositivo trabajador. Las opciones administrativas se limitan."
                )

        if hasattr(self, "child_devices_access_note"):
            self.child_devices_access_note.setVisible(not es_madre)

        if hasattr(self, "child_devices_tab_index") and hasattr(self, "tab_widget"):
            self.tab_widget.setTabEnabled(self.child_devices_tab_index, es_madre)
            if not es_madre and self.tab_widget.currentIndex() == self.child_devices_tab_index:
                self.tab_widget.setCurrentIndex(0)

        widget_names = [
            "entry_dispositivo_nombre",
            "entry_dispositivo_ciudad",
            "entry_dispositivo_codigo",
            "combo_dispositivo_estado",
            "btn_generar_codigo_dispositivo",
            "btn_guardar_dispositivo_hijo",
            "btn_limpiar_dispositivo_hijo",
            "list_dispositivos_hijos",
            "btn_eliminar_dispositivo_hijo",
            "btn_info_nube_dispositivos",
        ]
        for widget_name in widget_names:
            widget = getattr(self, widget_name, None)
            if widget is not None:
                widget.setEnabled(es_madre)

    def _get_child_devices_path(self):
        """Obtiene la ruta del archivo local de dispositivos hijos."""
        return os.path.join(VISO_DIR, self.username, "data", "dispositivos_hijos.json")

    def _normalizar_max_sucursales(self, value, default=0):
        """Normaliza max_sucursales a entero valido."""
        try:
            parsed = int(value)
            if parsed > 0:
                return parsed
        except Exception:
            pass
        return max(0, int(default or 0))

    def _leer_max_sucursales_local(self):
        """Intenta leer max_sucursales cacheado en .usuarios.json para este usuario."""
        try:
            from utils.file_handler import cargar_usuarios
            usuarios = cargar_usuarios()
            if not isinstance(usuarios, dict):
                return 0

            # 1) Buscar por username dentro de los entries
            for _uid, entry in usuarios.items():
                if not isinstance(entry, dict):
                    continue
                if str(entry.get("username", "")).strip() == str(self.username or "").strip():
                    if "max_sucursales" in entry:
                        return self._normalizar_max_sucursales(entry.get("max_sucursales"), default=0)

            # 2) Fallback por clave directa (si user_id == username)
            direct = usuarios.get(str(self.username or ""))
            if isinstance(direct, dict) and "max_sucursales" in direct:
                return self._normalizar_max_sucursales(direct.get("max_sucursales"), default=0)
        except Exception:
            pass
        return 0

    def _guardar_max_sucursales_local(self, max_sucursales):
        """Guarda max_sucursales en .usuarios.json para reutilizarlo en UI."""
        max_sucursales = self._normalizar_max_sucursales(max_sucursales, default=0)
        if max_sucursales <= 0:
            return
        try:
            from utils.file_handler import cargar_usuarios, guardar_usuarios
            usuarios = cargar_usuarios()
            if not isinstance(usuarios, dict):
                usuarios = {}

            target_key = None
            username = str(self.username or "").strip()

            for uid, entry in usuarios.items():
                if isinstance(entry, dict) and str(entry.get("username", "")).strip() == username:
                    target_key = uid
                    break

            if target_key is None:
                target_key = username
                if target_key not in usuarios or not isinstance(usuarios.get(target_key), dict):
                    usuarios[target_key] = {"username": username}

            if isinstance(usuarios.get(target_key), dict):
                usuarios[target_key]["max_sucursales"] = int(max_sucursales)
                guardar_usuarios(usuarios)
        except Exception:
            pass

    def _actualizar_etiqueta_limite_sucursales(self, registrados=0):
        """Actualiza texto de limite en la UI de dispositivos hijos."""
        limite = self._normalizar_max_sucursales(getattr(self, "_max_sucursales", 0), default=0)
        try:
            usados = int(registrados or 0)
        except Exception:
            usados = 0

        if hasattr(self, "lbl_dispositivos_limite"):
            if limite > 0:
                disponibles = max(0, limite - usados)
                self.lbl_dispositivos_limite.setText(
                    f"Límite de sucursales: {usados}/{limite} (disponibles: {disponibles})"
                )
                self.lbl_dispositivos_limite.setStyleSheet(
                    "color: #2e7d32; font-size: 12px;" if disponibles > 0 else "color: #c62828; font-size: 12px; font-weight: bold;"
                )
            else:
                limite_estimado = max(1, usados)
                self.lbl_dispositivos_limite.setText(
                    f"Límite de sucursales: {usados}/{limite_estimado} (estimado)"
                )
                self.lbl_dispositivos_limite.setStyleSheet("color: #546e7a; font-size: 12px;")

    def _cargar_dispositivos_hijos_remoto(self, show_error: bool = False):
        """
        Carga dispositivos hijos desde la BD remota (API cloud).
        Retorna: (ok, dispositivos, mensaje)
        """
        try:
            max_sucursales = self._leer_max_sucursales_local()
            try:
                from utils.api_handler import listar_dispositivos_hijos_remoto_con_limite
                ok, remote_devices, max_remote, msg = listar_dispositivos_hijos_remoto_con_limite(self.username)
                max_sucursales = self._normalizar_max_sucursales(max_remote, default=max_sucursales)
            except Exception:
                from utils.api_handler import listar_dispositivos_hijos_remoto
                ok, remote_devices, msg = listar_dispositivos_hijos_remoto(self.username)

            if not ok:
                self._max_sucursales = self._normalizar_max_sucursales(max_sucursales, default=0)
                if show_error:
                    QMessageBox.warning(
                        self,
                        "No se pudo cargar dispositivos hijos",
                        f"Error consultando la base de datos remota.\nDetalle: {msg}"
                    )
                return False, [], str(msg or "Error de consulta remota")

            devices = remote_devices if isinstance(remote_devices, list) else []
            devices = [d for d in devices if isinstance(d, dict)]
            resolved_limit = self._normalizar_max_sucursales(max_sucursales, default=0)
            self._max_sucursales = resolved_limit
            self._guardar_max_sucursales_local(self._max_sucursales)
            self.guardar_dispositivos_hijos(devices)  # cache local auxiliar
            self._dispositivos_hijos_cache = devices
            activos = len([
                d for d in devices
                if str(d.get("estado", "activo")).strip().lower() != "bloqueado"
            ])
            self._actualizar_etiqueta_limite_sucursales(activos)
            return True, devices, "OK"
        except Exception as e:
            self._max_sucursales = self._normalizar_max_sucursales(
                self._leer_max_sucursales_local(),
                default=0
            )
            if show_error:
                QMessageBox.warning(
                    self,
                    "No se pudo cargar dispositivos hijos",
                    f"Error consultando la base de datos remota.\nDetalle: {e}"
                )
            return False, [], str(e)

    def cargar_dispositivos_hijos(self, show_error: bool = False):
        """
        Carga dispositivos hijos desde la BD remota.
        No usa JSON local como fuente de verdad.
        """
        ok, devices, _msg = self._cargar_dispositivos_hijos_remoto(show_error=show_error)
        if ok:
            return devices
        return []

    def guardar_dispositivos_hijos(self, dispositivos):
        """Guarda cache local de dispositivos hijos (auxiliar para lectura offline)."""
        try:
            path = self._get_child_devices_path()
            payload = dispositivos if isinstance(dispositivos, list) else []
            if os.path.exists(path):
                pass
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def toggle_edicion_codigo_hijo(self):
        'Alterna entre modo solo lectura y edición manual del código de vinculación.'
        is_readonly = self.entry_dispositivo_codigo.isReadOnly()
        new_state = not is_readonly
        
        self.entry_dispositivo_codigo.setReadOnly(new_state)
        
        if new_state:
            # Volver a bloquear
            self.entry_dispositivo_codigo.setStyleSheet("background-color: #f5f5f5; color: #555;")
            self.btn_cambiar_codigo_manual.setText("Cambiar")
        else:
            # Desbloquear para edición
            self.entry_dispositivo_codigo.setStyleSheet("background-color: #ffffff; color: #000;")
            self.btn_cambiar_codigo_manual.setText("Bloquear")
            self.entry_dispositivo_codigo.setFocus()

    def generar_codigo_dispositivo_hijo(self):
        """Genera codigo de vinculacion para un dispositivo hijo."""
        code = f"VISO-{datetime.datetime.now().strftime('%y%m%d')}-{str(uuid.uuid4())[:6].upper()}"
        self.entry_dispositivo_codigo.setText(code)

    def _animar_boton_agregar_dispositivo_hijo(self):
        """Animacion de pulso al presionar el boton de agregar/actualizar."""
        try:
            btn = getattr(self, "btn_guardar_dispositivo_hijo", None)
            if btn is None:
                return

            effect = btn.graphicsEffect()
            if not isinstance(effect, QtWidgets.QGraphicsOpacityEffect):
                effect = QtWidgets.QGraphicsOpacityEffect(btn)
                effect.setOpacity(1.0)
                btn.setGraphicsEffect(effect)

            current_anim = getattr(self, "_anim_add_child_btn", None)
            if current_anim is not None:
                try:
                    current_anim.stop()
                except Exception:
                    pass

            anim = QtCore.QPropertyAnimation(effect, b"opacity", self)
            anim.setDuration(220)
            anim.setStartValue(1.0)
            anim.setKeyValueAt(0.5, 0.45)
            anim.setEndValue(1.0)
            anim.setEasingCurve(QtCore.QEasingCurve.InOutQuad)

            self._anim_add_child_btn = anim
            anim.start()
        except Exception:
            pass

    def _set_child_device_form_enabled(self, enabled: bool):
        for widget_name in (
            "entry_dispositivo_nombre",
            "entry_dispositivo_ciudad",
            "entry_dispositivo_codigo",
            "combo_dispositivo_estado",
            "btn_generar_codigo_dispositivo",
            "btn_guardar_dispositivo_hijo",
            "btn_limpiar_dispositivo_hijo",
        ):
            widget = getattr(self, widget_name, None)
            if widget is not None:
                widget.setEnabled(bool(enabled))

    def _start_child_device_loader(self):
        """Activa loader animado (puntos circulares) para guardado en nube."""
        self._stop_child_device_loader()

        self._set_child_device_form_enabled(False)
        if hasattr(self, "lbl_guardar_dispositivo_loader"):
            self.lbl_guardar_dispositivo_loader.setVisible(True)
            self.lbl_guardar_dispositivo_loader.setText("Guardando sucursal...")

        self._child_save_loader_step = 0
        self._child_save_loader_timer = QtCore.QTimer(self)
        self._child_save_loader_timer.setInterval(160)
        self._child_save_loader_timer.timeout.connect(self._tick_child_device_loader)
        self._child_save_loader_timer.start()
        self._tick_child_device_loader()

    def _stop_child_device_loader(self):
        try:
            if self._child_save_loader_timer is not None:
                self._child_save_loader_timer.stop()
                self._child_save_loader_timer.deleteLater()
        except Exception:
            pass
        self._child_save_loader_timer = None
        self._child_save_loader_step = 0

        if hasattr(self, "lbl_guardar_dispositivo_loader"):
            self.lbl_guardar_dispositivo_loader.setVisible(False)

        # Respetar modo madre/trabajador al reactivar controles.
        self.actualizar_modo_dispositivo_ui()

    def _tick_child_device_loader(self):
        if not hasattr(self, "lbl_guardar_dispositivo_loader") or self.lbl_guardar_dispositivo_loader is None:
            return
        frames = ("|", "/", "-", "\\")
        idx = int(getattr(self, "_child_save_loader_step", 0)) % len(frames)
        self.lbl_guardar_dispositivo_loader.setText(f"Guardando sucursal  {frames[idx]}")
        self._child_save_loader_step = idx + 1

    def _on_guardar_dispositivo_hijo_async_result(self, result: dict):
        self._stop_child_device_loader()

        if not isinstance(result, dict):
            QMessageBox.warning(self, "Error", 'Respuesta inválida al guardar dispositivo hijo.')
            return

        max_remote = self._normalizar_max_sucursales(result.get("max_sucursales", 0), default=0)
        if max_remote > 0:
            self._max_sucursales = max_remote
            self._guardar_max_sucursales_local(max_remote)

        if result.get("ok"):
            # --- INICIALIZAR ESTRUCTURA EN NUBE ---
            try:
                import requests
                import urllib3
                urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
                
                # Obtener código que estaba en el form
                codigo_sucursal = self.entry_dispositivo_codigo.text().strip().upper()
                if codigo_sucursal:
                    init_url = "https://api.yhana.cloud/win/new/init_branch.php"
                    requests.post(init_url, json={
                        "usuario_madre": self.username,
                        "codigo_sucursal": codigo_sucursal
                    }, timeout=5, verify=False)
            except Exception as e_cloud:
                print(f"[CLOUD] No se pudo inicializar carpeta: {e_cloud}")

            self.update_dispositivos_hijos_list()
            if self.parent_app and hasattr(self.parent_app, "on_device_role_changed"):
                try:
                    self.parent_app.on_device_role_changed("madre")
                except Exception:
                    pass
            self.limpiar_form_dispositivo_hijo()
            QMessageBox.information(self, 'Éxito', str(result.get("message", "Guardado correctamente.")))
            return

        kind = str(result.get("kind", "error"))
        message = str(result.get("message", "No se pudo guardar el dispositivo hijo."))
        if kind == "limit_reached":
            self._actualizar_etiqueta_limite_sucursales(result.get("total_actual", 0))
            QMessageBox.warning(self, 'Límite alcanzado', message)
            return
        if kind == "single_branch_plan":
            QMessageBox.information(self, "Dispositivo unico", message)
            return
        if kind == "duplicate":
            QMessageBox.warning(self, 'Código duplicado', message)
            return

        QMessageBox.warning(self, "Error", message)

    def _on_guardar_dispositivo_hijo_worker_finished(self):
        self._save_child_worker = None

    def limpiar_form_dispositivo_hijo(self):
        """Limpia el formulario y resetea modo edicion."""
        self._selected_child_device_id = None
        self.entry_dispositivo_nombre.clear()
        self.entry_dispositivo_ciudad.clear()
        # Generar código automáticamente al limpiar/iniciar
        self.generar_codigo_dispositivo_hijo()
        self.combo_dispositivo_estado.setCurrentIndex(0)
        self.btn_guardar_dispositivo_hijo.setText("Agregar dispositivo hijo")

    def update_dispositivos_hijos_list(self):
        """Refresca la lista visual de dispositivos hijos."""
        if not hasattr(self, "list_dispositivos_hijos"):
            return

        dispositivos = self.cargar_dispositivos_hijos(show_error=False)
        self._dispositivos_hijos_cache = dispositivos
        self.list_dispositivos_hijos.clear()

        for dispositivo in dispositivos:
            estado = str(dispositivo.get("estado", "activo")).lower()
            estado_tag = "ACTIVO" if estado == "activo" else "BLOQUEADO"
            nombre = dispositivo.get("nombre_optica", "Sin nombre")
            ciudad = dispositivo.get("ciudad", "Sin ciudad")
            codigo = dispositivo.get("codigo_dispositivo", "Sin codigo")
            text = f"[{estado_tag}] {nombre} | {ciudad} | Código: {codigo}"

            item = QtWidgets.QListWidgetItem(text)
            item.setData(Qt.UserRole, dispositivo.get("id"))
            self.list_dispositivos_hijos.addItem(item)

        total = len(dispositivos)
        activos = len([
            d for d in dispositivos
            if str(d.get("estado", "activo")).strip().lower() != "bloqueado"
        ])
        bloqueados = max(0, total - activos)
        limite = self._normalizar_max_sucursales(getattr(self, "_max_sucursales", 0), default=0)
        if hasattr(self, "lbl_dispositivos_total"):
            if limite > 0:
                disponibles = max(0, limite - activos)
                self.lbl_dispositivos_total.setText(
                    f"Sucursales activas: {activos}/{limite} | Bloqueadas: {bloqueados} | Total: {total} | Disponibles: {disponibles}"
                )
            else:
                limite_estimado = max(1, total)
                self.lbl_dispositivos_total.setText(
                    f"Dispositivos registrados: {total}/{limite_estimado} | Límite estimado"
                )
        self._actualizar_etiqueta_limite_sucursales(activos)

    def _load_dispositivos_hijos_with_loader(self, force=False):
        """Carga dispositivos hijos mostrando el loader de la página."""
        if getattr(self, "_loading_dispositivos_hijos", False):
            return
        if not force and getattr(self, "_dispositivos_hijos_loaded_once", False):
            return

        self._loading_dispositivos_hijos = True
        self._show_config_page_loader(
            title="Cargando dispositivos hijos",
            subtitle="Consultando sucursales y estado en la nube. Esto puede tardar unos segundos."
        )
        self._yield_ui_for_loader()

        def _do_load():
            try:
                self.update_dispositivos_hijos_list()
                self._dispositivos_hijos_loaded_once = True
            finally:
                self._loading_dispositivos_hijos = False
                self._hide_config_page_loader()

        QTimer.singleShot(25, _do_load)

    def recargar_dispositivos_hijos_ui(self):
        """Recarga manualmente la lista de dispositivos hijos."""
        try:
            self._load_dispositivos_hijos_with_loader(force=True)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo recargar la lista.\n{e}")

    def seleccionar_dispositivo_hijo(self):
        """Carga en formulario el dispositivo seleccionado."""
        item = self.list_dispositivos_hijos.currentItem()
        if not item:
            return

        device_id = item.data(Qt.UserRole)
        dispositivos = getattr(self, "_dispositivos_hijos_cache", None)
        if not isinstance(dispositivos, list):
            dispositivos = self.cargar_dispositivos_hijos(show_error=False)
        target = next((d for d in dispositivos if d.get("id") == device_id), None)
        if not target:
            return

        self._selected_child_device_id = device_id
        self.entry_dispositivo_nombre.setText(str(target.get("nombre_optica", "")))
        self.entry_dispositivo_ciudad.setText(str(target.get("ciudad", "")))
        self.entry_dispositivo_codigo.setText(str(target.get("codigo_dispositivo", "")))

        estado = str(target.get("estado", "activo")).lower()
        self.combo_dispositivo_estado.setCurrentIndex(0 if estado == "activo" else 1)
        self.btn_guardar_dispositivo_hijo.setText("Actualizar dispositivo hijo")

    def guardar_dispositivo_hijo_ui(self):
        """Agrega o actualiza dispositivo hijo."""
        try:
            if hasattr(self, "device_type_combo") and self.device_type_combo.currentIndex() != 0:
                QMessageBox.warning(
                    self,
                    "Solo dispositivo madre",
                    "Para gestionar dispositivos hijos, este equipo debe estar configurado como 'Dispositivo madre'."
                )
                return

            if self._save_child_worker is not None and self._save_child_worker.isRunning():
                return

            limite = self._normalizar_max_sucursales(getattr(self, "_max_sucursales", 0), default=0)
            if not self._selected_child_device_id and limite == 1:
                QMessageBox.information(
                    self,
                    "Dispositivo unico",
                    (
                        "Tu cuenta es de una sola sucursal.\n\n"
                        "No se pueden crear dispositivos hijos porque este equipo ya corresponde a esa sucursal."
                    )
                )
                return

            nombre_optica = self.entry_dispositivo_nombre.text().strip()
            ciudad = self.entry_dispositivo_ciudad.text().strip()
            codigo = self.entry_dispositivo_codigo.text().strip().upper()
            estado = "activo" if self.combo_dispositivo_estado.currentIndex() == 0 else "bloqueado"

            if not nombre_optica:
                QMessageBox.warning(self, "Campos incompletos", "Ingrese el nombre de la optica/sucursal.")
                return

            if not codigo:
                self.generar_codigo_dispositivo_hijo()
                codigo = self.entry_dispositivo_codigo.text().strip().upper()
            self._start_child_device_loader()
            worker = SaveChildDeviceWorker(
                username=self.username,
                selected_child_id=self._selected_child_device_id,
                nombre_optica=nombre_optica,
                ciudad=ciudad,
                codigo=codigo,
                estado=estado,
                parent=self
            )
            worker.result_ready.connect(self._on_guardar_dispositivo_hijo_async_result)
            worker.finished.connect(self._on_guardar_dispositivo_hijo_worker_finished)
            worker.finished.connect(worker.deleteLater)
            self._save_child_worker = worker
            worker.start()
        except Exception as e:
            self._stop_child_device_loader()
            QMessageBox.warning(self, "Error", f"No se pudo guardar el dispositivo hijo.\n{e}")

    def copiar_codigo_dispositivo_hijo_ui(self):
        """Copia el código de vinculación actual al portapapeles."""
        codigo = self.entry_dispositivo_codigo.text().strip()
        if not codigo:
            QMessageBox.warning(self, "Sin código", "No hay ningún código seleccionado para copiar.")
            return

        try:
            from PyQt5.QtWidgets import QApplication
            clipboard = QApplication.clipboard()
            clipboard.setText(codigo)
            
            # Feedback visual temporal en el botón
            original_text = self.btn_copiar_codigo_dispositivo.text()
            self.btn_copiar_codigo_dispositivo.setText("¡Copiado!")
            self.btn_copiar_codigo_dispositivo.setStyleSheet(self.btn_copiar_codigo_dispositivo.styleSheet().replace("#00796b", "#2e7d32"))
            
            def restore_button():
                self.btn_copiar_codigo_dispositivo.setText(original_text)
                self.btn_copiar_codigo_dispositivo.setStyleSheet(self.btn_copiar_codigo_dispositivo.styleSheet().replace("#2e7d32", "#00796b"))
            
            QtCore.QTimer.singleShot(1500, restore_button)
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo copiar al portapapeles: {e}")

    def _eliminar_cache_local_sucursal(self, codigo_dispositivo: str):
        """Elimina cache local de branch para la sucursal borrada."""
        try:
            code = str(codigo_dispositivo or "").strip().upper()
            if not code:
                return
            cache_dir = os.path.join(VISO_DIR, str(self.username), "branch_cache", code)
            if os.path.isdir(cache_dir):
                shutil.rmtree(cache_dir, ignore_errors=True)
        except Exception:
            pass

    def _reset_contexto_sucursal_si_eliminada(self, codigo_dispositivo: str):
        """Si estaba seleccionada, vuelve a 'Todas las sucursales'."""
        try:
            code = str(codigo_dispositivo or "").strip().upper()
            if not code:
                return
            parent = getattr(self, "parent_app", None)
            if parent is None:
                return

            selected = str(getattr(parent, "selected_branch_code", "") or "").strip().upper()
            if selected != code:
                return

            try:
                from utils.file_handler import clear_active_branch_context
                clear_active_branch_context(self.username)
            except Exception:
                pass

            if hasattr(parent, "on_branch_context_changed"):
                try:
                    parent.on_branch_context_changed("", "Todas las sucursales")
                except Exception:
                    pass
        except Exception:
            pass

    def eliminar_dispositivo_hijo_ui(self):
        """Elimina el dispositivo hijo seleccionado en la lista."""
        try:
            if hasattr(self, "device_type_combo") and self.device_type_combo.currentIndex() != 0:
                QMessageBox.warning(
                    self,
                    "Solo dispositivo madre",
                    "Para gestionar dispositivos hijos, este equipo debe estar configurado como 'Dispositivo madre'."
                )
                return

            item = self.list_dispositivos_hijos.currentItem()
            if not item:
                QMessageBox.warning(self, "Seleccion requerida", "Seleccione un dispositivo hijo para eliminar.")
                return

            device_id = item.data(Qt.UserRole)
            dispositivos = getattr(self, "_dispositivos_hijos_cache", None)
            if not isinstance(dispositivos, list):
                dispositivos = self.cargar_dispositivos_hijos(show_error=True)
            seleccionado = next((d for d in dispositivos if d.get("id") == device_id), None)

            if not seleccionado:
                QMessageBox.warning(self, "Error", "No se pudo encontrar el dispositivo seleccionado.")
                return

            remote_ok = False
            remote_msg = ""
            try:
                from utils.api_handler import eliminar_dispositivo_hijo_remoto
                remote_ok, remote_msg = eliminar_dispositivo_hijo_remoto(
                    self.username,
                    device_id=str(device_id) if device_id else None,
                    codigo_dispositivo=(seleccionado or {}).get("codigo_dispositivo")
                )
            except Exception as sync_error:
                remote_msg = str(sync_error)

            if remote_ok:
                codigo_eliminado = str((seleccionado or {}).get("codigo_dispositivo", "")).strip().upper()
                self._eliminar_cache_local_sucursal(codigo_eliminado)
                self._reset_contexto_sucursal_si_eliminada(codigo_eliminado)
                self.update_dispositivos_hijos_list()
                if self.parent_app and hasattr(self.parent_app, "on_device_role_changed"):
                    try:
                        self.parent_app.on_device_role_changed("madre")
                    except Exception:
                        pass
                self.limpiar_form_dispositivo_hijo()
                QMessageBox.information(self, "Exito", f"Dispositivo hijo eliminado y sincronizado en nube.\n{remote_msg}")
            else:
                QMessageBox.warning(
                    self,
                    "Error eliminando en nube",
                    f"No se pudo eliminar el dispositivo en la base de datos remota.\nDetalle: {remote_msg}"
                )
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo eliminar el dispositivo.\n{e}")

    def info_nube_dispositivos_ui(self):
        """Muestra resumen cloud por dispositivo hijo (clientes/pacientes/productos)."""
        try:
            from utils.api_handler import obtener_resumen_nube_dispositivos

            ok, resumen, msg = obtener_resumen_nube_dispositivos(self.username)
            if not ok:
                QMessageBox.warning(
                    self,
                    "Resumen cloud no disponible",
                    f"No se pudo consultar la nube.\n\nDetalle: {msg}"
                )
                return

            if not resumen:
                QMessageBox.information(
                    self,
                    "Resumen cloud",
                    'No hay dispositivos hijos con datos en nube todavía.'
                )
                return

            lineas = ["Resumen por dispositivo hijo en nube:\n"]
            total_clientes = 0
            total_pacientes = 0
            total_productos = 0

            for item in resumen:
                nombre = item.get("nombre_optica") or "Sin nombre"
                ciudad = item.get("ciudad") or "Sin ciudad"
                codigo = item.get("codigo_dispositivo") or "Sin codigo"
                estado = str(item.get("estado", "activo")).upper()
                clientes = int(item.get("clientes", 0))
                pacientes = int(item.get("pacientes", 0))
                productos = int(item.get("productos", 0))

                total_clientes += clientes
                total_pacientes += pacientes
                total_productos += productos

                lineas.append(
                    f"[{estado}] {nombre} ({ciudad})\n"
                    f"Código: {codigo}\n"
                    f"Clientes: {clientes} | Pacientes: {pacientes} | Productos: {productos}\n"
                )

            lineas.append(
                f"Totales nube -> Clientes: {total_clientes}, Pacientes: {total_pacientes}, Productos: {total_productos}"
            )

            QMessageBox.information(self, "Resumen cloud", "\n".join(lineas))
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo obtener el resumen cloud.\n{e}")

    def importar_datos_generales_ui(self):
        try:
            # Mostrar diálogo para seleccionar archivo
            ruta_archivo, _ = QFileDialog.getOpenFileName(
                self,
                "Seleccionar Archivo de Respaldo",
                os.path.join(os.path.expanduser("~"), "Documents"),
                "Archivos ZIP (*.zip)"
            )
            
            if not ruta_archivo:
                return

            reply = QMessageBox.question(
                self,
                'Confirmar Importación',
                "¿Desea reemplazar todos los datos actuales con los del respaldo?\n\n"
                'Nota: Esta acción no afectará su clave de activación actual.',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                # Mostrar diálogo de progreso
                progress = QMessageBox(self)
                progress.setIcon(QMessageBox.Information)
                progress.setWindowTitle("Importando Datos")
                progress.setText("Restaurando copia de seguridad...\nPor favor espere.")
                progress.setStandardButtons(QMessageBox.NoButton)
                progress.show()
                QApplication.processEvents()

                try:
                    # Crear directorio temporal para extracción
                    temp_dir = os.path.join(VISO_DIR, "temp_import")
                    if os.path.exists(temp_dir):
                        shutil.rmtree(temp_dir)
                    os.makedirs(temp_dir)

                    # Extraer ZIP
                    with zipfile.ZipFile(ruta_archivo, 'r') as zipf:
                        zipf.extractall(temp_dir)

                    # Respaldar clave de activación actual si existe
                    user_data_dir = os.path.join(VISO_DIR, str(self.username))
                    activation_key_path = os.path.join(user_data_dir, "data", "clave_activacion.txt")
                    temp_key = None
                    if os.path.exists(activation_key_path):
                        with open(activation_key_path, 'r') as f:
                            temp_key = f.read()

                    # Limpiar directorio actual (excepto clave de activación)
                    if os.path.exists(user_data_dir):
                        for root, dirs, files in os.walk(user_data_dir):
                            for file in files:
                                if file.lower() != "clave_activacion.txt":
                                    os.remove(os.path.join(root, file))

                    # Copiar archivos del backup
                    for root, dirs, files in os.walk(temp_dir):
                        for file in files:
                            src_path = os.path.join(root, file)
                            rel_path = os.path.relpath(src_path, temp_dir)
                            dst_path = os.path.join(user_data_dir, rel_path)
                            os.makedirs(os.path.dirname(dst_path), exist_ok=True)
                            shutil.copy2(src_path, dst_path)

                    # Restaurar clave de activación
                    if temp_key is not None:
                        os.makedirs(os.path.dirname(activation_key_path), exist_ok=True)
                        with open(activation_key_path, 'w') as f:
                            f.write(temp_key)

                    progress.close()
                    reply = QMessageBox.information(
                        self,
                        'Éxito',
                        'Datos restaurados exitosamente.\nLa aplicación necesita reiniciarse para aplicar los cambios.',
                        QMessageBox.Ok
                    )

                    if reply == QMessageBox.Ok and hasattr(self.parent_app, 'restart_app'):
                        self.parent_app.restart_app()

                except Exception as e:
                    progress.close()
                    QMessageBox.critical(
                        self,
                        "Error",
                        f"Error al restaurar la copia de seguridad:\n{str(e)}"
                    )
                finally:
                    # Limpiar archivos temporales
                    if os.path.exists(temp_dir):
                        shutil.rmtree(temp_dir)

        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error inesperado:\n{str(e)}"
            )
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("MainContent")
        self.setMinimumSize(0, 0)
        self.setSizePolicy(QtWidgets.QSizePolicy.Ignored, QtWidgets.QSizePolicy.Expanding)
        self.parent_app = parent
        self.username = getattr(parent, 'username', None)
        self._max_sucursales = 0
        self._dispositivos_hijos_cache = []
        self._save_child_worker = None
        self._child_save_loader_timer = None
        self._child_save_loader_step = 0
        self._config_page_loader_widget = None
        self._config_page_loader_status_label = None
        self._config_page_loader_timer = None
        self._config_page_loader_step = 0
        self.auto_sync_thread = None  # Referencia al hilo de auto-sync
        self.lan_server = None        # Referencia al servidor LAN
        self._loading_dispositivos_hijos = False
        self._dispositivos_hijos_loaded_once = False
        
        # --- OPTIMIZACIÓN CRÍTICA: Lazy UI Construction ---
        self._initialized_tabs = {} 
        self._ui_fully_loaded = False
        
        # Diferir la construcción pesada para que la navegación sea instantánea
        QTimer.singleShot(10, self._deferred_setup_ui)

    def _deferred_setup_ui(self):
        """Construcción de la UI integrada en el widget actual"""
        try:
            # Si ya se cargó, no hacer nada
            if self._ui_fully_loaded: return
            
            self.setup_ui()
            self._ui_fully_loaded = True
            
            # Notificar que los textos deben normalizarse
            self._schedule_ui_texts_local(delay_ms=100)
            print("✅ ConfigPage: UI construida de forma integrada.")
        except Exception as e:
            print(f"❌ Error en deferred setup: {e}")

    def setup_ui(self):
        self._schedule_ui_texts_local(delay_ms=0)
        self._schedule_ui_texts_local(delay_ms=350)
        self._schedule_ui_texts_local(delay_ms=1200)


    def _yield_ui_for_loader(self):
        """Permite refrescar la UI durante cargas pesadas."""
        try:
            QtWidgets.QApplication.processEvents(QtCore.QEventLoop.AllEvents, 12)
        except Exception:
            pass


    def _build_config_page_loader_widget(self):
        container = QWidget(self)
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(24, 24, 24, 24)
        container_layout.setAlignment(Qt.AlignCenter)

        card = QFrame(container)
        card.setObjectName("ConfigLoaderCard")
        card.setFixedWidth(460)
        card.setStyleSheet(
            """
            QFrame#ConfigLoaderCard {
                background: qlineargradient(
                    x1:0, y1:0, x2:1, y2:1,
                    stop:0 #ffffff,
                    stop:1 #f7f9fc
                );
                border: 1px solid #e6ebf2;
                border-radius: 16px;
            }
            QLabel#ConfigLoaderTitle {
                font-size: 20px;
                font-weight: 700;
                color: #172b4d;
            }
            QLabel#ConfigLoaderSubtitle {
                font-size: 13px;
                color: #5e6c84;
            }
            QLabel#ConfigLoaderStatus {
                font-size: 12px;
                color: #4c9aff;
                font-weight: 600;
            }
            QProgressBar#ConfigLoaderBar {
                border: none;
                border-radius: 6px;
                background-color: #e9eef5;
                height: 12px;
            }
            QProgressBar#ConfigLoaderBar::chunk {
                border-radius: 6px;
                background: qlineargradient(
                    x1:0, y1:0, x2:1, y2:0,
                    stop:0 #36b37e,
                    stop:1 #00b8d9
                );
            }
            """
        )

        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(24, 24, 24, 24)
        card_layout.setSpacing(12)

        self._config_page_loader_title_label = QLabel("Cargando configuracion")
        self._config_page_loader_title_label.setObjectName("ConfigLoaderTitle")
        self._config_page_loader_title_label.setAlignment(Qt.AlignCenter)

        self._config_page_loader_subtitle_label = QLabel("Sincronizando con la nube. Esto puede tardar unos segundos.")
        self._config_page_loader_subtitle_label.setObjectName("ConfigLoaderSubtitle")
        self._config_page_loader_subtitle_label.setWordWrap(True)
        self._config_page_loader_subtitle_label.setAlignment(Qt.AlignCenter)

        bar = QtWidgets.QProgressBar()
        bar.setObjectName("ConfigLoaderBar")
        bar.setRange(0, 0)
        bar.setTextVisible(False)
        bar.setFixedHeight(12)

        self._config_page_loader_status_label = QLabel("Preparando datos")
        self._config_page_loader_status_label.setObjectName("ConfigLoaderStatus")
        self._config_page_loader_status_label.setAlignment(Qt.AlignCenter)

        card_layout.addWidget(self._config_page_loader_title_label)
        card_layout.addWidget(self._config_page_loader_subtitle_label)
        card_layout.addWidget(bar)
        card_layout.addWidget(self._config_page_loader_status_label)

        container_layout.addWidget(card, alignment=Qt.AlignCenter)
        return container


    def _show_config_page_loader(self, title="Cargando configuracion", subtitle="Sincronizando con la nube. Esto puede tardar unos segundos."):
        if self._config_page_loader_widget is None:
            return
        if getattr(self, "_config_page_loader_title_label", None) is not None:
            self._config_page_loader_title_label.setText(str(title or "Cargando configuracion"))
        if getattr(self, "_config_page_loader_subtitle_label", None) is not None:
            self._config_page_loader_subtitle_label.setText(str(subtitle or "Sincronizando con la nube. Esto puede tardar unos segundos."))
        self._config_page_loader_widget.setVisible(True)
        self._start_config_page_loader_animation()


    def _hide_config_page_loader(self):
        self._stop_config_page_loader_animation()
        if self._config_page_loader_widget is not None:
            self._config_page_loader_widget.setVisible(False)


    def _start_config_page_loader_animation(self):
        self._stop_config_page_loader_animation()
        self._config_page_loader_step = 0
        self._config_page_loader_timer = QtCore.QTimer(self)
        self._config_page_loader_timer.setInterval(340)
        self._config_page_loader_timer.timeout.connect(self._tick_config_page_loader)
        self._config_page_loader_timer.start()
        self._tick_config_page_loader()


    def _stop_config_page_loader_animation(self):
        try:
            if self._config_page_loader_timer is not None:
                self._config_page_loader_timer.stop()
                self._config_page_loader_timer.deleteLater()
        except Exception:
            pass
        self._config_page_loader_timer = None
        self._config_page_loader_step = 0


    def _tick_config_page_loader(self):
        label = getattr(self, "_config_page_loader_status_label", None)
        if label is None:
            return
        frames = (
            "Preparando datos",
            "Preparando datos.",
            "Preparando datos..",
            "Preparando datos...",
            "Cargando tarjetas",
            "Cargando tarjetas.",
            "Cargando tarjetas..",
            "Cargando tarjetas...",
        )
        idx = int(getattr(self, "_config_page_loader_step", 0)) % len(frames)
        label.setText(frames[idx])
        self._config_page_loader_step = idx + 1


    def setup_ui(self):
        page = self
        layout = QVBoxLayout(page)
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self._config_page_loader_widget = self._build_config_page_loader_widget()
        layout.addWidget(self._config_page_loader_widget, alignment=Qt.AlignCenter)
        self._show_config_page_loader()
        self._yield_ui_for_loader()

        # Inicializar el QTabWidget antes de usarlo
        self.tab_widget = QTabWidget()
        self.tab_widget.setMinimumSize(0, 0)
        self.tab_widget.setSizePolicy(QtWidgets.QSizePolicy.Ignored, QtWidgets.QSizePolicy.Expanding)
        # Evitar que la barra de tabs fuerce un ancho global mayor al viewport.
        # Cuando hay muchos tabs, usamos botones de desplazamiento y elisión.
        self.tab_widget.setUsesScrollButtons(True)
        tab_bar = self.tab_widget.tabBar()
        if tab_bar is not None:
            tab_bar.setUsesScrollButtons(True)
            tab_bar.setExpanding(False)
            tab_bar.setElideMode(Qt.ElideRight)
        self.tab_widget.setVisible(False)
        layout.addWidget(self.tab_widget)

        # Sección del logo
        logo_label = QLabel()
        logo_pixmap = QtGui.QPixmap(resource_path('icon.ico'))
        logo_escalado = logo_pixmap.scaledToWidth(150, Qt.TransformationMode.SmoothTransformation)
        logo_label.setPixmap(logo_escalado)
        logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(logo_label)

        layout.addWidget(QLabel('<h1>Configuración del Sistema</h1>', alignment=Qt.AlignmentFlag.AlignCenter))

        # --- TAB: Optómetras (rediseñado) ---
        optometras_tab = QWidget()
        optometras_layout = QVBoxLayout(optometras_tab)
        optometras_layout.setContentsMargins(20, 20, 20, 20)
        optometras_layout.setSpacing(20)

        # Título con icono
        titulo_widget = QWidget()
        titulo_layout = QHBoxLayout(titulo_widget)
        titulo_layout.setContentsMargins(0, 0, 0, 0)
        titulo_icon = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_DialogApplyButton)
        titulo_icon.setPixmap(icon.pixmap(32, 32))
        titulo_layout.addWidget(titulo_icon)
        titulo = QLabel('<h2>Gestión de Optómetras</h2>')
        titulo.setStyleSheet("color: #0d47a1; margin-bottom: 10px; font-weight: bold;")
        titulo_layout.addWidget(titulo)
        titulo_layout.addStretch()
        optometras_layout.addWidget(titulo_widget)

        # Tarjeta principal
        card = QGroupBox()
        card.setStyleSheet("""
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #bbdefb;
                border-radius: 10px;
                margin-top: 0px;
                /* box-shadow removed for Qt compatibility */
            }
        """)
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(15)
        card_layout.setContentsMargins(20, 20, 20, 20)

        # Formulario de agregar
        form_widget = QWidget()
        form_widget.setStyleSheet("""
            QWidget {
                background: #e3f2fd;
                border-radius: 8px;
                padding: 10px;
                border: 1px solid #bbdefb;
            }
        """)
        form_layout = QHBoxLayout(form_widget)
        form_layout.setContentsMargins(15, 15, 15, 15)
        form_layout.setSpacing(10)

        icon_label = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_FileIcon)
        icon_label.setPixmap(icon.pixmap(24, 24))
        form_layout.addWidget(icon_label)

        self.entry_optometra = QLineEdit()
        self.entry_optometra.setPlaceholderText('Nombre del nuevo optómetra')
        self.entry_optometra.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                min-width: 250px;
                color: #0d47a1;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #0d47a1;
                border-width: 2px;
                background: #fff;
                color: #0d47a1;
            }
        """)
        form_layout.addWidget(self.entry_optometra)

        btn_agregar_optometra = QPushButton('Agregar Optómetra')
        btn_agregar_optometra.setStyleSheet("""
            QPushButton {
                background: #1565c0;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #0d47a1;
            }
            QPushButton:pressed {
                background: #0a2472;
            }
        """)
        btn_agregar_optometra.clicked.connect(self.agregar_optometra)
        form_layout.addWidget(btn_agregar_optometra)
        card_layout.addWidget(form_widget)

        # Separador
        separador = QFrame()
        separador.setFrameShape(QFrame.HLine)
        separador.setFrameShadow(QFrame.Sunken)
        separador.setStyleSheet("background-color: #e0e0e0;")
        card_layout.addWidget(separador)

        # Lista de optómetras
        list_label = QLabel('<b>Optómetras Registrados</b>')
        list_label.setStyleSheet("color: #424242; font-size: 14px; margin-top: 10px;")
        card_layout.addWidget(list_label)

        self.list_optometras = QListWidget()
        self.list_optometras.setStyleSheet("""
            QListWidget {
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                padding: 5px;
                min-height: 200px;
                font-size: 13px;
            }
            QListWidget::item {
                padding: 10px;
                border-bottom: 1px solid #bbdefb;
                color: #0d47a1;
            }
            QListWidget::item:selected {
                background: #1976d2;
                color: white;
                font-weight: bold;
            }
            QListWidget::item:hover:!selected {
                background: #bbdefb;
                color: #0d47a1;
            }
        """)
        self.list_optometras.setSelectionMode(QAbstractItemView.ExtendedSelection)
        card_layout.addWidget(self.list_optometras)

        # Botones de acción
        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)
        buttons_layout.setContentsMargins(0, 10, 0, 0)
        buttons_layout.addStretch()

        btn_eliminar_optometra = QPushButton("Eliminar Seleccionados")
        btn_eliminar_optometra.setStyleSheet("""
            QPushButton {
                background: #d32f2f;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #b71c1c;
            }
            QPushButton:pressed {
                background: #7f0000;
            }
        """)
        btn_eliminar_optometra.clicked.connect(self.eliminar_optometra)
        buttons_layout.addWidget(btn_eliminar_optometra)
        card_layout.addWidget(buttons_widget)

        # Nota informativa
        nota = QLabel()
        nota.setWordWrap(True)
        nota.setStyleSheet("color: #1565c0; font-size: 12px; margin-top: 10px; background: #e3f2fd; padding: 10px; border-radius: 5px;")
        nota.setText(
            'Nota: Los optómetras registrados aquí aparecerán en los formularios de creación '
            "de pacientes y en los registros de consultas."
        )
        card_layout.addWidget(nota)

        optometras_layout.addWidget(card)
        optometras_layout.addStretch()
        self.optometras_tab_index = self.tab_widget.addTab(optometras_tab, 'Optómetras')
        self._yield_ui_for_loader()
        self.update_optometras_list()

        # --- TABs: Materiales/Tallas/Tipos de Lente removed from General Config
        # These sections now live under Inventario > Configuración.

        # --- TAB: Métodos de Pago ---
        pagos_tab = QWidget()
        pagos_layout = QVBoxLayout(pagos_tab)
        pagos_layout.setContentsMargins(20, 20, 20, 20)
        pagos_layout.setSpacing(20)

        # Título con icono
        titulo_widget = QWidget()
        titulo_layout = QHBoxLayout(titulo_widget)
        titulo_layout.setContentsMargins(0, 0, 0, 0)
        titulo_icon = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_DialogApplyButton)
        titulo_icon.setPixmap(icon.pixmap(32, 32))
        titulo_layout.addWidget(titulo_icon)
        titulo = QLabel('<h2>Métodos de Pago</h2>')
        titulo.setStyleSheet("color: #0d47a1; margin-bottom: 10px; font-weight: bold;")
        titulo_layout.addWidget(titulo)
        titulo_layout.addStretch()
        pagos_layout.addWidget(titulo_widget)

        # Tarjeta principal
        card = QGroupBox()
        card.setStyleSheet("""
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #bbdefb;
                border-radius: 10px;
                margin-top: 0px;
                /* box-shadow removed for Qt compatibility */
            }
        """)
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(15)
        card_layout.setContentsMargins(20, 20, 20, 20)

        # Formulario de agregar
        form_widget = QWidget()
        form_widget.setStyleSheet("""
            QWidget {
                background: #e3f2fd;
                border-radius: 8px;
                padding: 10px;
                border: 1px solid #bbdefb;
            }
        """)
        form_layout = QHBoxLayout(form_widget)
        form_layout.setContentsMargins(15, 15, 15, 15)
        form_layout.setSpacing(10)

        icon_label = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_FileIcon)
        icon_label.setPixmap(icon.pixmap(24, 24))
        form_layout.addWidget(icon_label)

        self.entry_pago = QLineEdit()
        self.entry_pago.setPlaceholderText("Ej. Efectivo, Yape, Plin")
        self.entry_pago.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                min-width: 250px;
                color: #0d47a1;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #0d47a1;
                border-width: 2px;
                background: #fff;
            }
        """)
        form_layout.addWidget(self.entry_pago)

        btn_agregar_pago = QPushButton('Agregar Método de Pago')
        btn_agregar_pago.setStyleSheet("""
            QPushButton {
                background: #1565c0;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #0d47a1;
            }
            QPushButton:pressed {
                background: #0a2472;
            }
        """)
        btn_agregar_pago.clicked.connect(self.agregar_metodo_pago)
        form_layout.addWidget(btn_agregar_pago)
        card_layout.addWidget(form_widget)

        # Lista de métodos de pago
        list_label = QLabel('<b>Métodos de Pago Registrados</b>')
        list_label.setStyleSheet("color: #424242; font-size: 14px; margin-top: 10px;")
        card_layout.addWidget(list_label)

        self.list_pagos = QListWidget()
        self.list_pagos.setStyleSheet("""
            QListWidget {
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                padding: 5px;
                min-height: 200px;
                font-size: 13px;
            }
            QListWidget::item {
                padding: 10px;
                border-bottom: 1px solid #bbdefb;
                color: #0d47a1;
            }
            QListWidget::item:selected {
                background: #1976d2;
                color: white;
                font-weight: bold;
            }
            QListWidget::item:hover:!selected {
                background: #bbdefb;
                color: #0d47a1;
            }
        """)
        self.list_pagos.setSelectionMode(QAbstractItemView.ExtendedSelection)
        card_layout.addWidget(self.list_pagos)

        # Botón eliminar
        btn_eliminar_pago = QPushButton("Eliminar Seleccionados")
        btn_eliminar_pago.setStyleSheet("""
            QPushButton {
                background: #d32f2f;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #b71c1c;
            }
            QPushButton:pressed {
                background: #7f0000;
            }
        """)
        btn_eliminar_pago.clicked.connect(self.eliminar_metodo_pago)
        card_layout.addWidget(btn_eliminar_pago)

        # Nota informativa
        nota = QLabel()
        nota.setWordWrap(True)
        nota.setStyleSheet("color: #1565c0; font-size: 12px; margin-top: 10px; background: #e3f2fd; padding: 10px; border-radius: 5px;")
        nota.setText(
            'Nota: Los métodos de pago registrados aquí estarán disponibles al momento de registrar ventas y transacciones.'
        )
        card_layout.addWidget(nota)

        pagos_layout.addWidget(card)
        pagos_layout.addStretch()
        self.tab_widget.addTab(pagos_tab, "Pagos")
        self._yield_ui_for_loader()
        self.update_pagos_list()

        # --- TAB: Configuración General ---
        config_tab = QWidget()
        config_layout = QVBoxLayout(config_tab)
        config_layout.setContentsMargins(20, 20, 20, 20)
        config_layout.setSpacing(20)

        # Título con icono
        titulo_widget = QWidget()
        titulo_layout = QHBoxLayout(titulo_widget)
        titulo_layout.setContentsMargins(0, 0, 0, 0)
        titulo_icon = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_FileDialogDetailedView)
        titulo_icon.setPixmap(icon.pixmap(32, 32))
        titulo_layout.addWidget(titulo_icon)
        titulo = QLabel('<h2>Configuración General</h2>')
        titulo.setStyleSheet("color: #0d47a1; margin-bottom: 10px; font-weight: bold;")
        titulo_layout.addWidget(titulo)
        titulo_layout.addStretch()
        config_layout.addWidget(titulo_widget)

        # Tarjeta de Información de la Óptica
        optica_card = QGroupBox('Información de la Óptica')
        optica_card.setStyleSheet("""
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #bbdefb;
                border-radius: 10px;
                margin-top: 20px;
                padding-top: 25px;
                font-size: 14px;
                color: #0d47a1;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 10px;
                background-color: #ffffff;
                color: #0d47a1;
                font-weight: bold;
            }
        """)
        optica_layout = QVBoxLayout(optica_card)
        optica_layout.setSpacing(15)
        optica_layout.setContentsMargins(20, 20, 20, 20)

        # Leer configuracion preferentemente desde MySQL y caer a cache local.
        datos_optica = cargar_datos_optica(self.username, prefer_remote=True)
        nombre_optica_val = str(datos_optica.get("nombre_optica", "") or cargar_nombre_optica(self.username) or "").strip()
        slogan_val = str(datos_optica.get("slogan", "") or "").strip()
        direccion_val = str(datos_optica.get("direccion", "") or "").strip()
        correo_val = str(datos_optica.get("correo_electronico", "") or "").strip()
        whatsapp_val = str(datos_optica.get("whatsapp", "") or cargar_whatsapp_optica(self.username) or "").strip()

        # Campo Nombre de Óptica
        nombre_widget = QWidget()
        nombre_layout = QHBoxLayout(nombre_widget)
        nombre_layout.setContentsMargins(0, 0, 0, 0)
        
        nombre_icon = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_FileDialogNewFolder)
        nombre_icon.setPixmap(icon.pixmap(24, 24))
        nombre_layout.addWidget(nombre_icon)
        
        nombre_label = QLabel("Nombre de la Óptica:")
        nombre_label.setStyleSheet("color: #0d47a1; font-weight: bold;")
        nombre_layout.addWidget(nombre_label)
        
        self.entry_nombre_optica = QLineEdit(nombre_optica_val)
        self.entry_nombre_optica.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                min-width: 250px;
                color: #0d47a1;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #0d47a1;
                border-width: 2px;
                background: #fff;
            }
        """)
        nombre_layout.addWidget(self.entry_nombre_optica)
        
        btn_guardar_nombre = QPushButton("Guardar")
        btn_guardar_nombre.setStyleSheet("""
            QPushButton {
                background: #1565c0;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #0d47a1;
            }
        """)
        btn_guardar_nombre.clicked.connect(self.guardar_nombre_optica)
        nombre_layout.addWidget(btn_guardar_nombre)
        
        optica_layout.addWidget(nombre_widget)

        # Campo Slogan
        slogan_widget = QWidget()
        slogan_layout = QHBoxLayout(slogan_widget)
        slogan_layout.setContentsMargins(0, 0, 0, 0)

        slogan_icon = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_MessageBoxInformation)
        slogan_icon.setPixmap(icon.pixmap(24, 24))
        slogan_layout.addWidget(slogan_icon)

        slogan_label = QLabel("Slogan:")
        slogan_label.setStyleSheet("color: #0d47a1; font-weight: bold;")
        slogan_layout.addWidget(slogan_label)

        self.entry_slogan_optica = QLineEdit(slogan_val)
        self.entry_slogan_optica.setPlaceholderText("Ej. Cuidamos tu visión con precisión")
        self.entry_slogan_optica.setStyleSheet(self.entry_nombre_optica.styleSheet())
        slogan_layout.addWidget(self.entry_slogan_optica)

        btn_guardar_slogan = QPushButton("Guardar")
        btn_guardar_slogan.setStyleSheet(btn_guardar_nombre.styleSheet())
        btn_guardar_slogan.clicked.connect(self.guardar_nombre_optica)
        slogan_layout.addWidget(btn_guardar_slogan)

        optica_layout.addWidget(slogan_widget)

        # Campo Dirección
        direccion_widget = QWidget()
        direccion_layout = QHBoxLayout(direccion_widget)
        direccion_layout.setContentsMargins(0, 0, 0, 0)

        direccion_icon = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_DialogOpenButton)
        direccion_icon.setPixmap(icon.pixmap(24, 24))
        direccion_layout.addWidget(direccion_icon)

        direccion_label = QLabel("Dirección:")
        direccion_label.setStyleSheet("color: #0d47a1; font-weight: bold;")
        direccion_layout.addWidget(direccion_label)

        self.entry_direccion_optica = QLineEdit(direccion_val)
        self.entry_direccion_optica.setPlaceholderText("Dirección principal de la óptica")
        self.entry_direccion_optica.setStyleSheet(self.entry_nombre_optica.styleSheet())
        direccion_layout.addWidget(self.entry_direccion_optica)

        btn_guardar_direccion = QPushButton("Guardar")
        btn_guardar_direccion.setStyleSheet(btn_guardar_nombre.styleSheet())
        btn_guardar_direccion.clicked.connect(self.guardar_nombre_optica)
        direccion_layout.addWidget(btn_guardar_direccion)

        optica_layout.addWidget(direccion_widget)

        # Campo Correo
        correo_widget = QWidget()
        correo_layout = QHBoxLayout(correo_widget)
        correo_layout.setContentsMargins(0, 0, 0, 0)

        correo_icon = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_FileDialogDetailedView)
        correo_icon.setPixmap(icon.pixmap(24, 24))
        correo_layout.addWidget(correo_icon)

        correo_label = QLabel("Correo electrónico:")
        correo_label.setStyleSheet("color: #0d47a1; font-weight: bold;")
        correo_layout.addWidget(correo_label)

        self.entry_correo_optica = QLineEdit(correo_val)
        self.entry_correo_optica.setPlaceholderText("Ej. contacto@optica.com")
        self.entry_correo_optica.setStyleSheet(self.entry_nombre_optica.styleSheet())
        correo_layout.addWidget(self.entry_correo_optica)

        btn_guardar_correo = QPushButton("Guardar")
        btn_guardar_correo.setStyleSheet(btn_guardar_nombre.styleSheet())
        btn_guardar_correo.clicked.connect(self.guardar_nombre_optica)
        correo_layout.addWidget(btn_guardar_correo)

        optica_layout.addWidget(correo_widget)

        # Campo WhatsApp
        whatsapp_widget = QWidget()
        whatsapp_layout = QHBoxLayout(whatsapp_widget)
        whatsapp_layout.setContentsMargins(0, 0, 0, 0)
        
        whatsapp_icon = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_CommandLink)
        whatsapp_icon.setPixmap(icon.pixmap(24, 24))
        whatsapp_layout.addWidget(whatsapp_icon)
        
        whatsapp_label = QLabel("WhatsApp personal:")
        whatsapp_label.setStyleSheet("color: #0d47a1; font-weight: bold;")
        whatsapp_layout.addWidget(whatsapp_label)
        
        self.entry_whatsapp = QLineEdit(whatsapp_val)
        self.entry_whatsapp.setPlaceholderText("Ej. 51987654321")
        self.entry_whatsapp.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                min-width: 250px;
                color: #0d47a1;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #0d47a1;
                border-width: 2px;
                background: #fff;
            }
        """)
        whatsapp_layout.addWidget(self.entry_whatsapp)
        
        btn_guardar_whatsapp = QPushButton("Guardar")
        btn_guardar_whatsapp.setStyleSheet("""
            QPushButton {
                background: #1565c0;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #0d47a1;
            }
        """)
        btn_guardar_whatsapp.clicked.connect(self.guardar_numero_whatsapp)
        whatsapp_layout.addWidget(btn_guardar_whatsapp)
        
        optica_layout.addWidget(whatsapp_widget)

        config_layout.addWidget(optica_card)

        # Tarjeta de Configurar Dispositivo
        device_card = QGroupBox("Configurar Dispositivo")
        device_card.setStyleSheet("""
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #bbdefb;
                border-radius: 10px;
                margin-top: 20px;
                padding-top: 25px;
                font-size: 14px;
                color: #0d47a1;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 10px;
                background-color: #ffffff;
                color: #0d47a1;
                font-weight: bold;
            }
        """)
        device_layout = QVBoxLayout(device_card)
        device_layout.setSpacing(12)
        device_layout.setContentsMargins(20, 20, 20, 20)

        device_row = QHBoxLayout()
        device_label = QLabel("Tipo de dispositivo:")
        device_label.setStyleSheet("color: #0d47a1; font-weight: bold;")
        device_row.addWidget(device_label)

        self.device_type_combo = QtWidgets.QComboBox()
        self.device_type_combo.addItems(["Dispositivo madre", "Dispositivo trabajador"])
        self.device_type_combo.currentIndexChanged.connect(self.actualizar_modo_dispositivo_ui)
        self.device_type_combo.setStyleSheet("""
            QComboBox {
                padding: 8px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                min-width: 250px;
                color: #0d47a1;
                font-size: 13px;
                font-weight: 600;
            }
            QComboBox:focus {
                border-color: #0d47a1;
                border-width: 2px;
            }
        """)
        device_row.addWidget(self.device_type_combo)
        device_row.addStretch()
        device_layout.addLayout(device_row)

        device_btn_row = QHBoxLayout()
        device_btn_row.addStretch()

        self.btn_guardar_tipo_dispositivo = QPushButton('Guardar configuración de dispositivo')
        self.btn_guardar_tipo_dispositivo.setStyleSheet("""
            QPushButton {
                background: #1565c0;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #0d47a1;
            }
        """)
        self.btn_guardar_tipo_dispositivo.clicked.connect(self.guardar_tipo_dispositivo_ui)
        device_btn_row.addWidget(self.btn_guardar_tipo_dispositivo)
        device_layout.addLayout(device_btn_row)

        self.device_mode_note = QLabel("")
        self.device_mode_note.setWordWrap(True)
        self.device_mode_note.setStyleSheet(
            "color: #1565c0; font-size: 12px; margin-top: 6px; "
            "background: #e3f2fd; padding: 10px; border-radius: 5px;"
        )
        device_layout.addWidget(self.device_mode_note)

        config_layout.addWidget(device_card)
        self.cargar_tipo_dispositivo_ui()

        # Tarjeta de Seguridad
        security_card = QGroupBox("Seguridad")
        security_card.setStyleSheet("""
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #bbdefb;
                border-radius: 10px;
                margin-top: 20px;
                padding-top: 25px;
                font-size: 14px;
                color: #0d47a1;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 10px;
                background-color: #ffffff;
                color: #0d47a1;
                font-weight: bold;
            }
        """)
        security_layout = QVBoxLayout(security_card)
        security_layout.setSpacing(15)
        security_layout.setContentsMargins(20, 20, 20, 20)

        # Campo Contraseña
        password_widget = QWidget()
        password_layout = QHBoxLayout(password_widget)
        password_layout.setContentsMargins(0, 0, 0, 0)
        
        password_icon = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_DialogApplyButton)
        password_icon.setPixmap(icon.pixmap(24, 24))
        password_layout.addWidget(password_icon)
        
        password_label = QLabel('Contraseña de configuración:')
        password_label.setStyleSheet("color: #0d47a1; font-weight: bold;")
        password_layout.addWidget(password_label)
        
        self.entry_password_setup = QLineEdit()
        self.entry_password_setup.setPlaceholderText('Establecer contrasena de 6 dígitos')
        self.entry_password_setup.setEchoMode(QLineEdit.Password)
        self.entry_password_setup.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                min-width: 250px;
                color: #0d47a1;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #0d47a1;
                border-width: 2px;
                background: #fff;
            }
        """)
        password_layout.addWidget(self.entry_password_setup)
        
        self.btn_guardar_password_setup = QPushButton('Establecer Contraseña')
        self.btn_guardar_password_setup.setStyleSheet("""
            QPushButton {
                background: #1565c0;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #0d47a1;
            }
            QPushButton:disabled {
                background: #90caf9;
                color: #e3f2fd;
            }
        """)
        self.btn_guardar_password_setup.clicked.connect(self.guardar_password_setup_ui)
        password_layout.addWidget(self.btn_guardar_password_setup)
        
        security_layout.addWidget(password_widget)

        # Nota de seguridad
        nota = QLabel()
        nota.setWordWrap(True)
        nota.setStyleSheet("color: #1565c0; font-size: 12px; margin-top: 10px; background: #e3f2fd; padding: 10px; border-radius: 5px;")
        nota.setText(
            'Nota: La contrasena de configuración se establece una sola vez y se usa para proteger '
            'el acceso a ajustes sensibles del sistema. Asegúrate de recordarla.'
        )
        security_layout.addWidget(nota)

        config_layout.addWidget(security_card)

        # ============================================================================
        # SECCIÓN DE MODOS (BÁSICO / NORMAL)
        # ============================================================================
        modos_card = QGroupBox("Modos de Visualización")
        modos_card.setStyleSheet("""
            QGroupBox {
                background-color: #FAFAFA;
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                margin-top: 16px;
                padding-top: 25px;
                font-size: 14px;
                color: #2C2C2C;
                font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 5px;
            }
        """)
        modos_layout = QVBoxLayout(modos_card)
        modos_layout.setContentsMargins(20, 20, 20, 20)

        self.checkbox_modo_basico = QCheckBox("Activar Modo Básico")
        self.checkbox_modo_basico.setStyleSheet("""
            QCheckBox {
                font-size: 14px;
                color: #333333;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
        """)
        # Cargar estado actual
        self.checkbox_modo_basico.setChecked(is_modo_basico(self.username))
        self.checkbox_modo_basico.toggled.connect(self._toggle_modo_basico)
        modos_layout.addWidget(self.checkbox_modo_basico)

        modos_info = QLabel(
            "El Modo Básico simplifica la interfaz para usuarios de la tercera edad o sin experiencia tecnológica.\n"
            "Ahora también puedes decidir qué módulos y botones rápidos mostrar."
        )
        modos_info.setWordWrap(True)
        modos_info.setStyleSheet("color: #666; font-size: 12px; font-style: italic;")
        modos_layout.addWidget(modos_info)

        self._modo_basico_page_checks = {}
        self._modo_basico_action_checks = {}
        basic_config = get_modo_basico_config(self.username)

        visible_group = QGroupBox("Módulos visibles en modo fácil")
        visible_group.setStyleSheet("QGroupBox { font-size: 13px; color: #334155; }")
        visible_layout = QGridLayout(visible_group)
        visible_layout.setHorizontalSpacing(16)
        visible_layout.setVerticalSpacing(10)

        for idx, (page_index, label) in enumerate(MODO_BASICO_PAGE_OPTIONS.items()):
            checkbox = QCheckBox(label)
            checkbox.setChecked(page_index in basic_config.get("visible_pages", []))
            checkbox.toggled.connect(self._save_modo_basico_preferences)
            if page_index in (0, 10):
                checkbox.setEnabled(False)
                checkbox.setToolTip("Siempre visible para evitar dejar la app sin navegación.")
            self._modo_basico_page_checks[page_index] = checkbox
            visible_layout.addWidget(checkbox, idx // 2, idx % 2)

        modos_layout.addWidget(visible_group)

        actions_group = QGroupBox("Botones rápidos de la pantalla Inicio")
        actions_group.setStyleSheet("QGroupBox { font-size: 13px; color: #334155; }")
        actions_layout = QGridLayout(actions_group)
        actions_layout.setHorizontalSpacing(16)
        actions_layout.setVerticalSpacing(10)

        for idx, (page_index, label) in enumerate(MODO_BASICO_HOME_ACTION_OPTIONS.items()):
            checkbox = QCheckBox(label)
            checkbox.setChecked(page_index in basic_config.get("quick_actions", []))
            checkbox.toggled.connect(self._save_modo_basico_preferences)
            self._modo_basico_action_checks[page_index] = checkbox
            actions_layout.addWidget(checkbox, idx // 2, idx % 2)

        modos_layout.addWidget(actions_group)
        self._sync_modo_basico_config_ui()

        config_layout.addWidget(modos_card)
        
        config_layout.addStretch()

        self.check_password_setup_state()
        self.tab_widget.addTab(config_tab, 'Configuración General')
        self._yield_ui_for_loader()
        
        # ========================
        # TAB: Información SUNAT
        # ========================
        sunat_tab = QWidget()
        sunat_tab_layout = QVBoxLayout(sunat_tab)
        sunat_tab_layout.setContentsMargins(30, 30, 30, 30)
        sunat_tab_layout.setSpacing(20)
        
        # Título
        sunat_title = QLabel('<h2>Información SUNAT / RUC</h2>')
        sunat_title.setStyleSheet("color: #1976D2; margin-bottom: 10px;")
        sunat_tab_layout.addWidget(sunat_title)
        
        # Card de SUNAT / RUC
        sunat_card = QGroupBox("Datos de la Empresa")
        sunat_card.setStyleSheet("""
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #bbdefb;
                border-radius: 10px;
                margin-top: 20px;
                padding-top: 25px;
                font-size: 14px;
                color: #0d47a1;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 10px;
                background-color: #ffffff;
                color: #0d47a1;
                font-weight: bold;
            }
        """)
        sunat_layout = QVBoxLayout(sunat_card)
        sunat_layout.setSpacing(15)
        sunat_layout.setContentsMargins(20, 20, 20, 20)

        # Cargar datos SUNAT (lazy loading - no cargar en setup_ui para mejor rendimiento)
        # Los datos se cargarán cuando el usuario navegue al tab
        datos_sunat = {}  # Valor vacío por defecto
        
        # Fila 1: RUC
        ruc_row = QHBoxLayout()
        ruc_row.setSpacing(10)
        ruc_label = QLabel("RUC:")
        ruc_label.setStyleSheet("color: #0d47a1; font-weight: bold; min-width: 120px;")
        self.entry_ruc = QLineEdit(datos_sunat.get("ruc", ""))
        self.entry_ruc.setPlaceholderText("Ej. 20131312955")
        self.entry_ruc.setMaxLength(11)
        self.entry_ruc.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                color: #0d47a1;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #0d47a1;
                border-width: 2px;
            }
        """)
        btn_consultar_ruc = QPushButton("Consultar SUNAT")
        btn_consultar_ruc.setMaximumWidth(150)
        btn_consultar_ruc.setStyleSheet("""
            QPushButton {
                background: #1565c0;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #0d47a1;
            }
        """)
        btn_consultar_ruc.clicked.connect(self.consultar_ruc_sunat)
        ruc_row.addWidget(ruc_label)
        ruc_row.addWidget(self.entry_ruc, 1)
        ruc_row.addWidget(btn_consultar_ruc)
        sunat_layout.addLayout(ruc_row)

        # Fila 2: Razón Social
        razon_row = QHBoxLayout()
        razon_row.setSpacing(10)
        razon_label = QLabel('Razón Social:')
        razon_label.setStyleSheet("color: #0d47a1; font-weight: bold; min-width: 120px;")
        self.entry_razon_social = QLineEdit(datos_sunat.get("razon_social", ""))
        self.entry_razon_social.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                color: #0d47a1;
                font-size: 13px;
            }
        """)
        razon_row.addWidget(razon_label)
        razon_row.addWidget(self.entry_razon_social, 1)
        sunat_layout.addLayout(razon_row)

        # Fila 3: Dirección
        direccion_row = QHBoxLayout()
        direccion_row.setSpacing(10)
        direccion_label = QLabel('Dirección:')
        direccion_label.setStyleSheet("color: #0d47a1; font-weight: bold; min-width: 120px;")
        self.entry_direccion = QLineEdit(datos_sunat.get("direccion", ""))
        self.entry_direccion.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                color: #0d47a1;
                font-size: 13px;
            }
        """)
        direccion_row.addWidget(direccion_label)
        direccion_row.addWidget(self.entry_direccion, 1)
        sunat_layout.addLayout(direccion_row)

        # Fila 4: Ubicación (Depto, Provincia, Distrito)
        ubicacion_row = QHBoxLayout()
        ubicacion_row.setSpacing(10)
        ubicacion_label = QLabel('Ubicación:')
        ubicacion_label.setStyleSheet("color: #0d47a1; font-weight: bold; min-width: 120px;")
        self.entry_departamento = QLineEdit(datos_sunat.get("departamento", ""))
        self.entry_departamento.setPlaceholderText("Departamento")
        self.entry_departamento.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                color: #0d47a1;
                font-size: 13px;
            }
        """)
        self.entry_provincia = QLineEdit(datos_sunat.get("provincia", ""))
        self.entry_provincia.setPlaceholderText("Provincia")
        self.entry_provincia.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                color: #0d47a1;
                font-size: 13px;
            }
        """)
        self.entry_distrito = QLineEdit(datos_sunat.get("distrito", ""))
        self.entry_distrito.setPlaceholderText("Distrito")
        self.entry_distrito.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                color: #0d47a1;
                font-size: 13px;
            }
        """)
        ubicacion_row.addWidget(ubicacion_label)
        ubicacion_row.addWidget(self.entry_departamento, 1)
        ubicacion_row.addWidget(self.entry_provincia, 1)
        ubicacion_row.addWidget(self.entry_distrito, 1)
        sunat_layout.addLayout(ubicacion_row)

        # Fila 5: Estado / Condición
        estado_row = QHBoxLayout()
        estado_row.setSpacing(10)
        estado_label = QLabel('Estado/Condición:')
        estado_label.setStyleSheet("color: #0d47a1; font-weight: bold; min-width: 120px;")
        self.entry_estado = QLineEdit(datos_sunat.get("estado", ""))
        self.entry_estado.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: #f5f5f5;
                color: #0d47a1;
                font-size: 13px;
            }
        """)
        self.entry_estado.setReadOnly(True)
        self.entry_condicion = QLineEdit(datos_sunat.get("condicion", ""))
        self.entry_condicion.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: #f5f5f5;
                color: #0d47a1;
                font-size: 13px;
            }
        """)
        self.entry_condicion.setReadOnly(True)
        estado_row.addWidget(estado_label)
        estado_row.addWidget(self.entry_estado, 1)
        estado_row.addWidget(self.entry_condicion, 1)
        sunat_layout.addLayout(estado_row)

        # Fila 6: Botón Guardar
        btn_row = QHBoxLayout()
        btn_guardar_sunat = QPushButton("Guardar Datos SUNAT")
        btn_guardar_sunat.setMaximumWidth(200)
        btn_guardar_sunat.setStyleSheet("""
            QPushButton {
                background: #0d7377;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #0a5f68;
            }
        """)
        btn_guardar_sunat.clicked.connect(self.guardar_datos_sunat)
        btn_row.addWidget(btn_guardar_sunat)
        btn_row.addStretch()
        sunat_layout.addLayout(btn_row)
        
        sunat_tab_layout.addWidget(sunat_card)
        
        # Información adicional
        info_label = QLabel(
            "<b>Consulta SUNAT:</b><br>"
            "Ingresa el RUC (11 dígitos) de tu empresa y haz clic en 'Consultar SUNAT' "
            'para obtener automáticamente los datos de la empresa.<br><br>'
            "<b style='color: #d32f2f;'>IMPORTANTE:</b> El uso de esta información es <b>exclusivamente para boletas sin validez legal</b>. "
            ""
        )
        info_label.setStyleSheet("""
            QLabel {
                background-color: #fff3e0;
                border: 2px solid #ff9800;
                border-radius: 5px;
                padding: 15px;
                color: #e65100;
                font-size: 12px;
            }
        """)
        info_label.setWordWrap(True)
        sunat_tab_layout.addWidget(info_label)
        
        # ============================================
        # NUEVA SECCIÓN: EMISIÓN ELECTRÓNICA
        # ============================================
        emision_sep = QFrame()
        emision_sep.setFrameShape(QFrame.HLine)
        emision_sep.setStyleSheet("background-color: #1976d2; height: 2px; margin: 20px 0;")
        sunat_tab_layout.addWidget(emision_sep)
        
        emision_title = QLabel('<h3>Emisión Electrónica a SUNAT</h3>')
        emision_title.setStyleSheet("color: #1565c0; margin-top: 20px; margin-bottom: 10px;")
        sunat_tab_layout.addWidget(emision_title)
        
        # AVISO: Sistema en desarrollo
        aviso_desarrollo = QLabel(
            '<b>SISTEMA EN DESARROLLO</b><br>'
            'La funcionalidad de emisión electrónica está siendo integrada. '
            'Próximamente podrá generar y enviar boletas y facturas electrónicas directamente a SUNAT. '
            'Por el momento, esta sección es solo para configuración futura.'
            "Importante si el usuario intenta probar el sistema es bajo su propia responsabilidad."
            "En futuras actualizaciones se vendra implementando las mejoras y este sistema no estara disponible para todos los usuarios"
        )
        aviso_desarrollo.setStyleSheet("""
            QLabel {
                background-color: #fff3cd;
                border: 2px solid #ffc107;
                border-radius: 8px;
                padding: 15px;
                color: #856404;
                font-size: 12px;
                font-weight: bold;
            }
        """)
        aviso_desarrollo.setWordWrap(True)
        sunat_tab_layout.addWidget(aviso_desarrollo)
        
        # Card de Configuración de Emisión Electrónica
        emision_card = QGroupBox('Configuración de Emisión Electrónica')
        emision_card.setStyleSheet("""
            QGroupBox {
                background-color: #e3f2fd;
                border: 2px solid #1976d2;
                border-radius: 10px;
                margin-top: 20px;
                padding-top: 25px;
                font-size: 14px;
                color: #0d47a1;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 10px;
                background-color: #e3f2fd;
                color: #1565c0;
                font-weight: bold;
            }
        """)
        emision_layout = QVBoxLayout(emision_card)
        emision_layout.setSpacing(15)
        emision_layout.setContentsMargins(20, 20, 20, 20)
        
        # Fila 1: Estado de Emisión Electrónica
        estado_emision_row = QHBoxLayout()
        estado_emision_label = QLabel('Estado de Emisión Electrónica:')
        estado_emision_label.setStyleSheet("color: #0d47a1; font-weight: bold; min-width: 180px;")
        
        self.btn_habilitar_emision = QPushButton('Deshabilitada')
        self.btn_habilitar_emision.setCheckable(True)
        self.btn_habilitar_emision.setMaximumWidth(180)
        self.btn_habilitar_emision.setStyleSheet("""
            QPushButton {
                background: #d32f2f;
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background: #b71c1c;
            }
            QPushButton:checked {
                background: #4caf50;
            }
            QPushButton:checked:hover {
                background: #388e3c;
            }
        """)
        self.btn_habilitar_emision.clicked.connect(self.toggle_emision_electronica)
        
        estado_emision_row.addWidget(estado_emision_label)
        estado_emision_row.addWidget(self.btn_habilitar_emision)
        estado_emision_row.addStretch()
        emision_layout.addLayout(estado_emision_row)
        
        # Separador
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("background-color: #bbdefb;")
        emision_layout.addWidget(sep)
        
        # Fila 2: Usuario SOL
        sol_row = QHBoxLayout()
        sol_label = QLabel("Usuario SOL:")
        sol_label.setStyleSheet("color: #0d47a1; font-weight: bold; min-width: 180px;")
        self.entry_usuario_sol = QLineEdit()
        self.entry_usuario_sol.setPlaceholderText("Ej. usuario.sol@sunat.gob.pe")
        self.entry_usuario_sol.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                color: #0d47a1;
                font-size: 13px;
            }
        """)
        sol_row.addWidget(sol_label)
        sol_row.addWidget(self.entry_usuario_sol, 1)
        emision_layout.addLayout(sol_row)
        
        # Fila 3: Contraseña SOL
        pass_row = QHBoxLayout()
        pass_label = QLabel('Contraseña SOL:')
        pass_label.setStyleSheet("color: #0d47a1; font-weight: bold; min-width: 180px;")
        self.entry_password_sol = QLineEdit()
        self.entry_password_sol.setPlaceholderText('Contraseña SOL')
        self.entry_password_sol.setEchoMode(QLineEdit.Password)
        self.entry_password_sol.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                color: #0d47a1;
                font-size: 13px;
            }
        """)
        pass_row.addWidget(pass_label)
        pass_row.addWidget(self.entry_password_sol, 1)
        emision_layout.addLayout(pass_row)
        
        # Fila 4: Certificado Digital
        cert_row = QHBoxLayout()
        cert_label = QLabel("Certificado Digital:")
        cert_label.setStyleSheet("color: #0d47a1; font-weight: bold; min-width: 180px;")
        
        self.label_cert_estado = QLabel("No cargado")
        self.label_cert_estado.setStyleSheet("color: #d32f2f; font-weight: bold;")
        
        btn_cargar_cert = QPushButton("Cargar Certificado (.pem/.pfx)")
        btn_cargar_cert.setMaximumWidth(200)
        btn_cargar_cert.setStyleSheet("""
            QPushButton {
                background: #1565c0;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background: #0d47a1;
            }
        """)
        btn_cargar_cert.clicked.connect(self.cargar_certificado_digital)
        
        cert_row.addWidget(cert_label)
        cert_row.addWidget(self.label_cert_estado, 1)
        cert_row.addWidget(btn_cargar_cert)
        emision_layout.addLayout(cert_row)
        
        # Fila 5: Clave Privada
        key_row = QHBoxLayout()
        key_label = QLabel("Clave Privada:")
        key_label.setStyleSheet("color: #0d47a1; font-weight: bold; min-width: 180px;")
        
        self.label_key_estado = QLabel("No cargada")
        self.label_key_estado.setStyleSheet("color: #d32f2f; font-weight: bold;")
        
        btn_cargar_key = QPushButton("Cargar Clave Privada (.key/.pem)")
        btn_cargar_key.setMaximumWidth(200)
        btn_cargar_key.setStyleSheet("""
            QPushButton {
                background: #1565c0;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background: #0d47a1;
            }
        """)
        btn_cargar_key.clicked.connect(self.cargar_clave_privada)
        
        key_row.addWidget(key_label)
        key_row.addWidget(self.label_key_estado, 1)
        key_row.addWidget(btn_cargar_key)
        emision_layout.addLayout(key_row)
        
        # Fila 6: Ambiente
        ambiente_row = QHBoxLayout()
        ambiente_label = QLabel("Ambiente:")
        ambiente_label.setStyleSheet("color: #0d47a1; font-weight: bold; min-width: 180px;")
        
        self.combo_ambiente = QtWidgets.QComboBox()
        self.combo_ambiente.addItem("Testing/Desarrollo")
        self.combo_ambiente.addItem('Producción')
        self.combo_ambiente.setStyleSheet("""
            QComboBox {
                padding: 8px 12px;
                border: 1px solid #1976d2;
                border-radius: 5px;
                background: white;
                color: #0d47a1;
            }
        """)
        
        ambiente_row.addWidget(ambiente_label)
        ambiente_row.addWidget(self.combo_ambiente)
        ambiente_row.addStretch()
        emision_layout.addLayout(ambiente_row)
        
        # Fila 7: Opciones
        opciones_row = QHBoxLayout()
        self.check_enviar_auto = QtWidgets.QCheckBox('Enviar automáticamente a SUNAT')
        self.check_enviar_auto.setStyleSheet("color: #0d47a1; font-weight: bold;")
        
        self.check_guardar_cdr = QtWidgets.QCheckBox("Guardar CDR localmente")
        self.check_guardar_cdr.setChecked(True)
        self.check_guardar_cdr.setStyleSheet("color: #0d47a1; font-weight: bold;")
        
        opciones_row.addWidget(self.check_enviar_auto)
        opciones_row.addWidget(self.check_guardar_cdr)
        opciones_row.addStretch()
        emision_layout.addLayout(opciones_row)
        
        # Botones de acción
        botones_row = QHBoxLayout()
        
        btn_probar_conexion = QPushButton('Probar Conexión SUNAT')
        btn_probar_conexion.setMaximumWidth(180)
        btn_probar_conexion.setStyleSheet("""
            QPushButton {
                background: #f57c00;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background: #e65100;
            }
        """)
        btn_probar_conexion.clicked.connect(self.probar_conexion_sunat)
        
        btn_guardar_emision = QPushButton('Guardar Configuración')
        btn_guardar_emision.setMaximumWidth(180)
        btn_guardar_emision.setStyleSheet("""
            QPushButton {
                background: #0d7377;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background: #0a5f68;
            }
        """)
        btn_guardar_emision.clicked.connect(self.guardar_config_emision_electronica)
        
        botones_row.addWidget(btn_probar_conexion)
        botones_row.addWidget(btn_guardar_emision)
        botones_row.addStretch()
        emision_layout.addLayout(botones_row)
        
        sunat_tab_layout.addWidget(emision_card)
        
        # Nota importante
        nota_emision = QLabel(
            '• <b>Emisión Electrónica SUNAT:</b><br>'
            "Necesitas:<br>"
            '• Certificado digital vigente (descargable desde SUNAT)<br>'
            "• Usuario y contraseña SOL<br>"
            "• Habilitación como emisor electrónico en SUNAT<br><br>"
            'Una vez configurado, podrás emitir boletas y facturas válidas ante SUNAT '
            'que se registrarán automáticamente en sus sistemas.'
        )
        nota_emision.setStyleSheet("""
            QLabel {
                background-color: #c8e6c9;
                border: 2px solid #4caf50;
                border-radius: 5px;
                padding: 15px;
                color: #1b5e20;
                font-size: 12px;
            }
        """)
        nota_emision.setWordWrap(True)
        sunat_tab_layout.addWidget(nota_emision)
        
        sunat_tab_layout.addStretch()
        
        # Envolver en QScrollArea
        sunat_scroll = QScrollArea()
        sunat_scroll.setWidget(sunat_tab)
        sunat_scroll.setWidgetResizable(True)
        sunat_scroll.setStyleSheet("""
            QScrollArea {
                background-color: white;
                border: none;
            }
            QScrollBar:vertical {
                border: none;
                background: #f5f5f5;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #2196F3;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #1976D2;
            }
        """)
        
        self.tab_widget.addTab(sunat_scroll, 'Información SUNAT')
        self._yield_ui_for_loader()
        
        # ========================
        # TAB: LAN (Compartir en red)
        # ========================
        lan_tab = QWidget()
        lan_layout = QVBoxLayout(lan_tab)
        lan_layout.setContentsMargins(30, 30, 30, 30)
        lan_layout.setSpacing(20)

        # Titulo
        lan_header = QWidget()
        lan_header_layout = QHBoxLayout(lan_header)
        lan_header_layout.setContentsMargins(0, 0, 0, 0)

        lan_icon = QLabel("LAN")
        lan_icon.setStyleSheet("font-size: 18px; font-weight: bold; color: #1976D2;")
        lan_title = QLabel("<h2>Configuracion de Red Local (LAN)</h2>")
        lan_title.setStyleSheet("color: #1976D2; margin-bottom: 5px;")

        lan_header_layout.addWidget(lan_icon)
        lan_header_layout.addWidget(lan_title)
        lan_header_layout.addStretch()
        lan_layout.addWidget(lan_header)

        # Card principal
        lan_card = QGroupBox("Conexion multiusuario")
        lan_card.setStyleSheet("""
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #bbdefb;
                border-radius: 10px;
                margin-top: 20px;
                padding-top: 25px;
                font-size: 14px;
                color: #0d47a1;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 10px;
                background-color: #ffffff;
                color: #0d47a1;
                font-weight: bold;
            }
        """)
        card_layout = QVBoxLayout(lan_card)
        card_layout.setSpacing(20)
        card_layout.setContentsMargins(25, 25, 25, 25)

        # Selector de modo
        mode_row = QHBoxLayout()
        mode_label = QLabel("Modo de operacion:")
        mode_label.setStyleSheet("font-weight: bold; color: #424242; font-size: 13px;")

        self.lan_mode_combo = QtWidgets.QComboBox()
        self.lan_mode_combo.addItems(["Servidor principal (Host)", "Terminal punto de venta (Cliente)"])
        self.lan_mode_combo.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 1px solid #1976D2;
                border-radius: 5px;
                min-width: 200px;
                font-weight: bold;
            }
        """)
        self.lan_mode_combo.currentIndexChanged.connect(self.update_lan_ui_state)

        mode_row.addWidget(mode_label)
        mode_row.addWidget(self.lan_mode_combo)
        mode_row.addStretch()
        card_layout.addLayout(mode_row)

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("background-color: #E0E0E0;")
        card_layout.addWidget(line)

        # Seccion servidor
        self.server_widget = QWidget()
        server_layout = QGridLayout(self.server_widget)
        server_layout.setContentsMargins(0, 0, 0, 0)
        server_layout.setSpacing(15)

        lbl_my_ip = QLabel("Tu direccion IP local:")
        lbl_my_ip.setStyleSheet("color: #424242; font-weight: 500;")

        self.entry_my_ip = QLineEdit()
        self.entry_my_ip.setReadOnly(True)
        self.entry_my_ip.setStyleSheet("background-color: #F5F5F5; border: 1px solid #BDBDBD; padding: 8px; border-radius: 4px; color: #616161;")

        lbl_port_server = QLabel("Puerto de comunicacion:")
        lbl_port_server.setStyleSheet("color: #424242; font-weight: 500;")

        self.entry_port_server = QLineEdit("5000")
        self.entry_port_server.setPlaceholderText("Ej: 5000")
        self.entry_port_server.setStyleSheet("border: 1px solid #1976D2; padding: 8px; border-radius: 4px;")

        info_server = QLabel(
            "Configura esta PC como 'Servidor' si aqui esta instalada la base de datos principal. "
            "Las otras computadoras se conectaran a esta IP."
        )
        info_server.setStyleSheet("color: #1565C0; font-size: 11px; font-style: italic;")
        info_server.setWordWrap(True)

        server_layout.addWidget(lbl_my_ip, 0, 0)
        server_layout.addWidget(self.entry_my_ip, 0, 1)
        server_layout.addWidget(lbl_port_server, 1, 0)
        server_layout.addWidget(self.entry_port_server, 1, 1)
        server_layout.addWidget(info_server, 2, 0, 1, 2)
        card_layout.addWidget(self.server_widget)

        # Seccion cliente
        self.client_widget = QWidget()
        client_layout = QGridLayout(self.client_widget)
        client_layout.setContentsMargins(0, 0, 0, 0)
        client_layout.setSpacing(15)

        lbl_host_ip = QLabel("IP del servidor principal:")
        lbl_host_ip.setStyleSheet("color: #424242; font-weight: 500;")

        self.entry_host_ip = QLineEdit()
        self.entry_host_ip.setPlaceholderText("Ej: 192.168.1.XX")
        self.entry_host_ip.setStyleSheet("border: 1px solid #1976D2; padding: 8px; border-radius: 4px;")

        lbl_port_client = QLabel("Puerto del servidor:")
        lbl_port_client.setStyleSheet("color: #424242; font-weight: 500;")

        self.entry_port_client = QLineEdit("5000")
        self.entry_port_client.setPlaceholderText("Ej: 5000")
        self.entry_port_client.setStyleSheet("border: 1px solid #1976D2; padding: 8px; border-radius: 4px;")

        info_client = QLabel(
            "Configura esta PC como 'Terminal' para conectarte a la computadora principal. "
            "Debes ingresar la IP que aparece en la configuracion del servidor."
        )
        info_client.setStyleSheet("color: #2E7D32; font-size: 11px; font-style: italic;")
        info_client.setWordWrap(True)

        client_layout.addWidget(lbl_host_ip, 0, 0)
        client_layout.addWidget(self.entry_host_ip, 0, 1)
        client_layout.addWidget(lbl_port_client, 1, 0)
        client_layout.addWidget(self.entry_port_client, 1, 1)
        client_layout.addWidget(info_client, 2, 0, 1, 2)

        self.check_auto_sync = QtWidgets.QCheckBox("Activar sincronizacion en tiempo real (Auto-Sync)")
        self.check_auto_sync.setStyleSheet("font-weight: bold; color: #1565C0;")
        self.check_auto_sync.setToolTip("Descarga datos del servidor automaticamente cada 5 segundos")
        client_layout.addWidget(self.check_auto_sync, 3, 0, 1, 2)

        card_layout.addWidget(self.client_widget)

        # Botones
        btn_layout = QHBoxLayout()

        self.btn_save_lan = QPushButton("Guardar configuracion")
        self.btn_save_lan.setStyleSheet("""
            QPushButton {
                background-color: #1976D2;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #1565C0; }
        """)
        self.btn_save_lan.clicked.connect(self.save_lan_config)

        self.btn_stop_lan = QPushButton("Detener / Desconectar")
        self.btn_stop_lan.setStyleSheet("""
            QPushButton {
                background-color: #D32F2F;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #B71C1C; }
        """)
        self.btn_stop_lan.clicked.connect(self.stop_lan_service)

        self.btn_test_connection = QPushButton("Probar conexion")
        self.btn_test_connection.setStyleSheet("""
            QPushButton {
                background-color: #F57C00;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #EF6C00; }
        """)
        self.btn_test_connection.clicked.connect(self.test_lan_connection)

        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_stop_lan)
        btn_layout.addWidget(self.btn_test_connection)
        btn_layout.addWidget(self.btn_save_lan)
        card_layout.addLayout(btn_layout)
        lan_layout.addWidget(lan_card)

        self.lbl_lan_status = QLabel("Estado: Sin configurar")
        self.lbl_lan_status.setAlignment(Qt.AlignCenter)
        self.lbl_lan_status.setStyleSheet("background-color: #EEEEEE; padding: 10px; border-radius: 5px; color: #757575;")
        lan_layout.addWidget(self.lbl_lan_status)

        self.btn_open_web_dashboard = QPushButton("Abrir Dashboard en navegador")
        self.btn_open_web_dashboard.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #1976D2;
                border: 1px solid #1976D2;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
                margin-top: 5px;
            }
            QPushButton:hover {
                background-color: #E3F2FD;
            }
        """)
        self.btn_open_web_dashboard.clicked.connect(self.open_lan_dashboard)
        lan_layout.addWidget(self.btn_open_web_dashboard)

        lan_disabled_notice = QLabel(LAN_DISABLED_MESSAGE)
        lan_disabled_notice.setWordWrap(True)
        lan_disabled_notice.setStyleSheet(
            "background-color: #FFF3E0; color: #B45309; border: 1px solid #F59E0B; "
            "padding: 12px; border-radius: 6px; font-weight: bold;"
        )
        lan_layout.addWidget(lan_disabled_notice)

        lan_layout.addStretch()
        self.tab_widget.addTab(lan_tab, "LAN")
        self._yield_ui_for_loader()

        # TAB: Dispositivos Hijos (Cloud-ready)
        # ========================
        child_devices_tab = QWidget()
        child_devices_layout = QVBoxLayout(child_devices_tab)
        child_devices_layout.setContentsMargins(20, 20, 20, 20)
        child_devices_layout.setSpacing(16)

        child_title = QLabel('<h2>Gestión de Dispositivos Hijos</h2>')
        child_title.setStyleSheet("color: #1976D2; margin-bottom: 2px;")
        child_devices_layout.addWidget(child_title)

        child_subtitle = QLabel(
            "Administra sucursales/dispositivos hijos desde el panel madre. "
            'Esta estructura ya queda lista para sincronización en nube.'
        )
        child_subtitle.setWordWrap(True)
        child_subtitle.setStyleSheet("color: #546e7a; font-size: 12px;")
        child_devices_layout.addWidget(child_subtitle)

        self.lbl_dispositivos_limite = QLabel('Límite de sucursales: cargando...')
        self.lbl_dispositivos_limite.setStyleSheet("color: #546e7a; font-size: 12px;")
        child_devices_layout.addWidget(self.lbl_dispositivos_limite)

        self.child_devices_access_note = QLabel(
            'Esta pestaña se habilita solo cuando el equipo está marcado como Dispositivo madre.'
        )
        self.child_devices_access_note.setWordWrap(True)
        self.child_devices_access_note.setStyleSheet(
            "color: #c62828; font-size: 12px; background: #ffebee; "
            "padding: 8px; border-radius: 4px;"
        )
        child_devices_layout.addWidget(self.child_devices_access_note)

        # Card: formulario
        child_form_card = QGroupBox("Registro de Sucursal / Dispositivo")
        child_form_card.setStyleSheet("""
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #bbdefb;
                border-radius: 10px;
                margin-top: 12px;
                padding-top: 20px;
                font-size: 13px;
                color: #0d47a1;
                font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 10px;
                background-color: #ffffff;
                color: #0d47a1;
            }
        """)
        child_form_layout = QGridLayout(child_form_card)
        child_form_layout.setContentsMargins(18, 16, 18, 16)
        child_form_layout.setHorizontalSpacing(12)
        child_form_layout.setVerticalSpacing(10)

        child_form_layout.addWidget(QLabel('Nombre de la óptica/sucursal:'), 0, 0)
        self.entry_dispositivo_nombre = QLineEdit()
        self.entry_dispositivo_nombre.setPlaceholderText('Ej: Óptica Centro')
        child_form_layout.addWidget(self.entry_dispositivo_nombre, 0, 1)

        child_form_layout.addWidget(QLabel("Ciudad / sede:"), 1, 0)
        self.entry_dispositivo_ciudad = QLineEdit()
        self.entry_dispositivo_ciudad.setPlaceholderText("Ej: Lima")
        child_form_layout.addWidget(self.entry_dispositivo_ciudad, 1, 1)

        child_form_layout.addWidget(QLabel('Código de vinculación:'), 2, 0)
        code_row = QWidget()
        code_row_layout = QHBoxLayout(code_row)
        code_row_layout.setContentsMargins(0, 0, 0, 0)
        code_row_layout.setSpacing(8)
        self.entry_dispositivo_codigo = QLineEdit()
        self.entry_dispositivo_codigo.setReadOnly(True)
        self.entry_dispositivo_codigo.setReadOnly(True)
        self.entry_dispositivo_codigo.setStyleSheet("background-color: #f5f5f5; color: #555;")
        code_row_layout.addWidget(self.entry_dispositivo_codigo)

        # Botón para Generar otro código aleatorio
        self.btn_generar_codigo_dispositivo = QPushButton("Generar")
        self.btn_generar_codigo_dispositivo.setToolTip('Generar un nuevo código aleatorio')
        self.btn_generar_codigo_dispositivo.setStyleSheet("QPushButton { padding: 6px 12px; font-weight: bold; }")
        self.btn_generar_codigo_dispositivo.clicked.connect(self.generar_codigo_dispositivo_hijo)
        code_row_layout.addWidget(self.btn_generar_codigo_dispositivo)

        child_form_layout.addWidget(code_row, 2, 1)
        child_form_layout.addWidget(QLabel("Estado:"), 3, 0)
        self.combo_dispositivo_estado = QtWidgets.QComboBox()
        self.combo_dispositivo_estado.addItems(["Activo", "Bloqueado"])
        child_form_layout.addWidget(self.combo_dispositivo_estado, 3, 1)

        child_buttons_row = QWidget()
        child_buttons_layout = QHBoxLayout(child_buttons_row)
        child_buttons_layout.setContentsMargins(0, 8, 0, 0)
        child_buttons_layout.setSpacing(8)
        self.btn_guardar_dispositivo_hijo = QPushButton("Agregar dispositivo hijo")
        self.btn_guardar_dispositivo_hijo.setStyleSheet("""
            QPushButton {
                background: #1565c0;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover { background: #0d47a1; }
        """)
        self.btn_guardar_dispositivo_hijo.pressed.connect(self._animar_boton_agregar_dispositivo_hijo)
        self.btn_guardar_dispositivo_hijo.clicked.connect(self.guardar_dispositivo_hijo_ui)
        child_buttons_layout.addWidget(self.btn_guardar_dispositivo_hijo)

        self.btn_limpiar_dispositivo_hijo = QPushButton("Limpiar")
        self.btn_limpiar_dispositivo_hijo.clicked.connect(self.limpiar_form_dispositivo_hijo)
        child_buttons_layout.addWidget(self.btn_limpiar_dispositivo_hijo)

        self.lbl_guardar_dispositivo_loader = QLabel("Guardando sucursal...")
        self.lbl_guardar_dispositivo_loader.setStyleSheet(
            "color: #1976D2; font-size: 12px; font-weight: 600; padding-left: 8px;"
        )
        self.lbl_guardar_dispositivo_loader.setVisible(False)
        child_buttons_layout.addWidget(self.lbl_guardar_dispositivo_loader)

        child_buttons_layout.addStretch()
        child_form_layout.addWidget(child_buttons_row, 4, 0, 1, 2)

        child_devices_layout.addWidget(child_form_card)

        # Card: lista de dispositivos
        child_list_card = QGroupBox("Dispositivos Hijos Registrados")
        child_list_card.setStyleSheet("""
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #e3f2fd;
                border-radius: 10px;
                margin-top: 10px;
                padding-top: 20px;
                font-size: 13px;
                color: #0d47a1;
                font-weight: bold;
            }
        """)
        child_list_layout = QVBoxLayout(child_list_card)
        child_list_layout.setContentsMargins(18, 16, 18, 16)
        child_list_layout.setSpacing(10)

        self.lbl_dispositivos_total = QLabel('Dispositivos registrados: 0 | Límite: no definido')
        self.lbl_dispositivos_total.setStyleSheet("color: #546e7a; font-size: 12px;")
        child_list_layout.addWidget(self.lbl_dispositivos_total)

        self.list_dispositivos_hijos = QListWidget()
        self.list_dispositivos_hijos.setSelectionMode(QAbstractItemView.SingleSelection)
        self.list_dispositivos_hijos.itemClicked.connect(self.seleccionar_dispositivo_hijo)
        child_list_layout.addWidget(self.list_dispositivos_hijos)

        child_actions_row = QWidget()
        child_actions_layout = QHBoxLayout(child_actions_row)
        child_actions_layout.setContentsMargins(0, 0, 0, 0)
        child_actions_layout.setSpacing(8)

        self.btn_eliminar_dispositivo_hijo = QPushButton("Eliminar seleccionado")
        self.btn_eliminar_dispositivo_hijo.setStyleSheet("""
            QPushButton {
                background: #d32f2f;
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover { background: #b71c1c; }
        """)
        self.btn_eliminar_dispositivo_hijo.clicked.connect(self.eliminar_dispositivo_hijo_ui)
        child_actions_layout.addWidget(self.btn_eliminar_dispositivo_hijo)

        self.btn_info_nube_dispositivos = QPushButton("Info nube")
        self.btn_info_nube_dispositivos.clicked.connect(self.info_nube_dispositivos_ui)
        child_actions_layout.addWidget(self.btn_info_nube_dispositivos)

        self.btn_recargar_dispositivos_hijos = QPushButton("Recargar dispositivos")
        self.btn_recargar_dispositivos_hijos.clicked.connect(self.recargar_dispositivos_hijos_ui)
        child_actions_layout.addWidget(self.btn_recargar_dispositivos_hijos)

        self.btn_copiar_codigo_dispositivo = QPushButton("Copiar Código")
        self.btn_copiar_codigo_dispositivo.setStyleSheet("""
            QPushButton {
                background: #00796b;
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover { background: #004d40; }
        """)
        self.btn_copiar_codigo_dispositivo.clicked.connect(self.copiar_codigo_dispositivo_hijo_ui)
        child_actions_layout.addWidget(self.btn_copiar_codigo_dispositivo)

        child_actions_layout.addStretch()
        child_list_layout.addWidget(child_actions_row)

        child_devices_layout.addWidget(child_list_card)
        child_devices_layout.addStretch()

        self.child_devices_tab_index = self.tab_widget.addTab(child_devices_tab, "Dispositivos Hijos")
        self._yield_ui_for_loader()
        self._selected_child_device_id = None
        self.limpiar_form_dispositivo_hijo()
        
        # ⚠️ OPTIMIZACIÓN: Se ha quitado update_dispositivos_hijos_list() de aquí
        # porque hacía peticiones de red síncronas que bloqueaban la UI 10 segundos.
        # Ahora se cargan en _on_tab_changed.
        
        self.actualizar_modo_dispositivo_ui()
        
        # Inicializar estado UI (sin carga pesada)
        # self.load_lan_config() # Movido a carga bajo demanda si es lento
        self._yield_ui_for_loader()
        
        # Conectar signal para cargar datos cuando se cambia de pestaña
        self.tab_widget.currentChanged.connect(self._on_tab_changed)
        
        # ========================
        # TAB: Gestión de Ayudantes
        # ========================
        try:
            helpers_tab = HelpersPage(parent=self, username=self.username)
            self.tab_widget.addTab(helpers_tab, "Ayudantes")
            self._yield_ui_for_loader()
        except Exception as e:
            # Si hay error, crear una pestaña con mensaje de error
            error_tab = QWidget()
            error_layout = QVBoxLayout(error_tab)
            error_label = QLabel(f"Error al cargar gestión de ayudantes: {str(e)}")
            error_label.setStyleSheet("color: red;")
            error_layout.addWidget(error_label)
            self.tab_widget.addTab(error_tab, "Ayudantes")
            self._yield_ui_for_loader()
        
        # Mover la creación de los tabs de base de datos y backup aquí
        self.setup_db_and_backup_tabs()
        self._yield_ui_for_loader()

        # ========================
        # TAB: Términos de Uso
        # ========================
        terms_tab = QWidget()
        terms_layout = QVBoxLayout(terms_tab)
        terms_layout.setContentsMargins(30, 30, 30, 30)
        terms_layout.setSpacing(20)

        title = QLabel('Términos de Uso')
        title.setStyleSheet("font-size: 22px; font-weight: bold; color: #1976d2; margin-bottom: 10px;")
        terms_layout.addWidget(title)

        # Texto largo de Términos de Uso (sincronizado con terms_dialog.py)
        terms_text = QtWidgets.QTextEdit()
        terms_text.setReadOnly(True)
        terms_text.setStyleSheet("""
            QTextEdit {
                color: #424242;
                font-size: 12px;
                background-color: #FFFFFF;
                border: 1px solid #E8E8E8;
                border-radius: 4px;
                padding: 12px;
            }
        """)
        
        # Importar el texto desde terms_dialog para mantener coherencia
        try:
            from gui.dialogs.terms_dialog import TermsDialog
            terms_text.setPlainText(TermsDialog.get_terms_text())
        except Exception:
            # Fallback si no se puede importar
            terms_text.setPlainText('Términos y Condiciones de Uso - Consulte la ventana de inicio para los términos completos.')
        
        terms_layout.addWidget(terms_text)

        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setStyleSheet("color: #E0E0E0;")
        terms_layout.addWidget(separator)

        acceptance_widget = QWidget()
        acceptance_layout = QHBoxLayout(acceptance_widget)
        acceptance_layout.setContentsMargins(0, 0, 0, 0)

        self.terms_accepted_checkbox = QPushButton("Acepto los Términos y Condiciones")
        self.terms_accepted_checkbox.setCheckable(True)
        self.terms_accepted_checkbox.setChecked(True)
        self.terms_accepted_checkbox.setStyleSheet("""
            QPushButton {
                background-color: #191919;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px;
                font-weight: bold;
                font-size: 12px;
                text-align: left;
                padding-left: 20px;
            }
            QPushButton:hover {
                background-color: #191919;
            }
            QPushButton:pressed {
                background-color: #191919;
            }
        """)
        self.terms_accepted_checkbox.setFlat(True)
        acceptance_layout.addWidget(self.terms_accepted_checkbox)
        acceptance_layout.addStretch()

        terms_layout.addWidget(acceptance_widget)

        confirmation_text = QLabel(
            'Al continuar usando esta aplicación, confirmas que has leído y aceptas estos términos.'
        )
        confirmation_text.setWordWrap(True)
        confirmation_text.setStyleSheet("""
            QLabel {
                color: #D32F2F;
                font-size: 10px;
                font-weight: bold;
                padding: 8px 0px;
                border-top: 1px solid #E0E0E0;
                margin-top: 8px;
            }
        """)
        terms_layout.addWidget(confirmation_text)

        self.tab_widget.addTab(terms_tab, 'Términos de Uso')
        self._yield_ui_for_loader()

        # ========================
        # CONTACTO DEL DESARROLLADOR
        # ========================
        contacto_tab = QWidget()
        main_layout = QVBoxLayout(contacto_tab)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        main_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Título con icono
        titulo_widget = QWidget()
        titulo_layout = QHBoxLayout(titulo_widget)
        titulo_layout.setContentsMargins(0, 0, 0, 0)
        titulo_icon = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_MessageBoxInformation)
        titulo_icon.setPixmap(icon.pixmap(32, 32))
        titulo_layout.addWidget(titulo_icon)
        titulo = QLabel('<h2>Soporte Técnico VISO</h2>')
        titulo.setStyleSheet("color: #1976D2; margin-bottom: 10px;")
        titulo_layout.addWidget(titulo)
        titulo_layout.addStretch()
        main_layout.addWidget(titulo_widget)

        # Tarjeta de contacto
        contact_card = QGroupBox()
        contact_card.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 10px;
                margin-top: 0px;
            }
        """)
        card_layout = QVBoxLayout(contact_card)
        card_layout.setSpacing(15)
        card_layout.setContentsMargins(20, 20, 20, 20)

        # Mensaje principal
        mensaje_principal = QLabel()
        mensaje_principal.setWordWrap(True)
        mensaje_principal.setStyleSheet("font-size: 14px; color: #424242;")
        mensaje_principal.setText(
            '¿Necesitas ayuda con VISO? Estamos aquí para apoyarte.\n'
            "Contacta directamente al desarrollador para resolver cualquier problema."
        )
        card_layout.addWidget(mensaje_principal)

        # Número de WhatsApp con icono
        whatsapp_widget = QWidget()
        whatsapp_layout = QHBoxLayout(whatsapp_widget)
        whatsapp_layout.setContentsMargins(0, 0, 0, 0)
        whatsapp_icon = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_CommandLink)
        whatsapp_icon.setPixmap(icon.pixmap(24, 24))
        whatsapp_layout.addWidget(whatsapp_icon)
        numero_label = QLabel("<b>WhatsApp:</b> +51 972330654")
        numero_label.setStyleSheet("font-size: 16px; color: #191919;")
        whatsapp_layout.addWidget(numero_label)
        whatsapp_layout.addStretch()
        
        # Botón para copiar número
        btn_copiar = QPushButton('Copiar número')
        btn_copiar.setFixedWidth(120)
        btn_copiar.setStyleSheet("""
            QPushButton {
                background: #f5f5f5;
                border: 1px solid #e0e0e0;
                padding: 5px 10px;
                color: #424242;
                border-radius: 5px;
            }
            QPushButton:hover {
                background: #eeeeee;
                border-color: #bdbdbd;
            }
        """)
        def copiar_numero():
            cb = QApplication.clipboard()
            cb.setText("+51972330654")
            btn_copiar.setText("Copiado!")
            QtCore.QTimer.singleShot(2000, lambda: btn_copiar.setText('Copiar número'))
        btn_copiar.clicked.connect(copiar_numero)
        whatsapp_layout.addWidget(btn_copiar)
        
        card_layout.addWidget(whatsapp_widget)

        # Separador
        separador = QFrame()
        separador.setFrameShape(QFrame.HLine)
        separador.setFrameShadow(QFrame.Sunken)
        separador.setStyleSheet("background-color: #e0e0e0;")
        card_layout.addWidget(separador)

        # Botones de acción
        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)
        buttons_layout.setContentsMargins(0, 0, 0, 0)
        buttons_layout.setSpacing(10)

        # Botón de WhatsApp Web
        btn_contactar = QPushButton("Abrir WhatsApp Web")
        btn_contactar.setObjectName("primaryButton")
        btn_contactar.setStyleSheet("""
            QPushButton {
                background: #2196F3;
                color: white;
                border: none;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #1976D2;
            }
        """)
        def open_whatsapp():
            numero = "51972330654"
            mensaje = "Hola%20tengo%20un%20problema%20con%20VISO%20-%20por%20favor%20ayuda"
            url = f"https://wa.me/{numero}?text={mensaje}"
            try:
                webbrowser.open(url)
            except Exception:
                QMessageBox.information(self, "Abrir WhatsApp", f"Abre este enlace en tu navegador: {url}")
        btn_contactar.clicked.connect(open_whatsapp)
        buttons_layout.addWidget(btn_contactar)

        # Botón de correo (opcional para futuro)
        btn_email = QPushButton("Enviar correo")
        btn_email.setStyleSheet("""
            QPushButton {
                background: #f5f5f5;
                color: #424242;
                border: 1px solid #e0e0e0;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #eeeeee;
                border-color: #bdbdbd;
            }
        """)
        def open_email():
            webbrowser.open("mailto:soporte@viso.com")
        btn_email.clicked.connect(open_email)
        buttons_layout.addWidget(btn_email)

        card_layout.addWidget(buttons_widget)

        # Nota final
        nota = QLabel()
        nota.setWordWrap(True)
        nota.setStyleSheet("color: #757575; font-size: 12px;")
        nota.setText(
            'Nota: El soporte está disponible de lunes a viernes de 9:00 AM a 6:00 PM (GMT-5). '
            'Para una respuesta más rápida, incluye una descripción clara del problema y capturas de pantalla si es posible.'
        )
        card_layout.addWidget(nota)
        
        # Footer con créditos
        footer_frame = QFrame()
        footer_frame.setStyleSheet("""
            QFrame {
                background-color: #f9f9f9;
                border-top: 1px solid #e0e0e0;
                border-radius: 10px;
                margin-top: 10px;
            }
        """)
        footer_layout = QVBoxLayout(footer_frame)
        footer_layout.setContentsMargins(15, 10, 15, 10)
        footer_layout.setSpacing(2)
        
        footer_text = QLabel("Hecho por emprendedores para ayudarte a crecer")
        footer_text.setStyleSheet("color: #1976D2; font-size: 12px; font-weight: bold; text-align: center;")
        footer_text.setAlignment(Qt.AlignLeft)
        footer_layout.addWidget(footer_text)
        
        footer_credit = QLabel("Att. Equipo Sistemas Peruanos A&M")
        footer_credit.setStyleSheet("color: #757575; font-size: 10px; text-align: center;")
        footer_credit.setAlignment(Qt.AlignLeft)
        footer_layout.addWidget(footer_credit)
        
        card_layout.addWidget(footer_frame)

        main_layout.addWidget(contact_card)
        main_layout.addStretch()
        self.tab_widget.addTab(contacto_tab, "Contacto")
        self._yield_ui_for_loader()

        # --- TAB: Importar ---
        importar_tab = QWidget()
        importar_layout = QVBoxLayout(importar_tab)
        importar_layout.setContentsMargins(0, 0, 0, 0)
        importar_layout.setSpacing(0)
        
        # Crear scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("""
            QScrollArea {
                background-color: white;
                border: none;
            }
            QScrollBar:vertical {
                border: none;
                background: #f5f5f5;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #2196F3;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #1976D2;
            }
        """)
        
        # Contenedor dentro del scroll
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setContentsMargins(20, 20, 20, 20)
        scroll_layout.setSpacing(20)
        
        # Título
        titulo_importar = QLabel("<h2>Importar</h2>")
        titulo_importar.setStyleSheet("color: #1976D2;")
        scroll_layout.addWidget(titulo_importar)
        
        # Contenedor de tarjetas en grid (3 columnas)
        cards_container_importar = QWidget()
        cards_grid = QGridLayout(cards_container_importar)
        cards_grid.setSpacing(20)
        cards_grid.setContentsMargins(0, 0, 0, 0)
        
        # ========== CARD 1: IMPORTAR CLIENTES ==========
        card_clientes = QGroupBox()
        card_clientes.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 2px solid #2196F3;
                border-radius: 10px;
                margin-top: 0px;
            }
        """)
        card_clientes.setMinimumHeight(250)
        layout_clientes = QVBoxLayout(card_clientes)
        layout_clientes.setSpacing(15)
        layout_clientes.setContentsMargins(20, 20, 20, 20)
        
        titulo_clientes = QLabel("Importar Clientes")
        titulo_clientes.setStyleSheet("font-size: 16px; font-weight: bold; color: #1976D2;")
        layout_clientes.addWidget(titulo_clientes)
        
        desc_clientes = QLabel("Importa una lista de clientes desde un archivo Excel o CSV.")
        desc_clientes.setWordWrap(True)
        desc_clientes.setStyleSheet("color: #666666; font-size: 12px;")
        layout_clientes.addWidget(desc_clientes)
        
        separador_clientes = QFrame()
        separador_clientes.setFrameShape(QFrame.HLine)
        separador_clientes.setStyleSheet("background-color: #e0e0e0;")
        layout_clientes.addWidget(separador_clientes)
        
        btn_clientes = QPushButton("Seleccionar Archivo")
        btn_clientes.setStyleSheet("""
            QPushButton {
                background: #2196F3;
                color: white;
                border: none;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #1976D2;
            }
        """)
        btn_clientes.setCursor(Qt.PointingHandCursor)
        btn_clientes.clicked.connect(self.importar_clientes_archivo)
        layout_clientes.addWidget(btn_clientes)
        
        # Botón para descargar plantilla
        btn_plantilla = QPushButton("Descargar Plantilla")
        btn_plantilla.setStyleSheet("""
            QPushButton {
                background: #4CAF50;
                color: white;
                border: none;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #45a049;
            }
        """)
        btn_plantilla.setCursor(Qt.PointingHandCursor)
        btn_plantilla.clicked.connect(self.descargar_plantilla_clientes)
        layout_clientes.addWidget(btn_plantilla)
        
        layout_clientes.addStretch()
        
        cards_grid.addWidget(card_clientes, 0, 0)
        
        # ========== CARD 2: IMPORTAR PACIENTES (DESHABILITADO) ==========
        card_pacientes = QGroupBox()
        card_pacientes.setStyleSheet("""
            QGroupBox {
                background-color: #f5f5f5;
                border: 2px solid #cccccc;
                border-radius: 10px;
                margin-top: 0px;
            }
        """)
        card_pacientes.setMinimumHeight(250)
        layout_pacientes = QVBoxLayout(card_pacientes)
        layout_pacientes.setSpacing(15)
        layout_pacientes.setContentsMargins(20, 20, 20, 20)
        
        titulo_pacientes = QLabel("Importar Pacientes")
        titulo_pacientes.setStyleSheet("font-size: 16px; font-weight: bold; color: #999999;")
        layout_pacientes.addWidget(titulo_pacientes)
        
        desc_pacientes = QLabel("Importa una lista de pacientes desde un archivo Excel o CSV.")
        desc_pacientes.setWordWrap(True)
        desc_pacientes.setStyleSheet("color: #999999; font-size: 12px;")
        layout_pacientes.addWidget(desc_pacientes)
        
        separador_pacientes = QFrame()
        separador_pacientes.setFrameShape(QFrame.HLine)
        separador_pacientes.setStyleSheet("background-color: #cccccc;")
        layout_pacientes.addWidget(separador_pacientes)
        
        btn_pacientes = QPushButton("Seleccionar Archivo")
        btn_pacientes.setStyleSheet("""
            QPushButton {
                background: #cccccc;
                color: #666666;
                border: none;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #bbbbbb;
            }
        """)
        btn_pacientes.setCursor(Qt.ForbiddenCursor)
        btn_pacientes.setEnabled(False)
        layout_pacientes.addWidget(btn_pacientes)
        
        label_deshabilitado_pacientes = QLabel('Esta función está deshabilitada por el momento')
        label_deshabilitado_pacientes.setStyleSheet("color: #ff9800; font-size: 11px; font-weight: bold;")
        label_deshabilitado_pacientes.setAlignment(Qt.AlignCenter)
        layout_pacientes.addWidget(label_deshabilitado_pacientes)
        layout_pacientes.addStretch()
        
        cards_grid.addWidget(card_pacientes, 0, 1)
        
        # ========== CARD 3: IMPORTAR INVENTARIO (DESHABILITADO) ==========
        card_inventario = QGroupBox()
        card_inventario.setStyleSheet("""
            QGroupBox {
                background-color: #f5f5f5;
                border: 2px solid #cccccc;
                border-radius: 10px;
                margin-top: 0px;
            }
        """)
        card_inventario.setMinimumHeight(250)
        layout_inventario = QVBoxLayout(card_inventario)
        layout_inventario.setSpacing(15)
        layout_inventario.setContentsMargins(20, 20, 20, 20)
        
        titulo_inventario = QLabel("Importar Inventario")
        titulo_inventario.setStyleSheet("font-size: 16px; font-weight: bold; color: #999999;")
        layout_inventario.addWidget(titulo_inventario)
        
        desc_inventario = QLabel("Importa una lista de productos del inventario desde un archivo Excel o CSV.")
        desc_inventario.setWordWrap(True)
        desc_inventario.setStyleSheet("color: #999999; font-size: 12px;")
        layout_inventario.addWidget(desc_inventario)
        
        separador_inventario = QFrame()
        separador_inventario.setFrameShape(QFrame.HLine)
        separador_inventario.setStyleSheet("background-color: #cccccc;")
        layout_inventario.addWidget(separador_inventario)
        
        btn_inventario = QPushButton("Seleccionar Archivo")
        btn_inventario.setStyleSheet("""
            QPushButton {
                background: #cccccc;
                color: #666666;
                border: none;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #bbbbbb;
            }
        """)
        btn_inventario.setCursor(Qt.ForbiddenCursor)
        btn_inventario.setEnabled(False)
        layout_inventario.addWidget(btn_inventario)
        
        label_deshabilitado_inventario = QLabel('Esta función está deshabilitada por el momento')
        label_deshabilitado_inventario.setStyleSheet("color: #ff9800; font-size: 11px; font-weight: bold;")
        label_deshabilitado_inventario.setAlignment(Qt.AlignCenter)
        layout_inventario.addWidget(label_deshabilitado_inventario)
        layout_inventario.addStretch()
        
        cards_grid.addWidget(card_inventario, 0, 2)
        
        scroll_layout.addWidget(cards_container_importar)
        scroll_layout.addStretch()
        
        scroll_area.setWidget(scroll_content)
        importar_layout.addWidget(scroll_area)
        
        self.tab_widget.addTab(importar_tab, "Importar")
        self._yield_ui_for_loader()

        # --- TAB: Plantilla ---
        plantilla_tab = QWidget()
        main_layout = QVBoxLayout(plantilla_tab)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)

        # Título
        titulo = QLabel("<h2>Plantilla</h2>")
        titulo.setStyleSheet("color: #1976D2;")
        main_layout.addWidget(titulo)

        # Cargar plantilla seleccionada
        from utils.generador_boletas_plantilla import GeneradorBoletasPlantilla
        generador = GeneradorBoletasPlantilla(self.username)
        plantilla_actual = generador.plantilla_seleccionada

        # Contenedor de tarjetas
        cards_container = QWidget()
        cards_layout = QHBoxLayout(cards_container)
        cards_layout.setSpacing(15)
        cards_layout.setContentsMargins(0, 0, 0, 0)

        # Card 1: Diseño Pequeño
        card_pequeno = QGroupBox()
        card_pequeno.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 10px;
                margin-top: 0px;
            }
        """)
        layout_pequeno = QVBoxLayout(card_pequeno)
        layout_pequeno.setSpacing(15)
        layout_pequeno.setContentsMargins(20, 20, 20, 20)

        titulo_pequeno = QLabel('Diseño Pequeño')
        titulo_pequeno.setStyleSheet("font-size: 16px; font-weight: bold; color: #1976D2;")
        layout_pequeno.addWidget(titulo_pequeno)

        desc_pequeno = QLabel('Plantilla de boleta compacta y condensada para impresoras térmicas.')
        desc_pequeno.setWordWrap(True)
        desc_pequeno.setStyleSheet("color: #666666; font-size: 13px;")
        layout_pequeno.addWidget(desc_pequeno)

        separador2 = QFrame()
        separador2.setFrameShape(QFrame.HLine)
        separador2.setStyleSheet("background-color: #e0e0e0;")
        layout_pequeno.addWidget(separador2)

        btn_pequeno = QPushButton("Seleccionada" if plantilla_actual == "pequena" else "Seleccionar Diseño Pequeño")
        estilo_pequeno = """
            QPushButton {
                background: #757575;
                color: white;
                border: none;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #616161;
            }
        """ if plantilla_actual == 'pequena' else """
            QPushButton {
                background: #2196F3;
                color: white;
                border: none;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #1976D2;
            }
        """
        btn_pequeno.setStyleSheet(estilo_pequeno)
        btn_pequeno.setCursor(Qt.PointingHandCursor)
        btn_pequeno.clicked.connect(lambda: self.seleccionar_plantilla('pequena'))
        layout_pequeno.addWidget(btn_pequeno)
        self.btn_pequeno_ref = btn_pequeno  # Guardar referencia

        cards_layout.addWidget(card_pequeno)

        # Card 2: Diseño Extra Largo
        card_extra_largo = QGroupBox()
        card_extra_largo.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 10px;
                margin-top: 0px;
            }
        """)
        layout_extra_largo = QVBoxLayout(card_extra_largo)
        layout_extra_largo.setSpacing(15)
        layout_extra_largo.setContentsMargins(20, 20, 20, 20)

        titulo_extra_largo = QLabel('Diseño Extra Largo')
        titulo_extra_largo.setStyleSheet("font-size: 16px; font-weight: bold; color: #1976D2;")
        layout_extra_largo.addWidget(titulo_extra_largo)

        desc_extra_largo = QLabel('Plantilla de boleta con formato completo y detallado con toda la información disponible.')
        desc_extra_largo.setWordWrap(True)
        desc_extra_largo.setStyleSheet("color: #666666; font-size: 13px;")
        layout_extra_largo.addWidget(desc_extra_largo)

        separador3 = QFrame()
        separador3.setFrameShape(QFrame.HLine)
        separador3.setStyleSheet("background-color: #e0e0e0;")
        layout_extra_largo.addWidget(separador3)

        btn_extra_largo = QPushButton("Seleccionada" if plantilla_actual == "extra_larga" else "Seleccionar Diseño Extra Largo")
        estilo_extra_largo = """
            QPushButton {
                background: #757575;
                color: white;
                border: none;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #616161;
            }
        """ if plantilla_actual == 'extra_larga' else """
            QPushButton {
                background: #2196F3;
                color: white;
                border: none;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #1976D2;
            }
        """
        btn_extra_largo.setStyleSheet(estilo_extra_largo)
        btn_extra_largo.setCursor(Qt.PointingHandCursor)
        btn_extra_largo.clicked.connect(lambda: self.seleccionar_plantilla('extra_larga'))
        layout_extra_largo.addWidget(btn_extra_largo)
        self.btn_extra_largo_ref = btn_extra_largo  # Guardar referencia

        cards_layout.addWidget(card_extra_largo)

        # Card 3: Diseño A4
        card_a4 = QGroupBox()
        card_a4.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 10px;
                margin-top: 0px;
            }
        """)
        layout_a4 = QVBoxLayout(card_a4)
        layout_a4.setSpacing(15)
        layout_a4.setContentsMargins(20, 20, 20, 20)

        titulo_a4 = QLabel('Diseño A4')
        titulo_a4.setStyleSheet("font-size: 16px; font-weight: bold; color: #1976D2;")
        layout_a4.addWidget(titulo_a4)

        desc_a4 = QLabel('Plantilla de boleta en formato A4 con toda la información disponible para impresoras convencionales.')
        desc_a4.setWordWrap(True)
        desc_a4.setStyleSheet("color: #666666; font-size: 13px;")
        layout_a4.addWidget(desc_a4)

        separador4 = QFrame()
        separador4.setFrameShape(QFrame.HLine)
        separador4.setStyleSheet("background-color: #e0e0e0;")
        layout_a4.addWidget(separador4)

        btn_a4 = QPushButton("Seleccionada" if plantilla_actual == "a4" else "Seleccionar Diseño A4")
        estilo_a4 = """
            QPushButton {
                background: #757575;
                color: white;
                border: none;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #616161;
            }
        """ if plantilla_actual == 'a4' else """
            QPushButton {
                background: #2196F3;
                color: white;
                border: none;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #1976D2;
            }
        """
        btn_a4.setStyleSheet(estilo_a4)
        btn_a4.setCursor(Qt.PointingHandCursor)
        btn_a4.clicked.connect(lambda: self.seleccionar_plantilla('a4'))
        layout_a4.addWidget(btn_a4)
        self.btn_a4_ref = btn_a4  # Guardar referencia

        cards_layout.addWidget(card_a4)

        main_layout.addWidget(cards_container)

        separador_reportes = QFrame()
        separador_reportes.setFrameShape(QFrame.HLine)
        separador_reportes.setStyleSheet("background-color: #e0e0e0;")
        main_layout.addWidget(separador_reportes)

        titulo_reportes = QLabel("Plantillas de Ventas")
        titulo_reportes.setStyleSheet("font-size: 14px; font-weight: bold; color: #1976D2; margin-top: 10px;")
        main_layout.addWidget(titulo_reportes)

        desc_reportes = QLabel(
            "Selecciona el diseño HTML que se usará para los reportes/exportaciones de ventas. "
            "Puedes ampliar esta galería agregando más archivos en DISEÑOSPDF."
        )
        desc_reportes.setWordWrap(True)
        desc_reportes.setStyleSheet("color: #666666; font-size: 12px; margin-bottom: 10px;")
        main_layout.addWidget(desc_reportes)

        self._ventas_template_buttons = {}
        plantilla_ventas_actual = cargar_plantilla_ventas_seleccionada(self.username)
        ventas_templates = get_plantillas_ventas_disponibles()

        ventas_templates_container = QWidget()
        ventas_templates_layout = QHBoxLayout(ventas_templates_container)
        ventas_templates_layout.setSpacing(15)
        ventas_templates_layout.setContentsMargins(0, 0, 0, 0)

        for template_key, template_info in ventas_templates.items():
            card = QGroupBox()
            card.setStyleSheet("""
                QGroupBox {
                    background-color: white;
                    border: 1px solid #e0e0e0;
                    border-radius: 10px;
                    margin-top: 0px;
                }
            """)
            card_layout = QVBoxLayout(card)
            card_layout.setSpacing(12)
            card_layout.setContentsMargins(20, 20, 20, 20)

            title = QLabel(str(template_info.get("nombre", template_key)))
            title.setStyleSheet("font-size: 16px; font-weight: bold; color: #1976D2;")
            card_layout.addWidget(title)

            description = QLabel(str(template_info.get("descripcion", "")))
            description.setWordWrap(True)
            description.setStyleSheet("color: #666666; font-size: 12px;")
            card_layout.addWidget(description)

            route_label = QLabel(str(template_info.get("ruta", "")))
            route_label.setWordWrap(True)
            route_label.setStyleSheet("color: #90A4AE; font-size: 11px;")
            card_layout.addWidget(route_label)
            card_layout.addStretch()

            button = QPushButton(
                "Seleccionada" if template_key == plantilla_ventas_actual else "Usar esta plantilla"
            )
            button.setCursor(Qt.PointingHandCursor)
            button.setStyleSheet(
                """
                QPushButton {
                    background: #757575;
                    color: white;
                    border: none;
                    padding: 12px 20px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background: #616161;
                }
                """
                if template_key == plantilla_ventas_actual
                else
                """
                QPushButton {
                    background: #2196F3;
                    color: white;
                    border: none;
                    padding: 12px 20px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background: #1976D2;
                }
                """
            )
            button.clicked.connect(lambda _, key=template_key: self.seleccionar_plantilla_ventas(key))
            card_layout.addWidget(button)
            self._ventas_template_buttons[template_key] = button
            ventas_templates_layout.addWidget(card)

        main_layout.addWidget(ventas_templates_container)

        # ===== SECCIÓN LOGO DE EMPRESA =====
        separador_logo = QFrame()
        separador_logo.setFrameShape(QFrame.HLine)
        separador_logo.setStyleSheet("background-color: #e0e0e0;")
        main_layout.addWidget(separador_logo)

        titulo_logo = QLabel("Logo de Empresa")
        titulo_logo.setStyleSheet("font-size: 14px; font-weight: bold; color: #1976D2; margin-top: 10px;")
        main_layout.addWidget(titulo_logo)

        desc_logo = QLabel("Sube el logo de tu empresa para que se muestre en las boletas grandes (Largo y Extra Largo).")
        desc_logo.setWordWrap(True)
        desc_logo.setStyleSheet("color: #666666; font-size: 12px; margin-bottom: 10px;")
        main_layout.addWidget(desc_logo)

        # Contenedor para logo
        logo_container = QWidget()
        logo_layout = QHBoxLayout(logo_container)
        logo_layout.setContentsMargins(0, 0, 0, 0)
        logo_layout.setSpacing(10)

        # Botón para subir logo
        btn_upload_logo = QPushButton("Seleccionar Logo")
        btn_upload_logo.setStyleSheet("""
            QPushButton {
                background: #4CAF50;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background: #45a049;
            }
        """)
        btn_upload_logo.setCursor(Qt.PointingHandCursor)
        btn_upload_logo.clicked.connect(self.cargar_logo_empresa)
        logo_layout.addWidget(btn_upload_logo)

        # Label para mostrar logo actual
        self.lbl_logo_actual = QLabel("Sin logo cargado")
        self.lbl_logo_actual.setStyleSheet("color: #666666; font-size: 12px; margin-left: 10px;")
        logo_layout.addWidget(self.lbl_logo_actual)

        # Botón para eliminar logo
        btn_delete_logo = QPushButton("Eliminar Logo")
        btn_delete_logo.setStyleSheet("""
            QPushButton {
                background: #f44336;
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background: #da190b;
            }
        """)
        btn_delete_logo.setCursor(Qt.PointingHandCursor)
        btn_delete_logo.clicked.connect(self.eliminar_logo_empresa)
        logo_layout.addWidget(btn_delete_logo)

        logo_layout.addStretch()
        main_layout.addWidget(logo_container)

        # ===== SECCIÓN TAMAÑO DEL LOGO =====
        tamano_logo_container = QWidget()
        tamano_logo_vertical = QVBoxLayout(tamano_logo_container)
        tamano_logo_vertical.setContentsMargins(15, 15, 15, 15)
        tamano_logo_vertical.setSpacing(12)
        
        # Encabezado
        lbl_tamano = QLabel('Ajustar Tamaño del Logo')
        lbl_tamano.setStyleSheet("color: #333333; font-weight: bold; font-size: 13px;")
        tamano_logo_vertical.addWidget(lbl_tamano)
        
        # Fila con Slider (input range)
        slider_row = QHBoxLayout()
        slider_row.setSpacing(15)
        
        lbl_pequeno = QLabel("Pequeño")
        lbl_pequeno.setStyleSheet("color: #666; font-size: 11px;")
        slider_row.addWidget(lbl_pequeno)
        
        # Slider para ajustar tamaño de forma visual
        slider_tamano = QtWidgets.QSlider(Qt.Horizontal)
        slider_tamano.setMinimum(50)
        slider_tamano.setMaximum(400)
        slider_tamano.setValue(100)  # Valor por defecto para mejor rendimiento en setup_ui
        slider_tamano.setTickPosition(QtWidgets.QSlider.TicksBelow)
        slider_tamano.setTickInterval(50)
        slider_tamano.setStyleSheet("""
            QSlider::groove:horizontal {
                border: 1px solid #ddd;
                background: linear-gradient(to right, #e0e0e0, #f5f5f5);
                height: 6px;
                border-radius: 3px;
            }
            QSlider::handle:horizontal {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #2196F3, stop:1 #1976D2);
                border: 2px solid #1565C0;
                width: 20px;
                margin: -7px 0;
                border-radius: 10px;
            }
            QSlider::handle:horizontal:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #1976D2, stop:1 #1565C0);
                border: 2px solid #0d47a1;
            }
            QSlider::sub-page:horizontal {
                background: qlineargradient(to right, #42a5f5, #1976D2);
                border-radius: 3px;
            }
        """)
        slider_row.addWidget(slider_tamano, 1)
        
        lbl_grande = QLabel("Grande")
        lbl_grande.setStyleSheet("color: #666; font-size: 11px;")
        slider_row.addWidget(lbl_grande)
        
        tamano_logo_vertical.addLayout(slider_row)
        
        # Fila con Spinbox y preview
        control_row = QHBoxLayout()
        control_row.setSpacing(15)
        
        # Spinbox para entrada numérica exacta
        self.spinbox_logo_size = QtWidgets.QSpinBox()
        self.spinbox_logo_size.setMinimum(50)
        self.spinbox_logo_size.setMaximum(400)
        self.spinbox_logo_size.setValue(slider_tamano.value())
        self.spinbox_logo_size.setSuffix(" px")
        self.spinbox_logo_size.setStyleSheet("""
            QSpinBox {
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 6px;
                font-size: 12px;
                min-width: 90px;
                background: white;
            }
            QSpinBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        self.spinbox_logo_size.setMaximumWidth(100)
        self.spinbox_logo_size.valueChanged.connect(self.guardar_tamano_logo)
        control_row.addWidget(self.spinbox_logo_size)
        
        # Label para preview del tamaño en tiempo real
        self.lbl_preview_tamano = QLabel(f"Tamaño: {slider_tamano.value()}px")
        self.lbl_preview_tamano.setStyleSheet("""
            color: #1976D2;
            font-weight: bold;
            font-size: 12px;
            padding: 6px 12px;
            background: #E3F2FD;
            border-radius: 4px;
            border: 1px solid #BBDEFB;
        """)
        control_row.addWidget(self.lbl_preview_tamano)
        control_row.addStretch()
        
        tamano_logo_vertical.addLayout(control_row)
        
        # Conectar controles entre s?
        slider_tamano.valueChanged.connect(self.spinbox_logo_size.setValue)
        self.spinbox_logo_size.valueChanged.connect(slider_tamano.setValue)
        self.spinbox_logo_size.valueChanged.connect(
            lambda v: self.lbl_preview_tamano.setText(f"Tamaño: {v}px")
        )
        
        tamano_logo_vertical.addStretch()
        main_layout.addWidget(tamano_logo_container)

        # Actualizar label de logo
        self.actualizar_estado_logo()

        main_layout.addStretch()
        self.tab_widget.addTab(plantilla_tab, "Plantilla")
        self._yield_ui_for_loader()
        self.tab_widget.setVisible(True)
        self._hide_config_page_loader()
        
    def create_horizontal_widget(self, widgets):
        container = QWidget()
        h_layout = QHBoxLayout(container)
        h_layout.setContentsMargins(0, 0, 0, 0)
        h_layout.setSpacing(10)
        
        for widget in widgets:
            h_layout.addWidget(widget)
        
        return container

    def _normalize_ui_texts_local(self, root_widget=None):
        """Normaliza textos mojibake dentro de Configuracion."""
        root = root_widget if isinstance(root_widget, QtWidgets.QWidget) else self
        if root is None:
            return

        def _normalize_get_set(getter, setter):
            try:
                original = getter()
            except Exception:
                return
            if not isinstance(original, str) or not original:
                return
            fixed = maybe_normalize_ui_text(original)
            if fixed != original:
                try:
                    setter(fixed)
                except Exception:
                    pass

        widgets = [root]
        try:
            widgets.extend(root.findChildren(QtWidgets.QWidget))
        except Exception:
            pass

        for widget in widgets:
            if isinstance(widget, (QtWidgets.QLabel, QtWidgets.QPushButton, QtWidgets.QCheckBox, QtWidgets.QRadioButton)):
                _normalize_get_set(widget.text, widget.setText)

            if isinstance(widget, QtWidgets.QGroupBox):
                _normalize_get_set(widget.title, widget.setTitle)

            if isinstance(widget, QtWidgets.QLineEdit):
                _normalize_get_set(widget.placeholderText, widget.setPlaceholderText)

            if isinstance(widget, QtWidgets.QComboBox):
                try:
                    for i in range(widget.count()):
                        txt = widget.itemText(i)
                        fixed = maybe_normalize_ui_text(txt)
                        if fixed != txt:
                            widget.setItemText(i, fixed)
                except Exception:
                    pass

            if isinstance(widget, QtWidgets.QTabWidget):
                try:
                    for i in range(widget.count()):
                        txt = widget.tabText(i)
                        fixed = maybe_normalize_ui_text(txt)
                        if fixed != txt:
                            widget.setTabText(i, fixed)
                except Exception:
                    pass

    def _schedule_ui_texts_local(self, root_widget=None, delay_ms=0):
        QtCore.QTimer.singleShot(int(delay_ms), lambda: self._normalize_ui_texts_local(root_widget))
    
    def _on_tab_changed(self, index):
        """Se ejecuta cuando cambia el tab activo."""
        tab_name_raw = self.tab_widget.tabText(index)
        tab_name = maybe_normalize_ui_text(tab_name_raw)
        if tab_name != tab_name_raw:
            try:
                self.tab_widget.setTabText(index, tab_name)
            except Exception:
                pass

        current_widget = self.tab_widget.widget(index)
        self._schedule_ui_texts_local(current_widget, delay_ms=0)
        self._schedule_ui_texts_local(self, delay_ms=90)

        if hasattr(self, "optometras_tab_index") and index == self.optometras_tab_index:
            self.update_optometras_list()

        if hasattr(self, "child_devices_tab_index") and index == self.child_devices_tab_index:
            self._load_dispositivos_hijos_with_loader(force=False)

        if "SUNAT" in tab_name.upper():
            self._load_sunat_data()

    def _load_sunat_data(self):
        """Carga los datos SUNAT de forma lazy"""
        try:
            from utils.file_handler import cargar_datos_generales
            
            # Cargar datos del archivo
            datos_sunat = cargar_datos_generales(self.username)
            
            # Rellenar los campos
            if hasattr(self, 'entry_ruc'):
                self.entry_ruc.setText(datos_sunat.get("ruc", ""))
            if hasattr(self, 'entry_razon_social'):
                self.entry_razon_social.setText(datos_sunat.get("razon_social", ""))
            if hasattr(self, 'entry_direccion'):
                self.entry_direccion.setText(datos_sunat.get("direccion", ""))
            if hasattr(self, 'entry_departamento'):
                self.entry_departamento.setText(datos_sunat.get("departamento", ""))
            if hasattr(self, 'entry_provincia'):
                self.entry_provincia.setText(datos_sunat.get("provincia", ""))
            if hasattr(self, 'entry_distrito'):
                self.entry_distrito.setText(datos_sunat.get("distrito", ""))
                
        except Exception as e:
            # No mostrar error, solo fallar silenciosamente
            print(f"[ConfigPage] Error cargando datos SUNAT: {e}")
    
    def update_logo_preview(self):
        logo_path = cargar_logo_optica(str(self.parent_app.user_id))
        if not logo_path:
            self.logo_preview_label.clear()
            self.logo_preview_label.setText("Sin logo cargado")
            self.logo_preview_label.setStyleSheet("color: #666666; font-size: 12px;")
            self.logo_preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            return

        logo_pixmap = QtGui.QPixmap(logo_path)
        if logo_pixmap.isNull():
            self.logo_preview_label.clear()
            self.logo_preview_label.setText("Sin logo cargado")
            self.logo_preview_label.setStyleSheet("color: #666666; font-size: 12px;")
            self.logo_preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            return

        logo_escalado = logo_pixmap.scaledToWidth(150, Qt.TransformationMode.SmoothTransformation)
        self.logo_preview_label.setText("")
        self.logo_preview_label.setPixmap(logo_escalado)
        self.logo_preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

    def subir_logo_ui(self):
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("Imágenes (*.png *.jpg *.jpeg)")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        if file_dialog.exec_():
            source_path = file_dialog.selectedFiles()[0]
            try:
                guardar_logo_optica(str(self.parent_app.user_id), source_path)
                self.update_logo_preview()
                self.parent_app.update_logo()
                QMessageBox.information(self, 'Éxito', "Logo actualizado correctamente.")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al subir el logo: {e}")

    def check_password_setup_state(self):
        password_existente = cargar_password_setup(str(self.parent_app.user_id))
        if password_existente:
            self.entry_password_setup.setText("******")
            self.btn_guardar_password_setup.setEnabled(False)
            self.btn_guardar_password_setup.setText('Contraseña Establecida')

    def guardar_password_setup_ui(self):
        password = self.entry_password_setup.text().strip()
        if not password or len(password) != 6 or not password.isdigit():
            QMessageBox.critical(self, "Error", 'La contrasena debe ser un número de 6 dígitos.')
            return
        guardar_password_setup(str(self.parent_app.user_id), password)
        self.check_password_setup_state()

    def agregar_optometra(self):
        nombre = self.entry_optometra.text().strip()
        if not nombre:
            QMessageBox.critical(self, "Error", 'El nombre del optómetra no puede estar vacío.')
            return

        if not agregar_optometra_json(self.username, nombre):
            QMessageBox.critical(self, "Error", "Este optometra ya existe o no se pudo guardar.")
            return
        self.update_optometras_list()
        self._refresh_optometras_dependents()
        self.entry_optometra.clear()
        QMessageBox.information(self, 'Éxito', 'Optómetra agregado correctamente.')

    def eliminar_optometra(self):
        selected_items = self.list_optometras.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Advertencia", 'Seleccione un optómetra para eliminar.')
            return

        optometra_a_eliminar = selected_items[0].text()

        if not eliminar_optometra_json(self.username, optometra_a_eliminar):
            QMessageBox.warning(self, "Advertencia", 'No se pudo eliminar el optómetra seleccionado.')
            return
        self.update_optometras_list()
        self._refresh_optometras_dependents()
        QMessageBox.information(self, "Exito", "Optometra eliminado correctamente.")

    def update_optometras_list(self):
        self.list_optometras.clear()
        optometras = cargar_optometras(self.username)
        self.list_optometras.addItems(optometras)

    def _refresh_optometras_dependents(self):
        """Refresca combos y vistas que dependen de optometras."""
        try:
            if self.parent_app and hasattr(self.parent_app, 'create_patient_page'):
                self.parent_app.create_patient_page.cargar_optometras_en_combo()
        except Exception:
            pass

    # =========================================================================
    # MÉTODOS PARA SUNAT / RUC (NUEVOS)
    # =========================================================================

    def consultar_ruc_sunat(self):
        """Consulta el RUC en SUNAT y obtiene datos de la empresa."""
        from utils.sunat_api import consultar_ruc, DEFAULT_TOKEN
        from utils.file_handler import actualizar_datos_sunat, cargar_token_sunat
        
        ruc = self.entry_ruc.text().strip()
        
        if not ruc:
            QMessageBox.warning(self, "Error", "Ingrese un RUC para consultar.")
            return
        
        # Mostrar diálogo de carga
        progress = QMessageBox()
        progress.setText("Consultando SUNAT...\nEsta operación puede tardar unos segundos.")
        progress.setWindowTitle("Consultando RUC")
        progress.setStandardButtons(QMessageBox.NoButton)
        progress.setIcon(QMessageBox.Information)
        progress.show()
        QApplication.processEvents()
        
        try:
            # Obtener token guardado o usar default
            token = cargar_token_sunat(self.username)
            if not token:
                token = DEFAULT_TOKEN
            
            # Consultar SUNAT
            success, data = consultar_ruc(ruc, token)
            
            progress.close()
            
            if success:
                # Actualizar campos con datos de SUNAT
                self.entry_razon_social.setText(data.get("razonSocial", ""))
                self.entry_direccion.setText(data.get("direccion", ""))
                self.entry_departamento.setText(data.get("departamento", ""))
                self.entry_provincia.setText(data.get("provincia", ""))
                self.entry_distrito.setText(data.get("distrito", ""))
                self.entry_estado.setText(data.get("estado", ""))
                self.entry_condicion.setText(data.get("condicion", ""))
                
                # Guardar en base de datos
                actualizar_datos_sunat(self.username, data)
                
                QMessageBox.information(
                    self, 
                    'Éxito', 
                    f"RUC consultado correctamente.\n\n"
                    f"Razón Social: {data.get('razonSocial', 'N/A')}\n"
                    f"Estado: {data.get('estado', 'N/A')}\n"
                    f"Condición: {data.get('condicion', 'N/A')}"
                )
            else:
                error_msg = data.get("error", "Error desconocido")
                QMessageBox.critical(self, "Error", f"No se pudo consultar el RUC:\n{error_msg}")
        
        except Exception as e:
            progress.close()
            QMessageBox.critical(self, "Error", f"Error al consultar SUNAT:\n{str(e)}")

    def guardar_datos_sunat(self):
        """Guarda los datos SUNAT ingresados manualmente."""
        from utils.file_handler import guardar_datos_generales
        
        try:
            datos = {
                "ruc": self.entry_ruc.text().strip(),
                "razon_social": self.entry_razon_social.text().strip(),
                "direccion": self.entry_direccion.text().strip(),
                "departamento": self.entry_departamento.text().strip(),
                "provincia": self.entry_provincia.text().strip(),
                "distrito": self.entry_distrito.text().strip(),
                "estado": self.entry_estado.text().strip(),
                "condicion": self.entry_condicion.text().strip(),
            }
            
            guardar_datos_generales(self.username, datos)
            
            QMessageBox.information(
                self,
                'Éxito',
                "Datos SUNAT guardados correctamente."
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al guardar datos:\n{str(e)}")
        if self.parent_app and hasattr(self.parent_app, 'create_patient_page'):
            self.parent_app.create_patient_page.cargar_optometras_en_combo()

    def guardar_nombre_optica(self):
        nombre_optica = self.entry_nombre_optica.text().strip()
        slogan = self.entry_slogan_optica.text().strip() if hasattr(self, "entry_slogan_optica") else ""
        direccion = self.entry_direccion_optica.text().strip() if hasattr(self, "entry_direccion_optica") else ""
        correo_electronico = self.entry_correo_optica.text().strip() if hasattr(self, "entry_correo_optica") else ""
        whatsapp = self.entry_whatsapp.text().strip() if hasattr(self, "entry_whatsapp") else ""
        if not nombre_optica:
            QMessageBox.critical(self, "Error", 'El nombre de la óptica no puede estar vacío.')
            return

        try:
            guardar_datos_optica(
                self.username,
                {
                    "nombre_optica": nombre_optica,
                    "slogan": slogan,
                    "direccion": direccion,
                    "correo_electronico": correo_electronico,
                    "whatsapp": whatsapp,
                },
                sync_remote=False,
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo guardar el archivo local: {e}")
            return

        remote_ok, remote_msg = self._guardar_datos_optica_en_bd(
            {
                "nombre_optica": nombre_optica,
                "slogan": slogan,
                "direccion": direccion,
                "correo_electronico": correo_electronico,
                "whatsapp": whatsapp,
            }
        )

        # 1. Sincronización en la Nube
        try:
            from utils.sync_manager import get_sync_manager
            sync_mgr = get_sync_manager()
            
            # Guardar en Datos Generales (JSON)
            datos_gen = cargar_datos_generales(self.username)
            datos_gen["nombre_optica"] = nombre_optica
            datos_gen["slogan"] = slogan
            datos_gen["direccion"] = direccion
            datos_gen["correo_electronico"] = correo_electronico
            guardar_datos_generales(self.username, datos_gen)

            optica_cfg = cargar_configuracion_optica(self.username)
            plain_lines = [
                f"nombre_optica={optica_cfg.get('nombre_optica', nombre_optica)}",
                f"slogan={optica_cfg.get('slogan', slogan)}",
                f"direccion={optica_cfg.get('direccion', direccion)}",
                f"correo_electronico={optica_cfg.get('correo_electronico', correo_electronico)}",
            ]
            plain_content = "\n".join(plain_lines) + "\n"
            
            # Obtener ID de usuario
            usuarios = cargar_usuarios() or {}
            usuario_id = self.username
            for uid, info in usuarios.items():
                if isinstance(info, dict) and info.get('username') == self.username:
                    usuario_id = uid
                    break
            
            # Encolar cambio para el archivo de texto en nube
            sync_mgr.queue_change(
                usuario_id=str(usuario_id),
                tipo_dato='config_optica',
                operacion='SYNC_ALL',
                registro_id='config',
                contenido={'config_optica': plain_content}
            )
            sync_mgr.sync_now(str(usuario_id))
            # Enviar TXT directamente al servidor (Point de Yhana Cloud)
            try:
                import requests
                txt_url = "https://api.yhana.cloud/win/new/upload_txt_file.php"
                payload = {
                    "usuario": self.username,
                    "filename": "configuracion_optica.txt",
                    "content": plain_content
                }
                requests.post(txt_url, json=payload, timeout=5)
            except Exception as e:
                print(f"[DIRECT_UPLOAD] Error subiendo TXT: {e}")
        except Exception as sync_e:
            print(f"[SYNC] Error en sincronización: {sync_e}")

        # 3. Actualizar UI
        try:
            if self.parent_app and hasattr(self.parent_app, 'home_page'):
                if hasattr(self.parent_app.home_page, 'home_widget') and hasattr(self.parent_app.home_page.home_widget, 'updateOpticalName'):
                    self.parent_app.home_page.home_widget.updateOpticalName(nombre_optica)
                elif hasattr(self.parent_app.home_page, 'nombre_optica_label'):
                    self.parent_app.home_page.nombre_optica_label.setText(f"Bienvenido al Sistema de Gestión de {nombre_optica}")
        except Exception:
            pass

        if remote_ok:
            QMessageBox.information(self, 'Éxito', 'Información de la óptica guardada y sincronizada correctamente.')
        else:
            QMessageBox.warning(
                self,
                'Guardado parcial',
                f'La información se guardó localmente, pero falló el guardado en la BD remota.\n\nDetalle: {remote_msg}'
            )
    
    def agregar_metodo_pago(self):
        nombre = self.entry_pago.text().strip()
        if not nombre:
            QMessageBox.critical(self, "Error", 'El nombre del método de pago no puede estar vacío.')
            return

        metodos = cargar_metodos_pago(self.username)
        if nombre in metodos:
            QMessageBox.critical(self, "Error", 'Este método de pago ya existe.')
            return
        
        metodos.append(nombre)
        guardar_metodos_pago(self.username, metodos)
        self.update_pagos_list()
        self.entry_pago.clear()
        QMessageBox.information(self, 'Éxito', 'Método de pago agregado correctamente.')
    
    def eliminar_metodo_pago(self):
        selected_items = self.list_pagos.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Advertencia", 'Seleccione un método de pago para eliminar.')
            return

        metodo_a_eliminar = selected_items[0].text()
        
        metodos = cargar_metodos_pago(self.username)
        metodos.remove(metodo_a_eliminar)
        guardar_metodos_pago(self.username, metodos)
        self.update_pagos_list()
        QMessageBox.information(self, 'Éxito', 'Método de pago eliminado correctamente.')

    def update_pagos_list(self):
        self.list_pagos.clear()
        metodos = cargar_metodos_pago(self.username)
        self.list_pagos.addItems(metodos)
        if self.parent_app and hasattr(self.parent_app, 'sales_page'):
            self.parent_app.sales_page.update_metodo_pago_combo()

    def setup_db_and_backup_tabs(self):
        'Crea la pestaña de Copia de Seguridad con diseño de tarjeta.'
        backup_tab = QWidget()
        main_layout = QVBoxLayout(backup_tab)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        main_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Título con icono
        titulo_widget = QWidget()
        titulo_layout = QHBoxLayout(titulo_widget)
        titulo_layout.setContentsMargins(0, 0, 0, 0)
        titulo_icon = QtWidgets.QLabel()
        icon = QtWidgets.QApplication.style().standardIcon(QtWidgets.QStyle.SP_DirIcon)
        titulo_icon.setPixmap(icon.pixmap(32, 32))
        titulo_layout.addWidget(titulo_icon)
        titulo = QLabel("<h2>Copia de Seguridad</h2>")
        titulo.setStyleSheet("color: #1976D2; margin-bottom: 10px;")
        titulo_layout.addWidget(titulo)
        titulo_layout.addStretch()
        main_layout.addWidget(titulo_widget)

        # Tarjeta de copia de seguridad
        backup_card = QGroupBox()
        backup_card.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 10px;
                margin-top: 0px;
            }
        """)
        card_layout = QVBoxLayout(backup_card)
        card_layout.setSpacing(15)
        card_layout.setContentsMargins(20, 20, 20, 20)

        # Mensaje principal
        mensaje_principal = QLabel()
        mensaje_principal.setWordWrap(True)
        mensaje_principal.setStyleSheet("font-size: 14px; color: #424242;")
        mensaje_principal.setText(
            "Protege tus datos creando una copia de seguridad local. "
            "Guarda tus archivos en un lugar seguro para recuperarlos cuando los necesites."
        )
        card_layout.addWidget(mensaje_principal)

        # Separador
        separador = QFrame()
        separador.setFrameShape(QFrame.HLine)
        separador.setFrameShadow(QFrame.Sunken)
        separador.setStyleSheet("background-color: #e0e0e0;")
        card_layout.addWidget(separador)

        # Botón de acción
        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)
        buttons_layout.setContentsMargins(0, 0, 0, 0)
        buttons_layout.setSpacing(10)

        btn_export = QPushButton("Crear Copia de Seguridad")
        btn_export.setObjectName("primaryButton")
        btn_export.setStyleSheet("""
            QPushButton {
                background: #2196F3;
                color: white;
                border: none;
                padding: 12px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #1976D2;
            }
        """)
        btn_export.setCursor(Qt.PointingHandCursor)
        btn_export.clicked.connect(self.exportar_datos_generales_ui)
        buttons_layout.addWidget(btn_export)
        buttons_layout.addStretch()

        card_layout.addWidget(buttons_widget)

        # Nota final
        nota = QLabel()
        nota.setWordWrap(True)
        nota.setStyleSheet("color: #757575; font-size: 12px;")
        nota.setText(
            "Se recomienda descargar y guardar las copias de seguridad en un lugar seguro. "
            'Las copias de seguridad contienen toda la información de tu negocio.'
        )
        card_layout.addWidget(nota)

        main_layout.addWidget(backup_card)
        main_layout.addStretch()
        self.tab_widget.addTab(backup_tab, "Copia de Seguridad")
        self._yield_ui_for_loader()

    def subir_datos_a_la_nube(self):
        self.sender().setDisabled(True)
        QtWidgets.QApplication.setOverrideCursor(Qt.WaitCursor)
        
        success, message = comprimir_y_subir_datos(self.username)
        
        QtWidgets.QApplication.restoreOverrideCursor()
        self.sender().setDisabled(False)

        if success:
            QMessageBox.information(self, 'Éxito', message)
        else:
            QMessageBox.critical(self, "Error", message)

    def recibir_datos_de_la_nube(self):
        self.sender().setDisabled(True)
        QtWidgets.QApplication.setOverrideCursor(Qt.WaitCursor)
        
        success, message, zip_path = descargar_y_descomprimir_datos(self.username)

        QtWidgets.QApplication.restoreOverrideCursor()
        self.sender().setDisabled(False)
        
        if success:
            msg_box = QMessageBox()
            msg_box.setWindowTitle("Recibir Datos")
            msg_box.setText('¿Qué deseas hacer con los datos recibidos de la nube?')
            btn_replace = msg_box.addButton("Reemplazar todo", QMessageBox.ButtonRole.YesRole)
            btn_add = msg_box.addButton("Agregar datos", QMessageBox.ButtonRole.NoRole)
            msg_box.exec_()
            
            if msg_box.clickedButton() == btn_replace:
                self.reemplazar_datos_con_backup(zip_path)
            elif msg_box.clickedButton() == btn_add:
                self.agregar_datos_con_backup(zip_path)
            
            if os.path.exists(zip_path):
                os.remove(zip_path)
        else:
            QMessageBox.critical(self, "Error", message)

    def reemplazar_datos_con_backup(self, zip_path):
        try:
            user_data_path = VISO_DIR / self.username / "data"
            if user_data_path.exists():
                shutil.rmtree(user_data_path)
            
            os.makedirs(user_data_path, exist_ok=True)
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(VISO_DIR / self.username)
            
            QMessageBox.information(self, 'Éxito', 'Datos reemplazados correctamente. La aplicación se reiniciará para aplicar los cambios.')
            self.parent_app.restart_app()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudieron reemplazar los datos: {e}")

    def agregar_datos_con_backup(self, zip_path):
        try:
            temp_dir = VISO_DIR / "temp_backup_extract"
            os.makedirs(temp_dir, exist_ok=True)
            
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            data_types = ['pacientes', 'productos', 'ventas', 'optometras', 'metodos_pago', 'clientes']
            for data_type in data_types:
                local_file_path = get_user_file_path(self.username, f"{data_type}.json")
                backup_file_path = temp_dir / self.username / "data" / f"{data_type}.json"
                
                if backup_file_path.exists() and local_file_path.exists():
                    with open(local_file_path, 'r') as local_file:
                        try:
                            local_data = json.load(local_file)
                        except json.JSONDecodeError:
                            local_data = []

                    with open(backup_file_path, 'r') as backup_file:
                        try:
                            backup_data = json.load(backup_file)
                        except json.JSONDecodeError:
                            backup_data = []

                    if isinstance(local_data, list) and isinstance(backup_data, list):
                        local_data.extend(backup_data)
                    
                    with open(local_file_path, 'w') as local_file:
                        json.dump(local_data, local_file, indent=4)
            
            shutil.rmtree(temp_dir)
            QMessageBox.information(self, 'Éxito', 'Datos agregados correctamente. La aplicación se reiniciará para aplicar los cambios.')
            self.parent_app.restart_app()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudieron agregar los datos: {e}")

    def seleccionar_plantilla(self, tipo_plantilla):
        """Guarda la plantilla seleccionada y actualiza los estilos de los botones."""
        from utils.generador_boletas_plantilla import GeneradorBoletasPlantilla
        
        try:
            generador = GeneradorBoletasPlantilla(self.username)
            generador.guardar_plantilla_seleccionada(tipo_plantilla)
            
            # Mapear nombres para mostrar al usuario
            nombres_plantillas = {
                'pequena': 'Diseño Pequeño',
                'extra_larga': 'Diseño Extra Largo',
                'a4': 'Diseño A4'
            }
            
            nombre_plantilla = nombres_plantillas.get(tipo_plantilla, tipo_plantilla)
            
            # Estilos
            estilo_seleccionado = """
                QPushButton {
                    background: #757575;
                    color: white;
                    border: none;
                    padding: 12px 20px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background: #616161;
                }
            """
            
            estilo_no_seleccionado = """
                QPushButton {
                    background: #2196F3;
                    color: white;
                    border: none;
                    padding: 12px 20px;
                    border-radius: 5px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background: #1976D2;
                }
            """
            
            # Actualizar estilos de botones
            if hasattr(self, 'btn_pequeno_ref'):
                if tipo_plantilla == 'pequena':
                    self.btn_pequeno_ref.setText('Seleccionada')
                    self.btn_pequeno_ref.setStyleSheet(estilo_seleccionado)
                else:
                    self.btn_pequeno_ref.setText('Seleccionar Diseño Pequeño')
                    self.btn_pequeno_ref.setStyleSheet(estilo_no_seleccionado)
            
            if hasattr(self, 'btn_extra_largo_ref'):
                if tipo_plantilla == 'extra_larga':
                    self.btn_extra_largo_ref.setText('Seleccionada')
                    self.btn_extra_largo_ref.setStyleSheet(estilo_seleccionado)
                else:
                    self.btn_extra_largo_ref.setText('Seleccionar Diseño Extra Largo')
                    self.btn_extra_largo_ref.setStyleSheet(estilo_no_seleccionado)
            
            if hasattr(self, 'btn_a4_ref'):
                if tipo_plantilla == 'a4':
                    self.btn_a4_ref.setText('Seleccionada')
                    self.btn_a4_ref.setStyleSheet(estilo_seleccionado)
                else:
                    self.btn_a4_ref.setText('Seleccionar Diseño A4')
                    self.btn_a4_ref.setStyleSheet(estilo_no_seleccionado)
            
            QMessageBox.information(
                self,
                "Plantilla Seleccionada",
                f"{nombre_plantilla} ha sido seleccionada correctamente.\n\n"
                f"Las nuevas boletas se generarán con este formato."
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al seleccionar plantilla:\n{str(e)}")

    def seleccionar_plantilla_ventas(self, template_key):
        """Guarda la plantilla HTML seleccionada para reportes de ventas."""
        try:
            guardar_plantilla_ventas_seleccionada(self.username, template_key)
            template_map = get_plantillas_ventas_disponibles()
            for key, button in getattr(self, "_ventas_template_buttons", {}).items():
                selected = key == template_key
                button.setText("Seleccionada" if selected else "Usar esta plantilla")
                button.setStyleSheet(
                    """
                    QPushButton {
                        background: #757575;
                        color: white;
                        border: none;
                        padding: 12px 20px;
                        border-radius: 5px;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background: #616161;
                    }
                    """
                    if selected
                    else
                    """
                    QPushButton {
                        background: #2196F3;
                        color: white;
                        border: none;
                        padding: 12px 20px;
                        border-radius: 5px;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background: #1976D2;
                    }
                    """
                )

            template_name = str((template_map.get(template_key) or {}).get("nombre", template_key))
            QMessageBox.information(
                self,
                "Plantilla de Ventas",
                f"{template_name} ha sido seleccionada correctamente.\n\nLos reportes de ventas usarán este diseño."
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al seleccionar plantilla de ventas:\n{str(e)}")
    
    def cargar_logo_empresa(self):
        'Abre un diálogo para seleccionar la imagen del logo.'
        from PyQt5.QtWidgets import QFileDialog
        import os
        import shutil
        
        # Obtener ruta del logo
        ruta_logos = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'logos')
        os.makedirs(ruta_logos, exist_ok=True)
        
        # Abrir diálogo de selección de archivo
        opciones = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar Logo de Empresa",
            "",
            "Imágenes (*.png *.jpg *.jpeg *.bmp);;Todos los archivos (*)",
            options=opciones
        )
        
        if file_path:
            try:
                # Copiar logo a carpeta data/logos
                nombre_logo = "logo_empresa.png"
                ruta_destino = os.path.join(ruta_logos, nombre_logo)
                shutil.copy(file_path, ruta_destino)
                
                # Guardar ruta en config del usuario
                from utils.generador_boletas_plantilla import GeneradorBoletasPlantilla
                generador = GeneradorBoletasPlantilla(self.username)
                
                # Actualizar estado
                self.actualizar_estado_logo()
                
                QMessageBox.information(
                    self,
                    "Logo Cargado",
                    "Logo cargado correctamente.\n\n"
                    "Se mostrará en las boletas grandes (Largo y Extra Largo)."
                )
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al cargar logo:\n{str(e)}")
    
    def eliminar_logo_empresa(self):
        """Elimina el logo cargado."""
        import os
        
        ruta_logos = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'logos')
        ruta_logo = os.path.join(ruta_logos, 'logo_empresa.png')
        
        if os.path.exists(ruta_logo):
            try:
                os.remove(ruta_logo)
                self.actualizar_estado_logo()
                QMessageBox.information(self, "Logo Eliminado", "Logo eliminado correctamente.")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al eliminar logo:\n{str(e)}")
        else:
            QMessageBox.information(self, 'Información', "No hay logo cargado actualmente.")
    
    def actualizar_estado_logo(self):
        """Actualiza el label que muestra el estado del logo."""
        import os
        
        ruta_logos = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'logos')
        ruta_logo = os.path.join(ruta_logos, 'logo_empresa.png')
        
        if os.path.exists(ruta_logo):
            try:
                tamano_kb = os.path.getsize(ruta_logo) / 1024
                self.lbl_logo_actual.setText(f"Logo cargado ({tamano_kb:.1f} KB)")
                self.lbl_logo_actual.setStyleSheet("color: #4CAF50; font-size: 12px; margin-left: 10px; font-weight: bold;")
            except:
                self.lbl_logo_actual.setText("Logo cargado")
                self.lbl_logo_actual.setStyleSheet("color: #4CAF50; font-size: 12px; margin-left: 10px; font-weight: bold;")
        else:
            self.lbl_logo_actual.setText("Sin logo cargado")
            self.lbl_logo_actual.setStyleSheet("color: #666666; font-size: 12px; margin-left: 10px;")

    def cargar_tamano_logo_guardado(self):
        'Carga el tamaño guardado del logo desde la configuración.'
        try:
            return cargar_tamano_logo(self.username)
        except Exception as e:
            print(f"Error al cargar tamaño del logo: {e}")
            return 150  # Default size

    def guardar_tamano_logo(self, tamano):
        'Guarda el tamaño del logo seleccionado por el usuario.'
        try:
            guardar_tamano_logo(self.username, tamano)
            # Actualizar label de preview
            if hasattr(self, 'lbl_preview_tamano'):
                self.lbl_preview_tamano.setText(f"Tamaño actual: {tamano}px")
        except Exception as e:
            print(f"Error al guardar tamaño del logo: {e}")

    def descargar_plantilla_clientes(self):
        """Genera y descarga una plantilla de ejemplo para importar clientes."""
        try:
            # Crear archivo temporal
            plantilla_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar plantilla de importación de clientes",
                os.path.expanduser("~\\Documents\\plantilla_clientes.xlsx"),
                "Archivos Excel (*.xlsx)"
            )
            
            if not plantilla_path:
                return
            
            # Crear workbook
            try:
                from openpyxl import Workbook # type: ignore
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # type: ignore
            except ImportError:
                QMessageBox.critical(self, "Error", "La librería 'openpyxl' no está instalada.")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Clientes"
            
            # Encabezados
            encabezados = ["Nombres", "Apellidos", "DNI", "Fecha de Nacimiento", "Edad", "Fecha de Registro", "Teléfono", "Correo"]
            
            # Aplicar estilos a los encabezados
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Escribir encabezados
            for col_num, encabezado in enumerate(encabezados, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = encabezado
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Datos de ejemplo
            datos_ejemplo = [
                ["Juan Alberto", 'Pérez García', "12345678", "15/03/1985", "40", "5/12/2024", "999123456", "juan.perez@email.com"],
                ['María Elena', 'Rodríguez Ruiz', "87654321", "22/07/1990", "34", "15/12/2024", "999234567", "maria.rodriguez@email.com"],
                ["Carlos Luis", 'Gómez Torres', "11223344", "10/11/1988", "36", "20/12/2024", "999345678", "carlos.gomez@email.com"],
                ['Ana Lucía', 'Fernández Lima', "44332211", "", "28", "10/01/2025", "999456789", ""],
                ["Luis Miguel", "Castro Soto", "55667788", "05/06/1992", "32", "18/01/2025", "", "luis.castro@email.com"],
            ]
            
            # Escribir datos de ejemplo
            for row_num, datos in enumerate(datos_ejemplo, 2):
                for col_num, valor in enumerate(datos, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = valor
                    cell.border = border
                    if col_num in [4, 6]:  # Columnas de fecha
                        cell.alignment = Alignment(horizontal="center")
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 25
            
            # Altura de encabezado
            ws.row_dimensions[1].height = 25
            
            # Crear hoja de instrucciones
            ws_instrucciones = wb.create_sheet("Instrucciones")
            
            instrucciones = [
                ["GUÍA DE IMPORTACIÓN DE CLIENTES", ""],
                ["", ""],
                ["COLUMNAS DISPONIBLES:", ""],
                ["", ""],
                ["Nombres *", "Obligatorio. Puede contener el nombre completo o solo el primer nombre."],
                ["Apellidos", "Opcional. Los apellidos del cliente. Se combina con Nombres si ambos existen."],
                ["DNI", "Opcional. Número de documento (DNI, RUC, Pasaporte, etc)."],
                ["Fecha de Nacimiento", "Opcional. Formato: DD/MM/YYYY (ej: 15/03/1985)"],
                ["Edad", 'Opcional. Edad en años (número entero).'],
                ["Fecha de Registro", "Opcional. Fecha en que el cliente se registró. Formato: D/MM/YYYY (ej: 5/12/2024)"],
                ["Teléfono", 'Opcional. Número telefónico o celular.'],
                ["Correo", 'Opcional. Dirección de correo electrónico.'],
                ["", ""],
                ["NOTAS IMPORTANTES:", ""],
                ["• Solo la columna 'Nombres' es obligatoria.", ""],
                ["• Puedes usar cualquier combinación de columnas.", ""],
                ["• El sistema detecta automáticamente el nombre de las columnas.", ""],
                ["• Variaciones aceptadas: 'Nombre'/'Nombres', 'Apellido'/'Apellidos', 'Documento'/'DNI', etc.", ""],
                ["• Las filas vacías o sin nombre serán ignoradas.", ""],
                ["• Los duplicados (mismo nombre) no se importarán.", ""],
            ]
            
            for row_num, datos in enumerate(instrucciones, 1):
                for col_num, valor in enumerate(datos, 1):
                    cell = ws_instrucciones.cell(row=row_num, column=col_num)
                    cell.value = valor
                    if row_num == 1:
                        cell.font = Font(bold=True, size=14, color="2196F3")
                    elif row_num in [3, 14]:
                        cell.font = Font(bold=True, size=12)
            
            ws_instrucciones.column_dimensions['A'].width = 30
            ws_instrucciones.column_dimensions['B'].width = 60
            
            # Guardar archivo
            wb.save(plantilla_path)
            
            QMessageBox.information(
                self,
                "Plantilla Descargada",
                f"Plantilla descargada correctamente en:\n{plantilla_path}\n\n"
                "Puedes usar este archivo como base para importar tus clientes.\n"
                "Completa los datos y vuelve a importar el archivo."
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar plantilla:\n{str(e)}")
    
    def importar_clientes_archivo(self):
        """Importa clientes desde un archivo Excel o CSV."""
        try:
            # Abrir diálogo de selección de archivo
            archivo_path, _ = QFileDialog.getOpenFileName(
                self,
                "Seleccionar archivo de clientes",
                os.path.expanduser("~\\Documents"),
                "Archivos Excel (*.xlsx *.xls);;Archivos CSV (*.csv);;Todos los archivos (*)"
            )
            
            if not archivo_path:
                return
            
            # Determinar el tipo de archivo
            if archivo_path.lower().endswith(('.xlsx', '.xls')):
                clientes, columnas_detectadas = self.leer_excel(archivo_path)
            elif archivo_path.lower().endswith('.csv'):
                clientes, columnas_detectadas = self.leer_csv(archivo_path)
            else:
                QMessageBox.critical(self, "Error", "Formato de archivo no soportado. Use Excel o CSV.")
                return
            
            if not clientes:
                QMessageBox.warning(self, "Advertencia", 'El archivo no contiene datos válidos o no tiene las columnas requeridas.')
                return
            
            # Mostrar columnas detectadas
            mensaje_columnas = "Columnas detectadas:\n\n"
            for campo, columna in columnas_detectadas.items():
                if columna:
                    mensaje_columnas += f"✓ {campo.replace('_', ' ').title()}: '{columna}'\n"
                else:
                    mensaje_columnas += f"✗ {campo.replace('_', ' ').title()}: No encontrada\n"
            
            mensaje_columnas += f"\n\nTotal de clientes a importar: {len(clientes)}"
            
            respuesta = QMessageBox.question(
                self,
                'Confirmación de Importación',
                mensaje_columnas + '\n¿Deseas continuar con la importación?',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if respuesta != QMessageBox.Yes:
                return
            
            # Cargar clientes existentes
            clientes_file = get_user_file_path(self.username, "clientes.json")
            clientes_existentes = []
            
            if clientes_file.exists():
                try:
                    with open(clientes_file, 'r', encoding='utf-8') as f:
                        clientes_existentes = json.load(f)
                except:
                    clientes_existentes = []
            
            # Agregar nuevos clientes
            clientes_agregados = 0
            clientes_duplicados = 0
            
            for cliente in clientes:
                # Buscar si el cliente ya existe por nombre
                existe = False
                nombre_nuevo = cliente.get('nombre', '').lower()
                
                if nombre_nuevo:
                    for cliente_existente in clientes_existentes:
                        if cliente_existente.get('nombre', '').lower() == nombre_nuevo:
                            existe = True
                            clientes_duplicados += 1
                            break
                
                if not existe and nombre_nuevo:
                    clientes_existentes.append(cliente)
                    clientes_agregados += 1
            
            # Guardar archivo actualizado
            os.makedirs(clientes_file.parent, exist_ok=True)
            with open(clientes_file, 'w', encoding='utf-8') as f:
                json.dump(clientes_existentes, f, ensure_ascii=False, indent=2)
            
            # Mostrar resultado
            mensaje = "Importación completada\n\n"
            mensaje += f"Clientes agregados: {clientes_agregados}\n"
            if clientes_duplicados > 0:
                mensaje += f"Clientes duplicados (no agregados): {clientes_duplicados}\n"
            
            QMessageBox.information(self, 'Importación exitosa', mensaje)
            
            # Actualizar la interfaz si es necesario
            if hasattr(self.parent_app, 'inventory_page') and hasattr(self.parent_app.inventory_page, 'actualizar_clientes'):
                self.parent_app.inventory_page.actualizar_clientes()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al importar clientes:\n{str(e)}")

    def normalizar_nombre_columna(self, nombre):
        'Normaliza un nombre de columna para comparación.'
        import unicodedata
        # Convertir a minúsculas y eliminar espacios
        nombre = nombre.lower().strip()
        # Remover acentos
        nombre = unicodedata.normalize('NFD', nombre)
        nombre = ''.join(c for c in nombre if unicodedata.category(c) != 'Mn')
        # Remover caracteres especiales excepto espacios
        nombre = ''.join(c if c.isalnum() or c == ' ' else '' for c in nombre)
        return nombre
    
    def obtener_mapeo_columnas(self):
        'Retorna un diccionario con los sinónimos para cada campo.'
        return {
            'nombre': [
                'nombre', 'nombres', 'name', 'first name', 'primer nombre', 
                'nombre completo', 'full name', 'cliente', 'customer', 'person',
                'nombre del cliente', 'nombre de la persona', 'apellido y nombre',
                'nombre y apellido', 'razón social', 'razon social', 'nombre cliente'
            ],
            'apellido': [
                'apellido', 'apellidos', 'last name', 'lastname', 'surname',
                'segundo nombre', 'segundo apellido', 'apellido paterno',
                'apellido materno', 'family name', 'apellido paterno materno'
            ],
            'dni': [
                'dni', 'document', 'documento', 'numero de documento', 
                'número de documento', 'numero documento', 'número documento',
                'id', 'identification', 'cedula', 'cédula', 'ruc', 'nif', 
                'pasaporte', 'carnet', 'identidad', 'doc', 'id number', 
                'documento numero', 'documento número', 'num documento'
            ],
            'edad': [
                'edad', 'eda', 'age', 'años', 'years', 'year of birth', 'año nacimiento',
                'año de nacimiento', 'años edad', 'age years', 'e', 'edades'
            ],
            'fecha_nacimiento': [
                'fecha de nacimiento', 'fecha nacimiento', 'nacimiento', 'birth',
                'birth date', 'birthdate', 'fecha de nac', 'f. nac', 'dob',
                'date of birth', 'birth day', 'cumpleaños', 'cumpleaños fecha',
                'fecha nac', 'fecha nac.', 'fec. nacimiento', 'fecha nac.', 
                'fechanacimiento', 'fecha_nacimiento'
            ],
            'fecha_registro': [
                'fecha de registro', 'fecha registro', 'fecha de alta', 'fecha alta',
                'fecha creacion', 'fecha creación', 'fecha de creacion', 'fecha de creación',
                'registration date', 'fecha ingreso', 'fecha de ingreso', 'fecha afiliacion',
                'fecha de afiliacion', 'fecha de afiliación', 'fecha afiliación',
                'fecharegistro', 'fecha_registro', 'registro fecha', 'f/registro', 'f registro',
                'f/alta', 'f alta', 'f creacion', 'f/creacion', 'fecha', 'f.registro',
                'f.alta', 'f.creacion'
            ],
            'telefono': [
                'telefono', 'teléfono', 'phone', 'celular', 'móvil', 'mobile',
                'numero telefonico', 'número telefónico', 'number', 'tel',
                'tel.', 'telefono 1', 'telefono 2', 'phone number', 'cell',
                'cell phone', 'contacto', 'contact', 'teléfono principal',
                'teléfono celular', 'numero celular', 'número celular',
                'telefónico', 'telefonica', 'telefonico', 'cell phone number',
                'telefo', 'telefo.', 'num celular', 'numero celular', 'numero',
                'numero de telefono', 'número de teléfono', 'tel', 'cel', 'celula'
            ],
            'correo': [
                'correo', 'email', 'e-mail', 'mail', 'e mail', 'correo electronico',
                'correo electrónico', 'email address', 'direccion email',
                'dirección email', 'dirección e-mail', 'correo principal', 
                'correo contacto', 'e-correo', 'electronico', 'electrónico',
                'correo_electronico', 'email_address'
            ]
        }
    
    def encontrar_columna(self, encabezados, palabras_clave):
        """Encuentra una columna por palabras clave de forma muy robusta."""
        if not encabezados or not palabras_clave:
            return None
        
        # Intentar cada encabezado contra cada palabra clave
        for encabezado in encabezados:
            encabezado_norm = self.normalizar_nombre_columna(encabezado)
            encabezado_sin_espacios = encabezado_norm.replace(" ", "")
            
            for palabra in palabras_clave:
                palabra_norm = self.normalizar_nombre_columna(palabra)
                palabra_sin_espacios = palabra_norm.replace(" ", "")
                
                # Nivel 1: Coincidencia exacta (normalizado)
                if palabra_norm == encabezado_norm:
                    return encabezado
                
                # Nivel 2: Coincidencia sin espacios
                if palabra_sin_espacios == encabezado_sin_espacios:
                    return encabezado
                
                # Nivel 3: Contención en cualquier dirección (pero solo para palabras largas)
                # Evitar falsos positivos como "apellido" coincidiendo con "dni"
                if len(palabra_norm) >= 4 and len(encabezado_norm) >= 4:
                    if (palabra_norm in encabezado_norm or encabezado_norm in palabra_norm):
                        return encabezado
        
        return None

    def leer_excel(self, archivo_path):
        'Lee clientes desde un archivo Excel usando pandas (más rápido y robusto).'
        clientes = []
        columnas_detectadas = {}
        
        try:
            # Intenta usar pandas primero (más eficiente)
            if pd is not None:
                return self._leer_excel_con_pandas(archivo_path)
            # Si no, fallback a openpyxl
            elif openpyxl is not None:
                return self._leer_excel_con_openpyxl(archivo_path)
            else:
                QMessageBox.warning(
                    self,
                    "Advertencia",
                    'Para importar archivos Excel, instala las librerías:\n'
                    "pandas (recomendado): pip install pandas\n"
                    "openpyxl: pip install openpyxl\n\n"
                    "Por ahora, convierte el archivo a CSV e intenta de nuevo."
                )
                return [], {}
            
        except Exception as e:
            print(f"Error al leer Excel: {e}")
            QMessageBox.critical(self, "Error", f"Error al leer archivo Excel:\n{str(e)}")
            return [], {}

    def _leer_excel_con_pandas(self, archivo_path):
        'Lee Excel con pandas - más rápido y robusto.'
        clientes = []
        
        try:
            # Leer Excel con pandas
            df = pd.read_excel(archivo_path, engine='openpyxl')
            
            # Normalizar nombres de columnas
            df.columns = df.columns.str.strip()
            encabezados = df.columns.tolist()
            
            if not encabezados:
                return [], {}
            
            # Obtener mapeo de columnas
            mapeo = self.obtener_mapeo_columnas()
            
            # Buscar columnas específicas
            col_nombre = self.encontrar_columna(encabezados, mapeo['nombre'])
            col_apellido = self.encontrar_columna(encabezados, mapeo['apellido'])
            col_dni = self.encontrar_columna(encabezados, mapeo['dni'])
            col_edad = self.encontrar_columna(encabezados, mapeo['edad'])
            col_fecha_nacimiento = self.encontrar_columna(encabezados, mapeo['fecha_nacimiento'])
            col_fecha_registro = self.encontrar_columna(encabezados, mapeo['fecha_registro'])
            col_telefono = self.encontrar_columna(encabezados, mapeo['telefono'])
            col_correo = self.encontrar_columna(encabezados, mapeo['correo'])
            
            # Validar que al menos tenga nombre
            if not col_nombre:
                QMessageBox.warning(self, "Error", "El archivo debe tener una columna 'Nombre'.")
                return [], {}
            
            # Procesar filas con pandas (mucho más eficiente)
            for idx, row in df.iterrows():
                cliente = {}
                nombre_parte = ""
                apellido_parte = ""
                
                # Extraer valores de columnas específicas
                if col_nombre and pd.notna(row[col_nombre]):
                    nombre_parte = str(row[col_nombre]).strip()
                
                if col_apellido and pd.notna(row[col_apellido]):
                    apellido_parte = str(row[col_apellido]).strip()
                
                # Procesar DNI - asegurarse de que sea un número válido
                if col_dni and pd.notna(row[col_dni]):
                    dni_valor = str(row[col_dni]).strip()
                    # Validar que el DNI tenga contenido y no sea solo texto
                    if dni_valor and (any(c.isdigit() for c in dni_valor) or len(dni_valor) <= 20):
                        cliente['dni'] = dni_valor
                    else:
                        # Si el DNI parece inválido, usar un valor por defecto
                        cliente['dni'] = '00000000'
                else:
                    cliente['dni'] = '00000000'
                
                if col_edad and pd.notna(row[col_edad]):
                    cliente['edad'] = str(row[col_edad]).strip()
                
                if col_fecha_nacimiento and pd.notna(row[col_fecha_nacimiento]):
                    cliente['fecha_nacimiento'] = str(row[col_fecha_nacimiento]).strip()
                
                if col_fecha_registro and pd.notna(row[col_fecha_registro]):
                    cliente['fecha_registro'] = str(row[col_fecha_registro]).strip()
                
                if col_telefono and pd.notna(row[col_telefono]):
                    cliente['telefono'] = str(row[col_telefono]).strip()
                
                if col_correo and pd.notna(row[col_correo]):
                    cliente['correo'] = str(row[col_correo]).strip()
                
                # Combinar nombre y apellido
                if nombre_parte and apellido_parte:
                    cliente['nombre'] = f"{nombre_parte} {apellido_parte}"
                elif nombre_parte:
                    cliente['nombre'] = nombre_parte
                
                # Asignar DNI por defecto si no tiene
                if not cliente.get('dni'):
                    cliente['dni'] = '00000000'
                
                # Validar que tenga al menos nombre
                if cliente.get('nombre'):
                    clientes.append(cliente)
            
            # Retornar tanto clientes como columnas detectadas
            columnas_detectadas = {
                'nombre': col_nombre,
                'apellido': col_apellido,
                'dni': col_dni,
                'edad': col_edad,
                'fecha_nacimiento': col_fecha_nacimiento,
                'fecha_registro': col_fecha_registro,
                'telefono': col_telefono,
                'correo': col_correo
            }
            return clientes, columnas_detectadas
            
        except Exception as e:
            print(f"Error al leer Excel con pandas: {e}")
            raise

    def _leer_excel_con_openpyxl(self, archivo_path):
        'Lee Excel con openpyxl - fallback si pandas no está disponible.'
        clientes = []
        
        try:
            from openpyxl import load_workbook # type: ignore
        except ImportError:
            return [], {}
            
        wb = load_workbook(archivo_path)
        ws = wb.active
        
        # Obtener encabezados
        encabezados = []
        for celda in ws[1]:
            if celda.value:
                encabezados.append(str(celda.value).strip())
        
        if not encabezados:
            return [], {}
        
        # Obtener mapeo de columnas
        mapeo = self.obtener_mapeo_columnas()
        
        # Buscar columnas específicas
        col_nombre = self.encontrar_columna(encabezados, mapeo['nombre'])
        col_apellido = self.encontrar_columna(encabezados, mapeo['apellido'])
        col_dni = self.encontrar_columna(encabezados, mapeo['dni'])
        col_edad = self.encontrar_columna(encabezados, mapeo['edad'])
        col_fecha_nacimiento = self.encontrar_columna(encabezados, mapeo['fecha_nacimiento'])
        col_fecha_registro = self.encontrar_columna(encabezados, mapeo['fecha_registro'])
        col_telefono = self.encontrar_columna(encabezados, mapeo['telefono'])
        col_correo = self.encontrar_columna(encabezados, mapeo['correo'])
        
        # Validar que al menos tenga nombre
        if not col_nombre:
            QMessageBox.warning(self, "Error", "El archivo debe tener una columna 'Nombre'.")
            return [], {}
        
        # Leer filas
        for fila in ws.iter_rows(min_row=2, values_only=False):
            cliente = {}
            nombre_parte = ""
            apellido_parte = ""
            
            # Leer solo las columnas permitidas
            for i, celda in enumerate(fila):
                if i >= len(encabezados):
                    break
                
                encabezado = encabezados[i]
                valor = celda.value
                
                if valor is None:
                    continue
                
                valor_str = str(valor).strip()
                
                if encabezado == col_nombre and valor_str:
                    nombre_parte = valor_str
                elif encabezado == col_apellido and valor_str:
                    apellido_parte = valor_str
                elif encabezado == col_dni and valor_str:
                    # Validar que el DNI sea válido (contenga dígitos)
                    if any(c.isdigit() for c in valor_str) or len(valor_str) <= 20:
                        cliente['dni'] = valor_str
                    else:
                        cliente['dni'] = '00000000'
                elif encabezado == col_edad and valor_str:
                    cliente['edad'] = valor_str
                elif encabezado == col_fecha_nacimiento and valor_str:
                    cliente['fecha_nacimiento'] = valor_str
                elif encabezado == col_fecha_registro and valor_str:
                    cliente['fecha_registro'] = valor_str
                elif encabezado == col_telefono and valor_str:
                    cliente['telefono'] = valor_str
                elif encabezado == col_correo and valor_str:
                    cliente['correo'] = valor_str
            
            # Combinar nombre y apellido si ambos existen
            if nombre_parte and apellido_parte:
                cliente['nombre'] = f"{nombre_parte} {apellido_parte}"
            elif nombre_parte:
                cliente['nombre'] = nombre_parte
            
            # Asignar DNI por defecto si no tiene
            if not cliente.get('dni'):
                cliente['dni'] = '00000000'
            
            # Validar que tenga al menos nombre
            if cliente.get('nombre'):
                clientes.append(cliente)
        
        # Retornar tanto clientes como columnas detectadas
        columnas_detectadas = {
            'nombre': col_nombre,
            'apellido': col_apellido,
            'dni': col_dni,
            'edad': col_edad,
            'fecha_nacimiento': col_fecha_nacimiento,
            'fecha_registro': col_fecha_registro,
            'telefono': col_telefono,
            'correo': col_correo
        }
        return clientes, columnas_detectadas

    def leer_csv(self, archivo_path):
        """Lee clientes desde un archivo CSV, solo aceptando columnas específicas."""
        clientes = []
        try:
            with open(archivo_path, 'r', encoding='utf-8-sig') as csvfile:
                lector = csv.DictReader(csvfile)
                
                if not lector.fieldnames:
                    return []
                
                # Buscar columnas específicas
                encabezados = [str(h).strip() for h in lector.fieldnames]
                
                # Obtener mapeo de columnas
                mapeo = self.obtener_mapeo_columnas()
                
                col_nombre = self.encontrar_columna(encabezados, mapeo['nombre'])
                col_apellido = self.encontrar_columna(encabezados, mapeo['apellido'])
                col_dni = self.encontrar_columna(encabezados, mapeo['dni'])
                col_edad = self.encontrar_columna(encabezados, mapeo['edad'])
                col_fecha_nacimiento = self.encontrar_columna(encabezados, mapeo['fecha_nacimiento'])
                col_fecha_registro = self.encontrar_columna(encabezados, mapeo['fecha_registro'])
                col_telefono = self.encontrar_columna(encabezados, mapeo['telefono'])
                col_correo = self.encontrar_columna(encabezados, mapeo['correo'])
                
                # Validar que al menos tenga nombre
                if not col_nombre:
                    QMessageBox.warning(self, "Error", "El archivo debe tener una columna 'Nombre'.")
                    return []
                
                for fila in lector:
                    cliente = {}
                    nombre_parte = ""
                    apellido_parte = ""
                    
                    # Extraer solo las columnas permitidas
                    if col_nombre and fila.get(col_nombre):
                        nombre_parte = fila.get(col_nombre).strip()
                    
                    if col_apellido and fila.get(col_apellido):
                        apellido_parte = fila.get(col_apellido).strip()
                    
                    if col_dni and fila.get(col_dni):
                        cliente['dni'] = fila.get(col_dni).strip()
                    
                    if col_edad and fila.get(col_edad):
                        cliente['edad'] = fila.get(col_edad).strip()
                    
                    if col_fecha_nacimiento and fila.get(col_fecha_nacimiento):
                        cliente['fecha_nacimiento'] = fila.get(col_fecha_nacimiento).strip()
                    
                    if col_fecha_registro and fila.get(col_fecha_registro):
                        cliente['fecha_registro'] = fila.get(col_fecha_registro).strip()
                    
                    if col_telefono and fila.get(col_telefono):
                        cliente['telefono'] = fila.get(col_telefono).strip()
                    
                    if col_correo and fila.get(col_correo):
                        cliente['correo'] = fila.get(col_correo).strip()
                    
                    # Combinar nombre y apellido si ambos existen
                    if nombre_parte and apellido_parte:
                        cliente['nombre'] = f"{nombre_parte} {apellido_parte}"
                    elif nombre_parte:
                        cliente['nombre'] = nombre_parte
                    
                    # Asignar DNI por defecto si no tiene
                    if not cliente.get('dni'):
                        cliente['dni'] = '00000000'
                    
                    # Validar que tenga al menos nombre
                    if cliente.get('nombre'):
                        clientes.append(cliente)
        
        except Exception as e:
            print(f"Error al leer CSV: {e}")
            QMessageBox.critical(self, "Error", f"Error al leer archivo CSV:\n{str(e)}")
        
        # Retornar tanto clientes como columnas detectadas
        columnas_detectadas = {
            'nombre': col_nombre,
            'apellido': col_apellido,
            'dni': col_dni,
            'edad': col_edad,
            'fecha_nacimiento': col_fecha_nacimiento,
            'fecha_registro': col_fecha_registro,
            'telefono': col_telefono,
            'correo': col_correo
        }
        return clientes, columnas_detectadas

    # ========== MÉTODOS DE EMISIÓN ELECTRÓNICA SUNAT ==========
    
    def toggle_emision_electronica(self):
        'Alterna el estado de emisión electrónica'
        try:
            from utils.configurador_sunat import ConfiguradorSUNAT
            
            configurador = ConfiguradorSUNAT(self.username, VISO_DIR)
            
            if self.btn_habilitar_emision.isChecked():
                # Validar que está todo configurado
                if not self.entry_usuario_sol.text().strip():
                    QMessageBox.warning(self, 'Validación', "Configura el Usuario SOL primero")
                    self.btn_habilitar_emision.setChecked(False)
                    return
                
                if not self.entry_password_sol.text().strip():
                    QMessageBox.warning(self, 'Validación', 'Configura la Contraseña SOL primero')
                    self.btn_habilitar_emision.setChecked(False)
                    return
                
                estado_cert = self.label_cert_estado.text().strip()
                if not (estado_cert.startswith("Válido") or estado_cert.startswith("✓")):
                    QMessageBox.warning(self, 'Validación', 'Carga un certificado válido primero')
                    self.btn_habilitar_emision.setChecked(False)
                    return
                
                # Habilitar
                success, msg = configurador.habilitar_emision_electronica(True)
                if success:
                    self.btn_habilitar_emision.setText('Habilitada')
                    QMessageBox.information(self, 'Éxito', msg)
                else:
                    QMessageBox.critical(self, "Error", msg)
                    self.btn_habilitar_emision.setChecked(False)
            else:
                # Deshabilitar
                success, msg = configurador.habilitar_emision_electronica(False)
                if success:
                    self.btn_habilitar_emision.setText('Deshabilitada')
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error: {str(e)}")

    def cargar_certificado_digital(self):
        """Carga archivo de certificado digital"""
        try:
            archivo, _ = QFileDialog.getOpenFileName(
                self,
                "Seleccionar Certificado",
                os.path.expanduser("~"),
                "Certificados (*.pem *.cer *.crt);;Todos (*)"
            )
            
            if archivo:
                from utils.sunat_digital_signer import SUNATDigitalSigner
                signer = SUNATDigitalSigner()
                is_valid, info = signer.verify_certificate(archivo)
                
                if is_valid:
                    fecha_vencimiento = info.get('not_valid_after', 'N/A')
                    self.label_cert_estado.setText(f"Válido hasta {fecha_vencimiento}")
                    self.label_cert_estado.setStyleSheet("color: #4caf50; font-weight: bold;")
                    # Guardar ruta en variable temporal
                    self._cert_path = archivo
                    QMessageBox.information(self, 'Éxito', 'Certificado válido cargado')
                else:
                    QMessageBox.critical(self, "Error", f"Certificado inválido: {info.get('error', 'Desconocido')}")
                    self.label_cert_estado.setText("Inválido")
                    self.label_cert_estado.setStyleSheet("color: #d32f2f; font-weight: bold;")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar certificado:\n{str(e)}")

    def cargar_clave_privada(self):
        """Carga archivo de clave privada"""
        try:
            archivo, _ = QFileDialog.getOpenFileName(
                self,
                "Seleccionar Clave Privada",
                os.path.expanduser("~"),
                "Claves Privadas (*.key *.pem *.pfx);;Todos (*)"
            )
            
            if archivo:
                if not os.path.exists(archivo):
                    raise FileNotFoundError("Archivo no encontrado")
                
                self.label_key_estado.setText(f"Cargada ({os.path.basename(archivo)})")
                self.label_key_estado.setStyleSheet("color: #4caf50; font-weight: bold;")
                # Guardar ruta en variable temporal
                self._key_path = archivo
                QMessageBox.information(self, 'Éxito', "Clave privada cargada")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar clave:\n{str(e)}")

    def probar_conexion_sunat(self):
        'Prueba la conexión con SUNAT'
        try:
            usuario_sol = self.entry_usuario_sol.text().strip()
            contrasena = self.entry_password_sol.text().strip()
            
            # VALIDACIÓN 1: Campos vacíos
            if not usuario_sol:
                QMessageBox.warning(
                    self, 
                    'Campo Vacío', 
                    'Usuario SOL no puede estar vacío'
                )
                return
            
            if not contrasena:
                QMessageBox.warning(
                    self, 
                    'Campo Vacío', 
                    'Contraseña SOL no puede estar vacía'
                )
                return
            
            # VALIDACIÓN 2: Longitud mínima
            if len(usuario_sol) < 5:
                QMessageBox.warning(
                    self, 
                    'Usuario SOL Inválido', 
                    f"Usuario SOL muy corto ({len(usuario_sol)} caracteres)\n"
                    f"Mínimo 5 caracteres requeridos"
                )
                return
            
            if len(contrasena) < 6:
                QMessageBox.warning(
                    self, 
                    'Contraseña SOL Inválida', 
                    f"Contraseña muy corta ({len(contrasena)} caracteres)\n"
                    f"Mínimo 6 caracteres requeridos"
                )
                return
            
            # Validaciones pasadas, mostrar conexión
            QMessageBox.information(
                self, 
                "Datos válidos",
                "Usuario y contraseña tienen formato correcto.\n"
                "Ahora intentando conectar con SUNAT..."
            )
            
            # VALIDACIÓN 3: Conectar con SUNAT
            from utils.sunat_client import SUNATClient
            
            ambiente = "testing" if self.combo_ambiente.currentIndex() == 0 else "produccion"
            client = SUNATClient(usuario_sol, contrasena, ambiente)
            
            is_valid, msg = client.validar_credenciales()
            if is_valid:
                QMessageBox.information(self, "Conexión Exitosa", msg)
            else:
                QMessageBox.warning(self, "Error de Conexión", msg)
        except Exception as e:
            error_msg = str(e)
            QMessageBox.critical(
                self, 
                "Error Crítico",
                f"Error al probar conexión:\n{error_msg}\n\n"
                f"Verifica que:\n"
                "• Tu usuario y contraseña SOL sean correctos\n"
                "• Tengas conexión a internet\n"
                "• El ambiente seleccionado sea el correcto"
            )

    def open_lan_dashboard(self):
        """Abre la dirección del servidor LAN en el navegador"""
        import webbrowser
        ip = self.get_local_ip()
        port = self.entry_port_server.text()
        url = f"http://{ip}:{port}"
        
        try:
            webbrowser.open(url)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo abrir el navegador: {e}")

    def update_lan_ui_state(self):
        """Actualiza la visibilidad de los widgets segun el modo seleccionado."""
        index = self.lan_mode_combo.currentIndex()
        if index == 0:  # Servidor
            self.server_widget.setVisible(True)
            self.client_widget.setVisible(False)
            self.btn_test_connection.setVisible(False)

            my_ip = self.get_local_ip()
            self.entry_my_ip.setText(my_ip)
            self.lbl_lan_status.setText("Modo: Servidor (Esperando configuracion)")
            self.lbl_lan_status.setStyleSheet("background-color: #E3F2FD; padding: 10px; border-radius: 5px; color: #1976D2;")
        else:  # Cliente
            self.server_widget.setVisible(False)
            self.client_widget.setVisible(True)
            self.btn_test_connection.setVisible(True)
            self.lbl_lan_status.setText("Modo: Terminal (Desconectado)")
            self.lbl_lan_status.setStyleSheet("background-color: #FFEBEE; padding: 10px; border-radius: 5px; color: #C62828;")

    def get_local_ip(self):
        """Detecta la direccion IP local de la maquina de forma rapida y sin bloqueos largos."""
        # Cachear IP para evitar re-calculo lento
        if hasattr(self, '_cached_ip') and self._cached_ip:
            return self._cached_ip
            
        import socket
        try:
            # Intento rapido con timeout minimo (0.2s)
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.settimeout(0.2)
            # No conecta realmente, solo elige la interfaz de red correcta
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            self._cached_ip = ip
            return ip
        except Exception:
            try:
                ip = socket.gethostbyname(socket.gethostname())
                self._cached_ip = ip
                return ip
            except Exception:
                return "127.0.0.1"

    def save_lan_config(self):
        """Guarda la configuracion LAN y gestiona el servidor."""
        QMessageBox.warning(self, "LAN deshabilitado", LAN_DISABLED_MESSAGE)
        self.stop_lan_service_process()
        return
        try:
            mode_idx = self.lan_mode_combo.currentIndex()
            mode = "server" if mode_idx == 0 else "client"
            port_text = self.entry_port_server.text() if mode == "server" else self.entry_port_client.text()

            unsafe_ports = [1, 7, 9, 11, 13, 15, 17, 19, 20, 21, 22, 23, 25, 37, 42, 43, 53, 69, 77, 79, 87, 95, 101, 102, 103, 104, 109, 110, 111, 113, 115, 117, 119, 123, 135, 137, 139, 143, 161, 179, 389, 427, 465, 512, 513, 514, 515, 526, 530, 531, 532, 540, 548, 554, 556, 563, 587, 601, 636, 989, 990, 993, 995, 1719, 1720, 1723, 2049, 3659, 4045, 5060, 5061, 6000, 6566, 6665, 6666, 6667, 6668, 6669, 6697, 10080]

            try:
                port_int = int(port_text)
                if port_int in unsafe_ports:
                    QMessageBox.warning(
                        self,
                        "Puerto no seguro",
                        f"El puerto {port_int} es bloqueado por navegadores (Chrome/Firefox).\n\n"
                        "Por favor usa otro, por ejemplo: 5000, 8000, 8080 o 9000."
                    )
                    return
            except ValueError:
                QMessageBox.warning(self, "Error de puerto", "Ingresa un numero de puerto valido.")
                return

            config = {
                "mode": mode,
                "port": port_text,
                "updated_at": datetime.datetime.now().isoformat(),
            }

            if mode == "client":
                config["host_ip"] = self.entry_host_ip.text()
                config["auto_sync"] = self.check_auto_sync.isChecked()
            if mode == "server":
                config["server_enabled"] = True

            lan_config_path = os.path.join(VISO_DIR, self.username, "data", "config_lan.json")
            os.makedirs(os.path.dirname(lan_config_path), exist_ok=True)
            with open(lan_config_path, "w") as f:
                json.dump(config, f, indent=2)

            if mode == "server":
                self._start_server_thread(port_text)
                self.manage_auto_sync(False)
                QMessageBox.information(
                    self,
                    "Servidor iniciado",
                    f"El servidor LAN esta activo en el puerto {port_text}.\nAhora otras PCs pueden conectarse."
                )
            else:
                if hasattr(self, "lan_server") and self.lan_server:
                    self.lan_server.stop_server()
                self.manage_auto_sync(self.check_auto_sync.isChecked())
                if self.check_auto_sync.isChecked():
                    QMessageBox.information(self, "Conectado", "Modo terminal activado con sincronizacion automatica.")
                else:
                    QMessageBox.information(
                        self,
                        "Configuracion guardada",
                        "Configurado como terminal. Usa 'Sincronizar datos ahora' manualmente."
                    )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error guardando configuracion: {e}")

    def _start_server_thread(self, port):
        """Inicia el servidor como proceso independiente en segundo plano."""
        QMessageBox.warning(self, "LAN deshabilitado", LAN_DISABLED_MESSAGE)
        return
        import subprocess
        import sys

        self.stop_lan_service_process()

        try:
            creation_flags = 0
            if os.name == "nt":
                creation_flags = 0x08000000  # CREATE_NO_WINDOW

            if getattr(sys, "frozen", False):
                cmd = [sys.executable, "server", self.username, str(port)]
            else:
                main_script = os.path.join(os.getcwd(), "main.py")
                cmd = [sys.executable, main_script, "server", self.username, str(port)]

            process = subprocess.Popen(cmd, creationflags=creation_flags, close_fds=True)

            pid_path = os.path.join(VISO_DIR, self.username, "data", "server.pid")
            os.makedirs(os.path.dirname(pid_path), exist_ok=True)
            with open(pid_path, "w") as f:
                f.write(str(process.pid))

            self.server_pid = process.pid

            ip = self.get_local_ip()
            self.lbl_lan_status.setText(f"Servidor ACTIVO (Fondo) en {ip}:{port}")
            self.lbl_lan_status.setStyleSheet("background-color: #E8F5E9; color: #2E7D32; padding: 10px; font-weight: bold;")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo iniciar el proceso de fondo: {e}")

    def stop_lan_service_process(self):
        """Mata el proceso del servidor externo si existe."""
        import subprocess
        import signal

        pid_path = os.path.join(VISO_DIR, self.username, "data", "server.pid")
        if os.path.exists(pid_path):
            try:
                with open(pid_path, "r") as f:
                    pid = int(f.read().strip())

                if os.name == "nt":
                    subprocess.call(["taskkill", "/F", "/T", "/PID", str(pid)], creationflags=0x08000000)
                else:
                    os.kill(pid, signal.SIGKILL)
            except Exception:
                pass

            try:
                os.remove(pid_path)
            except Exception:
                pass

    def stop_lan_service(self):
        """Detiene manualmente cualquier servicio LAN activo y guarda el estado."""
        self.stop_lan_service_process()
        if self.auto_sync_thread:
            try:
                self.auto_sync_thread.stop()
            except Exception:
                pass
            self.auto_sync_thread = None
        self.lbl_lan_status.setText("LAN deshabilitado.")
        self.lbl_lan_status.setStyleSheet("background-color: #EEEEEE; color: #616161; padding: 10px;")
        QMessageBox.information(self, "LAN deshabilitado", LAN_DISABLED_MESSAGE)
        return
        stopped_something = False
        updated_config = False

        try:
            lan_config_path = os.path.join(VISO_DIR, self.username, "data", "config_lan.json")
            if os.path.exists(lan_config_path):
                with open(lan_config_path, "r") as f:
                    config = json.load(f)
            else:
                config = {}
        except Exception:
            config = {}

        self.stop_lan_service_process()
        if self.lan_mode_combo.currentIndex() == 0:
            self.lbl_lan_status.setText("Servidor detenido manualmente.")
            self.lbl_lan_status.setStyleSheet("background-color: #FFEBEE; color: #C62828; padding: 10px; font-weight: bold;")
            config["server_enabled"] = False
            updated_config = True
            stopped_something = True

        if self.auto_sync_thread:
            self.manage_auto_sync(False)
            self.lbl_lan_status.setText("Sincronizacion desactivada.")
            self.lbl_lan_status.setStyleSheet("background-color: #EEEEEE; color: #616161; padding: 10px;")
            self.check_auto_sync.setChecked(False)
            stopped_something = True
            config["auto_sync"] = False
            updated_config = True

        if updated_config:
            try:
                os.makedirs(os.path.dirname(lan_config_path), exist_ok=True)
                with open(lan_config_path, "w") as f:
                    json.dump(config, f, indent=2)
            except Exception as e:
                print(f"Error actualizando config al detener: {e}")

        if not stopped_something:
            QMessageBox.information(self, "Informacion", "No hay servicios activos para detener.")
        else:
            QMessageBox.information(self, "Detenido", "Se han cortado todas las conexiones LAN y se ha guardado el estado.")

    def _start_server_thread_OLD_THREAD(self, port):
        # ... (Metodo antiguo reemplazado) ...
        pass

    def update_lan_status_label(self, msg):
        self.lbl_lan_status.setText(msg)
        if "ACTIVO" in msg:
            self.lbl_lan_status.setStyleSheet("background-color: #E8F5E9; color: #2E7D32; padding: 10px; font-weight: bold;")
        elif "Error" in msg:
            self.lbl_lan_status.setStyleSheet("background-color: #FFEBEE; color: #C62828; padding: 10px; font-weight: bold;")

    def _probe_lan_server(self, host, port, timeout=1.2):
        """Verifica si el servidor LAN responde (con intento HTTP)."""
        return False
        try:
            port = int(port)
            with socket.create_connection((host, port), timeout=timeout) as sock:
                sock.settimeout(timeout)
                req = f"GET / HTTP/1.1\r\nHost: {host}\r\nConnection: close\r\n\r\n"
                try:
                    sock.sendall(req.encode("ascii", errors="ignore"))
                    data = sock.recv(12)
                    return data.startswith(b"HTTP/")
                except socket.timeout:
                    return True
                except Exception:
                    return True
        except Exception:
            return False

    def _is_lan_server_active(self, port):
        """Confirma actividad del servidor haciendo una consulta HTTP local."""
        return False
        hosts = ["127.0.0.1"]
        try:
            local_ip = self.get_local_ip()
            if local_ip and local_ip not in hosts:
                hosts.append(local_ip)
        except Exception:
            pass
        for host in hosts:
            if self._probe_lan_server(host, port):
                return True
        return False

    def manage_auto_sync(self, enable):
        """Inicia o detiene el worker de sincronizacion automatica."""
        if self.auto_sync_thread:
            try:
                self.auto_sync_thread.stop()
            except Exception:
                pass
            self.auto_sync_thread = None
        self.lbl_lan_status.setText("LAN deshabilitado.")
        self.lbl_lan_status.setStyleSheet("background-color: #EEEEEE; color: #616161; padding: 10px;")
        return
        if self.auto_sync_thread:
            self.auto_sync_thread.stop()
            self.auto_sync_thread = None

        if enable:
            host = self.entry_host_ip.text()
            port = self.entry_port_client.text()
            if host and port:
                self.auto_sync_thread = LanAutoSyncWorker(self.username, host, port)
                self.auto_sync_thread.start()
                self.lbl_lan_status.setText("Sincronizando en tiempo real...")
                self.lbl_lan_status.setStyleSheet("background-color: #E8F5E9; color: #2E7D32; padding: 10px; font-weight: bold;")

    def test_lan_connection(self):
        """Sincroniza datos desde el servidor."""
        QMessageBox.warning(self, "LAN deshabilitado", LAN_DISABLED_MESSAGE)
        return
        host = self.entry_host_ip.text()
        port = self.entry_port_client.text()

        if not host or not port:
            QMessageBox.warning(self, "Datos incompletos", "Por favor ingresa IP y puerto del servidor.")
            return

        self.btn_test_connection.setText("Sincronizando...")
        self.btn_test_connection.setEnabled(False)
        QApplication.processEvents()

        try:
            client = LanClient(self.username)
            logs = client.sincronizar_todo(host, port)

            exitos = len([l for l in logs if "OK" in l])
            total = len(logs)
            details = "\n".join(logs)

            if exitos > 0:
                QMessageBox.information(
                    self,
                    "Sincronizacion completada",
                    f"Se han descargado {exitos}/{total} archivos correctamente.\n\n"
                    "Reinicia la aplicacion o usa 'Recargar Datos' para ver los cambios.\n\n"
                    f"Detalles:\n{details}"
                )
            else:
                QMessageBox.warning(
                    self,
                    "Fallo de conexion",
                    f"No se pudo conectar con {host}:{port}.\n"
                    "Verifica que el servidor este encendido y la IP sea correcta.\n\n"
                    f"Detalles:\n{details}"
                )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error en sincronizacion: {str(e)}")
        finally:
            self.btn_test_connection.setText("Sincronizar datos ahora")
            self.btn_test_connection.setEnabled(True)

    def load_lan_config(self):
        """Carga la configuracion LAN del archivo."""
        try:
            self.lan_mode_combo.setCurrentIndex(0)
            self.entry_port_server.setText("5000")
            self.entry_port_client.setText("5000")
            self.entry_host_ip.setText("")
            self.check_auto_sync.setChecked(False)
            self.lbl_lan_status.setText("LAN deshabilitado.")
            self.lbl_lan_status.setStyleSheet("background-color: #EEEEEE; color: #616161; padding: 10px;")
            for widget in (
                self.lan_mode_combo,
                self.entry_port_server,
                self.entry_port_client,
                self.entry_host_ip,
                self.check_auto_sync,
                self.btn_save_lan,
                self.btn_stop_lan,
                self.btn_test_connection,
                self.btn_open_web_dashboard,
            ):
                try:
                    widget.setEnabled(False)
                except Exception:
                    pass
            self.stop_lan_service_process()
        except Exception:
            pass
        return
        try:
            lan_config_path = os.path.join(VISO_DIR, self.username, "data", "config_lan.json")
            if os.path.exists(lan_config_path):
                with open(lan_config_path, "r") as f:
                    config = json.load(f)

                mode = config.get("mode", "server")
                port = config.get("port", "5000")
                host_ip = config.get("host_ip", "")
                auto_sync = config.get("auto_sync", False)
                server_enabled = config.get("server_enabled", True)

                if mode == "server":
                    self.lan_mode_combo.setCurrentIndex(0)
                    self.entry_port_server.setText(port)
                    is_running = self._is_lan_server_active(port)

                    if is_running:
                        ip = self.get_local_ip()
                        self.lbl_lan_status.setText(f"Servidor ACTIVO (Segundo plano) en {ip}:{port}")
                        self.lbl_lan_status.setStyleSheet("background-color: #E8F5E9; color: #2E7D32; padding: 10px; font-weight: bold;")
                    elif server_enabled:
                        self._start_server_thread(port)
                    else:
                        self.lbl_lan_status.setText("Modo: Servidor (Detenido manualmente)")
                        self.lbl_lan_status.setStyleSheet("background-color: #FFEBEE; color: #C62828; padding: 10px; font-weight: bold;")
                else:
                    self.lan_mode_combo.setCurrentIndex(1)
                    self.entry_port_client.setText(port)
                    self.entry_host_ip.setText(host_ip)
                    self.check_auto_sync.setChecked(auto_sync)
                    self.btn_test_connection.setText("Sincronizar datos ahora")
                    if auto_sync:
                        self.manage_auto_sync(True)
            else:
                self.lan_mode_combo.setCurrentIndex(0)

            self.update_lan_ui_state()
        except Exception as e:
            print(f"Error cargando config LAN: {e}")

    def guardar_config_emision_electronica(self):
        'Guarda la configuración de emisión electrónica'
        try:
            from utils.configurador_sunat import ConfiguradorSUNAT
            
            configurador = ConfiguradorSUNAT(self.username, VISO_DIR)
            
            # Guardar credenciales
            usuario = self.entry_usuario_sol.text().strip()
            contrasena = self.entry_password_sol.text().strip()
            
            if usuario and contrasena:
                success, msg = configurador.set_credenciales_sunat(usuario, contrasena)
                if not success:
                    QMessageBox.warning(self, "Aviso", f"Credenciales: {msg}")
            
            # Guardar datos empresa
            ruc = self.entry_ruc.text().strip()
            razon_social = self.entry_razon_social.text().strip()
            
            if ruc and razon_social:
                success, msg = configurador.set_datos_empresa(
                    ruc=ruc,
                    razon_social=razon_social,
                    direccion=self.entry_direccion.text().strip(),
                    departamento=self.entry_departamento.text().strip(),
                    provincia=self.entry_provincia.text().strip(),
                    distrito=self.entry_distrito.text().strip()
                )
                if not success:
                    QMessageBox.warning(self, "Aviso", f"Datos: {msg}")
            
            # Guardar opciones
            configurador.config['enviar_automaticamente'] = self.check_enviar_auto.isChecked()
            configurador.config['guardar_cdr'] = self.check_guardar_cdr.isChecked()
            configurador.config['ambiente'] = "testing" if self.combo_ambiente.currentIndex() == 0 else "produccion"
            
            success, msg = configurador.guardar_config()
            
            if success:
                QMessageBox.information(self, 'Éxito', 'Configuración SUNAT guardada correctamente')
            else:
                QMessageBox.critical(self, "Error", f"Error al guardar: {msg}")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al guardar configuración:\n{str(e)}")






    def _sync_modo_basico_config_ui(self):
        enabled = bool(self.checkbox_modo_basico.isChecked())
        for page_index, checkbox in getattr(self, "_modo_basico_page_checks", {}).items():
            if page_index in (0, 10):
                checkbox.setChecked(True)
                checkbox.setEnabled(False)
            else:
                checkbox.setEnabled(enabled)
        for checkbox in getattr(self, "_modo_basico_action_checks", {}).values():
            checkbox.setEnabled(enabled)

    def _save_modo_basico_preferences(self):
        page_checks = getattr(self, "_modo_basico_page_checks", {})
        action_checks = getattr(self, "_modo_basico_action_checks", {})
        visible_pages = [page for page, checkbox in page_checks.items() if checkbox.isChecked()]
        quick_actions = [page for page, checkbox in action_checks.items() if checkbox.isChecked()]

        visible_pages = list(dict.fromkeys([0] + visible_pages + [10]))
        quick_actions = [page for page in quick_actions if page in visible_pages or page in (4, 2, 1, 9, 3, 6, 10)]
        if not quick_actions:
            fallback = 4 if 4 in action_checks else next(iter(action_checks.keys()), 4)
            quick_actions = [fallback]
            if fallback in action_checks:
                action_checks[fallback].setChecked(True)

        save_modo_basico_config(
            self.username,
            {
                "visible_pages": visible_pages,
                "quick_actions": quick_actions,
            },
        )
        self._sync_modo_basico_config_ui()

    def _toggle_modo_basico(self, checked):
        set_modo_basico(self.username, checked)
        self._sync_modo_basico_config_ui()
        self._save_modo_basico_preferences()
        if checked:
            QMessageBox.information(
                self,
                'Modo Básico',
                'Modo Básico activado. Ya puedes elegir qué módulos y botones rápidos mostrar.\nReinicia la aplicación para refrescar la navegación lateral.'
            )
        else:
            QMessageBox.information(
                self,
                'Modo Normal',
                'Modo Básico desactivado. Reinicia la aplicación para volver a la navegación completa.'
            )
